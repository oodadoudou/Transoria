from __future__ import annotations

import asyncio
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

from transoria.domain import Language, TaskStatus
from transoria.llm import LlmClient, ModelConfig, ProviderFormat, ThinkingLevel
from transoria.llm.client import TransportResult
from transoria.prompts import PromptKind, default_preset
from transoria.runtime import TaskCache
from transoria.workflows.glossary import (
    GLOSSARY_FILENAME_JSON,
    GLOSSARY_FILENAME_REFERENCES,
    GLOSSARY_FILENAME_XLSX,
    GLOSSARY_STATISTICS_FILENAME_JSON,
    GlossaryConfig,
    GlossaryOrchestrator,
)


@dataclass
class CandidateEmittingTransport:
    """Each request returns a fixed list of glossary candidates as JSONL."""

    candidates: tuple[tuple[str, str, str], ...] = (
        ("신해범", "申海范", "Male Name"),
        ("공이", "孔二", "Author"),
        ("흑룡", "黑龙", "Creature"),
    )
    forced_failures: tuple[int, ...] = ()
    requests: list[dict[str, object]] = field(default_factory=list)

    async def execute(
        self,
        url: str,
        headers: Mapping[str, str],
        payload: Mapping[str, object],
        timeout: float,
    ) -> TransportResult:
        index = len(self.requests)
        self.requests.append(dict(payload))
        if index in self.forced_failures:
            return TransportResult(500, {"error": "boom"})
        lines = [
            json.dumps({"src": src, "dst": dst, "type": info}, ensure_ascii=False)
            for src, dst, info in self.candidates
        ]
        body = {
            "choices": [
                {"message": {"role": "assistant", "content": "\n".join(lines)}}
            ],
            "usage": {"prompt_tokens": 25, "completion_tokens": 55},
        }
        return TransportResult(200, body)


def _model(retry_attempts: int = 2) -> ModelConfig:
    return ModelConfig(
        id="m",
        display_name="m",
        provider_format=ProviderFormat.OPENAI,
        base_url="https://example/api/v1/",
        model_id="model-x",
        api_keys=("key",),
        thinking_level=ThinkingLevel.OFF,
        concurrency_limit=2,
        rpm_limit=0,
        retry_attempts=retry_attempts,
        retry_initial_backoff_seconds=0.0,
        retry_max_backoff_seconds=0.0,
    )


_TIMES = iter(range(120))


def _frozen_clock() -> str:
    return f"2026-04-27T00:00:{next(_TIMES, 119):02d}+00:00"


def _config(input_dir: Path, output_dir: Path) -> GlossaryConfig:
    return GlossaryConfig(
        input_dir=input_dir,
        output_dir=output_dir,
        source_language=Language.KOREAN,
        target_language=Language.CHINESE_SIMPLIFIED,
        model=_model(),
        prompt_preset=default_preset(PromptKind.GLOSSARY),
        chunk_char_limit=200,
    )


def _new_orchestrator(transport, cache_root: Path) -> GlossaryOrchestrator:
    counter = iter(range(1000))

    def id_factory() -> str:
        return f"task-{next(counter):04d}"

    return GlossaryOrchestrator(
        cache=TaskCache(root=cache_root),
        client=LlmClient(transport=transport),
        clock=_frozen_clock,
        id_factory=id_factory,
    )


def test_orchestrator_extracts_glossary_end_to_end_for_txt(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()

    (input_dir / "Novel.txt").write_text(
        "\n".join(
            [
                "신해범 entered the room.",
                "공이 followed silently.",
                "Then 흑룡 spoke.",
                "신해범 sat down.",
                "흑룡 nodded.",
                "신해범 was thinking.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    assert result.final_status is TaskStatus.COMPLETED
    assert len(result.glossary_outputs_per_file) == 1
    xlsx_path, json_path, references_path = result.glossary_outputs_per_file[0]
    assert xlsx_path.name == "Novel" + GLOSSARY_FILENAME_XLSX
    assert json_path.name == "Novel" + GLOSSARY_FILENAME_JSON
    assert references_path.name == "Novel" + GLOSSARY_FILENAME_REFERENCES

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    by_src = {row["src"]: row for row in payload}
    assert "신해범" in by_src
    assert by_src["신해범"]["dst"] == "申海范"
    # Frequency is segment-line count: 신해범 appears in 3 lines.
    assert by_src["신해범"]["frequency"] == 3
    assert by_src["흑룡"]["frequency"] == 2
    assert by_src["공이"]["frequency"] == 1

    workbook = load_workbook(xlsx_path)
    rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    assert rows[0] == ("src", "dst", "info", "regex", "frequency")
    assert {row[0] for row in rows[1:]} == {"신해범", "공이", "흑룡"}

    references_text = references_path.read_text(encoding="utf-8")
    assert "原文: 신해범" in references_text
    assert "참조" not in references_text  # ensure we use the design's labels
    assert "신해범 entered the room." in references_text


def test_orchestrator_writes_extraction_statistics_file(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Novel.txt").write_text(
        "\n".join(["신해범 line 1.", "신해범 line 2.", "공이 here."]) + "\n",
        encoding="utf-8",
    )

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    stats_path = output_dir / GLOSSARY_STATISTICS_FILENAME_JSON
    stats = json.loads(stats_path.read_text(encoding="utf-8"))
    assert stats["candidate_count"] >= 3
    assert stats["final_entry_count"] >= 1
    assert stats["failed_subtasks"] == 0
    assert stats["input_tokens"] >= 25


def test_orchestrator_records_failed_subtasks_in_statistics(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    big_body = "\n".join(f"신해범 line {n}." for n in range(40)) + "\n"
    (input_dir / "Novel.txt").write_text(big_body, encoding="utf-8")

    # Force the second of two chunks to fail.
    transport = CandidateEmittingTransport(forced_failures=(1,))
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(
        orchestrator.run(
            GlossaryConfig(
                input_dir=input_dir,
                output_dir=output_dir,
                source_language=Language.KOREAN,
                target_language=Language.CHINESE_SIMPLIFIED,
                model=_model(retry_attempts=0),
                prompt_preset=default_preset(PromptKind.GLOSSARY),
                chunk_char_limit=200,
            )
        )
    )

    assert result.final_status is TaskStatus.FAILED
    assert result.statistics.failed_subtasks == 1


def test_orchestrator_emits_one_artifact_set_per_input_file(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Alpha.txt").write_text("신해범 line\n", encoding="utf-8")
    (input_dir / "Bravo.txt").write_text("공이 line\n", encoding="utf-8")

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    output_basenames = {
        path.stem.split("-Glossary")[0]
        for triple in result.glossary_outputs_per_file
        for path in triple
    }
    assert output_basenames == {"Alpha", "Bravo"}


def test_orchestrator_returns_completed_status_for_empty_input(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    assert result.final_status is TaskStatus.COMPLETED
    assert result.glossary_outputs_per_file == ()
    assert result.statistics.candidate_count == 0
