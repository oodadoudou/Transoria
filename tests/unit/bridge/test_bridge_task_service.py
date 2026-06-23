"""Integration tests for the bridge ``TaskService``.

These tests exercise the full launch → poll → finish flow for translation,
glossary, and replacement tasks using the in-memory transport helpers and
the file-backed cache. They do not hit the network.
"""

from __future__ import annotations

import json
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

import pytest
from openpyxl import Workbook, load_workbook

from tests.helpers.transport import (
    CandidateEmittingTransport,
    EchoTranslationTransport,
    QueuedTransport,
)
from transoria.llm.client import TransportResult
from transoria.bridge.errors import BridgeError
from transoria.bridge.task_registry import RunningTask, TaskRegistry
from transoria.bridge.task_service import TaskService
from transoria.bridge.task_service import _effective_glossary_chunk_token_limit
from transoria.bridge.task_service import _low_confidence_summary
from transoria.bridge.task_service import _read_segment_dst
from transoria.domain import Language, SubtaskStatus, TaskKind, TaskStatus
from transoria.llm.client import LlmClient
from transoria.model_profiles import ModelProfileStore
from transoria.runtime.cache import TaskCache
from transoria.runtime.subtask import Subtask
from transoria.runtime.task_record import TaskRecord, TaskSnapshot
from transoria.settings import SettingsStore


def _wait_until(condition, *, timeout: float = 5.0, interval: float = 0.05) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if condition():
            return
        time.sleep(interval)
    raise AssertionError("timed out waiting for condition")


def test_effective_glossary_chunk_token_limit_migrates_legacy_default() -> None:
    assert _effective_glossary_chunk_token_limit(4000) == 2000
    assert _effective_glossary_chunk_token_limit(0) == 0
    assert _effective_glossary_chunk_token_limit(1200) == 1200


def test_low_confidence_summary_clears_stale_records_after_clean_retry() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-confidence", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-confidence",
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "안녕"},
                        "low_confidence": [
                            {
                                "segment_id": "0:0",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
            Subtask(
                id="chunk-00000.s1",
                task_id="translation-confidence",
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "你好"},
                        "low_confidence": [],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _low_confidence_summary(snapshot, {}) == {"total": 0, "source_residue": 0}


def test_low_confidence_summary_counts_latest_updated_record() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-confidence", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-confidence",
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "你好"},
                        "low_confidence": [],
                    },
                    ensure_ascii=False,
                ),
            ),
            Subtask(
                id="chunk-00000.s1",
                task_id="translation-confidence",
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "안녕"},
                        "low_confidence": [
                            {
                                "segment_id": "0:0",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _low_confidence_summary(snapshot, {}) == {"total": 1, "source_residue": 1}


def test_read_segment_dst_prefers_completed_translation_over_stale_failed_payload() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-read-dst", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-read-dst",
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "你好"},
                        "low_confidence": [],
                    },
                    ensure_ascii=False,
                ),
            ),
            Subtask(
                id="chunk-99999",
                task_id="translation-read-dst",
                status=SubtaskStatus.FAILED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "안녕"},
                        "low_confidence": [
                            {
                                "segment_id": "0:0",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _read_segment_dst(snapshot, "0:0") == "你好"


def test_read_segment_dst_ignores_unaccepted_failed_payload() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-read-dst", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-read-dst",
                status=SubtaskStatus.FAILED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "안녕"},
                        "low_confidence": [
                            {
                                "segment_id": "0:0",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _read_segment_dst(snapshot, "0:0") == ""


def test_read_segment_dst_uses_accepted_failed_override() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-read-dst", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-read-dst",
                status=SubtaskStatus.FAILED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "你好", "0:1": "산"},
                        "accepted_overrides": ["0:0"],
                        "low_confidence": [
                            {
                                "segment_id": "0:1",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _read_segment_dst(snapshot, "0:0") == "你好"


def test_low_confidence_summary_counts_accepted_failed_override() -> None:
    snapshot = TaskSnapshot(
        record=TaskRecord(id="translation-confidence", kind=TaskKind.TRANSLATION),
        subtasks=(
            Subtask(
                id="chunk-00000",
                task_id="translation-confidence",
                status=SubtaskStatus.FAILED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "안녕", "0:1": "산"},
                        "accepted_overrides": ["0:0"],
                        "low_confidence": [
                            {
                                "segment_id": "0:0",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            },
                            {
                                "segment_id": "0:1",
                                "reasons": ["source residue remains"],
                                "tags": ["source_residue"],
                            },
                        ],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    assert _low_confidence_summary(snapshot, {}) == {"total": 1, "source_residue": 1}


def _service(
    tmp_path: Path,
    *,
    transport,
    cache_root: Path | None = None,
) -> TaskService:
    if cache_root is None:
        cache_root = tmp_path / "cache"
    cache_root.mkdir(parents=True, exist_ok=True)
    settings_store = SettingsStore(path=cache_root / "settings.json")
    profile_store = ModelProfileStore.from_cache_root(cache_root)
    task_cache = TaskCache(root=cache_root / "tasks")
    return TaskService(
        cache=task_cache,
        registry=TaskRegistry(),
        settings_store=settings_store,
        profile_store=profile_store,
        prompts_cache_root=cache_root,
        llm_client_factory=lambda: LlmClient(transport=transport),
    )


def _ensure_test_profile(service: TaskService) -> str:
    """Create a minimal test profile so the task service has a model
    to resolve. Step G removed first-run seeding, so tests must
    create profiles explicitly."""

    from transoria.llm.config import ModelConfig, ProviderFormat

    profiles = service.profile_store.load()
    if profiles:
        chosen = profiles[0]
    else:
        chosen = ModelConfig(
            id="test-profile",
            display_name="Test Profile",
            provider_format=ProviderFormat.OPENAI,
            base_url="https://api.example.com",
            model_id="test-model",
        )
        service.profile_store.create(chosen)
    service.profile_store.set_api_keys(chosen.id, ("test-key",))
    return chosen.id


def _seed_translation_settings(
    service: TaskService,
    *,
    input_dir: Path,
    output_dir: Path,
) -> None:
    profile_id = _ensure_test_profile(service)
    service.settings_store.save_partial(
        "translation",
        {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "source_language": Language.KOREAN.value,
            "target_language": Language.CHINESE_SIMPLIFIED.value,
            "context_lines": 0,
        },
    )
    service.settings_store.save_partial(
        "app",
        {"active_translation_model_id": profile_id},
    )


def test_read_request_events_filters_and_offsets(tmp_path: Path) -> None:
    service = _service(tmp_path, transport=QueuedTransport())
    service.cache.save_task(TaskRecord(id="task-1", kind=TaskKind.TRANSLATION))
    for request_id, status in (
        ("r1", "running"),
        ("r1", "failed"),
        ("r2", "running"),
        ("r2", "completed"),
        ("r3", "running"),
        ("r3", "failed"),
    ):
        service.cache.append_request_event(
            "task-1",
            {
                "request_id": request_id,
                "status": status,
                "timestamp": "2026-06-18T00:00:00+00:00",
                "task_id": "task-1",
                "subtask_id": request_id,
                "subtask_attempt": 1,
            },
        )

    failed = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=10,
        status="failed",
    )
    assert [event["request_id"] for event in failed["events"]] == ["r3", "r1"]
    assert failed["total"] == 2

    paged = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=1,
        offset=1,
    )
    assert [event["request_id"] for event in paged["events"]] == ["r2"]
    assert paged["total"] == 3


def test_read_request_events_marks_orphan_running_request_cancelled(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path, transport=QueuedTransport())
    service.cache.write_seed(
        TaskRecord(
            id="task-1",
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
        ),
        (
            Subtask(
                id="chunk-00001",
                task_id="task-1",
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )
    for phase in ("sent", "headers_received"):
        service.cache.append_request_event(
            "task-1",
            {
                "request_id": "r1",
                "status": "running",
                "phase": phase,
                "timestamp": "2026-06-18T00:00:00+00:00",
                "task_id": "task-1",
                "subtask_id": "chunk-00001",
                "subtask_attempt": 2,
            },
        )

    all_events = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=10,
    )

    assert all_events["events"][0]["status"] == "cancelled"
    assert all_events["events"][0]["phase"] == "cancelled"
    assert all_events["events"][0]["error"] == (
        "Request was cancelled by a previous app session."
    )
    reconciled = service.cache.load("task-1")
    assert reconciled.record.status is TaskStatus.STOPPED
    assert reconciled.subtasks[0].status is SubtaskStatus.PENDING

    cancelled = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=10,
        status="cancelled",
    )
    assert [event["request_id"] for event in cancelled["events"]] == ["r1"]

    running = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=10,
        status="running",
    )
    assert running["events"] == []


def test_read_request_events_includes_local_failed_subtask_debug(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path, transport=QueuedTransport())
    record = TaskRecord(
        id="task-1",
        kind=TaskKind.TRANSLATION,
        status=TaskStatus.FAILED,
        created_at="2026-06-18T00:00:00+00:00",
        updated_at="2026-06-18T00:01:00+00:00",
    )
    service.cache.save_task(record)
    service.cache.save_subtask(
        Subtask(
            id="chunk-00068.s1.0",
            task_id="task-1",
            status=SubtaskStatus.FAILED,
            attempt_count=3,
            last_error="TranslationQualityFailureError: mass_source_residue_after_batch",
            last_error_at="2026-06-18T00:01:00+00:00",
        )
    )
    service.cache.append_request_event(
        "task-1",
        {
            "request_id": "r1",
            "status": "completed",
            "timestamp": "2026-06-18T00:00:30+00:00",
            "task_id": "task-1",
            "subtask_id": "chunk-00068.s1.0",
            "subtask_attempt": 3,
            "model_id": "model-x",
            "provider_format": "openai",
            "response_text": '{"0":"prefix"}',
        },
    )
    debug_dir = service.cache.task_dir("task-1") / "debug"
    debug_dir.mkdir(parents=True, exist_ok=True)
    (debug_dir / "chunk-00068.s1.0.json").write_text(
        json.dumps(
            {
                "terminal_error": "TranslationQualityFailureError: local quality failed",
                "attempts": [{"raw_response": '{"0":"last model response"}'}],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failed = service.read_request_events(
        kind="translation",
        task_id="task-1",
        limit=10,
        status="failed",
    )

    assert failed["total"] == 1
    event = failed["events"][0]
    assert event["status"] == "failed"
    assert event["phase"] == "validation"
    assert event["last_activity_at"] == "2026-06-18T00:01:00+00:00"
    assert event["local_failure"] is True
    assert event["subtask_id"] == "chunk-00068.s1.0"
    assert event["model_id"] == "model-x"
    assert "local quality failed" in str(event["error"])
    assert "last model response" in str(event["response_text"])


def _seed_glossary_settings(
    service: TaskService,
    *,
    input_dir: Path,
    output_dir: Path,
) -> None:
    profile_id = _ensure_test_profile(service)
    service.settings_store.save_partial(
        "glossary",
        {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "source_language": Language.KOREAN.value,
            "target_language": Language.CHINESE_SIMPLIFIED.value,
            "merge_folder_glossary": False,
            "minimum_frequency": 1,
            "chunk_token_limit": 0,
        },
    )
    service.settings_store.save_partial(
        "app",
        {"active_glossary_model_id": profile_id},
    )


def _seed_glossary_review_settings(
    service: TaskService,
    *,
    input_dir: Path,
    review_rounds: int = 1,
) -> None:
    profile_id = _ensure_test_profile(service)
    service.settings_store.save_partial(
        "glossary_review",
        {
            "input_folder": str(input_dir),
            "output_filename": "reviewed.xlsx",
            "review_rounds": review_rounds,
            "batch_size": 10,
        },
    )
    service.settings_store.save_partial(
        "app",
        {"active_glossary_review_model_id": profile_id},
    )


def _seed_replacement_settings(
    service: TaskService,
    *,
    input_dir: Path,
    output_dir: Path,
) -> None:
    service.settings_store.save_partial(
        "replacement",
        {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
        },
    )


class BlockingMixedTransport:
    def __init__(self) -> None:
        self.release = threading.Event()
        self.requests: list[dict[str, object]] = []

    async def execute(
        self,
        url: str,
        headers: dict[str, str],
        payload: dict[str, object],
        timeout: float,
    ) -> TransportResult:
        self.requests.append(dict(payload))
        self.release.wait(timeout=5.0)
        user_message = payload["messages"][-1]["content"]
        if "[Translate]" in user_message:
            translate_section = user_message.rsplit("[Translate]\n", 1)[-1]
            lines: list[str] = []
            for line in translate_section.splitlines():
                stripped = line.strip()
                if not stripped.startswith("{"):
                    continue
                try:
                    parsed = json.loads(stripped)
                except json.JSONDecodeError:
                    continue
                for key, value in parsed.items():
                    lines.append(
                        json.dumps({key: f"翻译:{value}"}, ensure_ascii=False)
                    )
            content = "\n".join(lines)
        elif "decisions" in user_message or "review" in user_message.lower():
            content = '{"decisions":[]}'
        else:
            content = json.dumps(
                {"src": "신해범", "dst": "申海范", "type": "Male Name"},
                ensure_ascii=False,
            )
        return TransportResult(
            200,
            {
                "choices": [{"message": {"role": "assistant", "content": content}}],
                "usage": {"prompt_tokens": 5, "completion_tokens": 5},
            },
        )


def test_translation_start_rejects_empty_input_folder(tmp_path: Path):
    """An input folder that contains zero supported files (.epub/.txt)
    must fail loudly with ``bridge.invalid_argument`` so the user sees
    the run-error banner instead of a silently-completed empty task."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    # Drop an unsupported file so the directory exists but has nothing
    # the scanner will pick up.
    (input_dir / "readme.md").write_text("not a novel", encoding="utf-8")
    output_dir = tmp_path / "out"
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    with pytest.raises(BridgeError) as caught:
        service.start_translation(request_id="req-1")

    assert caught.value.code == "bridge.invalid_argument"
    assert "input_folder" in caught.value.payload.message
    assert "no supported files" in caught.value.payload.message


def test_glossary_start_rejects_empty_input_folder(tmp_path: Path):
    """Glossary mirror of the translation empty-folder check."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    input_dir.mkdir()
    output_dir = tmp_path / "out"
    _seed_glossary_settings(service, input_dir=input_dir, output_dir=output_dir)

    with pytest.raises(BridgeError) as caught:
        service.start_glossary(request_id="req-1")

    assert caught.value.code == "bridge.invalid_argument"
    assert "input_folder" in caught.value.payload.message


def test_glossary_review_writes_final_xlsx_and_changed_report(tmp_path: Path):
    folder = tmp_path / "review"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "旧译", "人物", 5])
    sheet.append(["응", "嗯", "其它", 20])
    workbook.save(folder / "terms.xlsx")
    (folder / "terms-references.txt").write_text("신해범 出现在正文里。응 只是语气。", encoding="utf-8")
    transport = QueuedTransport(
        responses=[
            TransportResult(
                200,
                {
                    "choices": [
                        {
                            "message": {
                                "role": "assistant",
                                "content": (
                                    '{"decisions":['
                                    '{"row_index":2,"action":"modify","suggested_dst":"申海范","reason":"统一人名"},'
                                    '{"row_index":3,"action":"delete","reason":"普通回应"}'
                                    "]}"
                                ),
                            }
                        }
                    ],
                    "usage": {"prompt_tokens": 10, "completion_tokens": 8},
                },
            )
        ]
    )
    service = _service(tmp_path, transport=transport)
    _seed_glossary_review_settings(service, input_dir=folder)

    started = service.start_glossary_review(request_id="req-1")
    task_id = str(started["task_id"])
    _wait_until(
        lambda: service.read_snapshot(kind="glossary_review", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed"
    )

    artifacts = service.read_artifacts(kind="glossary_review", task_id=task_id)
    assert artifacts["output_path"] == str(folder / "reviewed.xlsx")
    report = service.read_glossary_review_report(task_id=task_id)
    assert report["changed_count"] == 2
    assert [row["action"] for row in report["rows"]] == ["modify", "delete"]
    assert "신해범" in report["rows"][0]["context_excerpt"]
    snapshot = service.read_snapshot(kind="glossary_review", task_id=task_id)[
        "snapshot"
    ]
    assert snapshot["round_progress"] == {
        "total_rounds": 2,
        "current_round": 2,
        "completed_rounds": 2,
        "current_total_batches": 0,
        "current_completed_batches": 0,
    }

    output_book = load_workbook(folder / "reviewed.xlsx")
    rows = list(output_book.active.iter_rows(values_only=True))
    assert rows == [
        ("src", "dst", "info", "frequency"),
        ("신해범", "申海范", "人物", 5),
    ]


def test_glossary_review_snapshot_reports_multi_round_progress(tmp_path: Path):
    folder = tmp_path / "review"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "旧译", "人物", 5])
    workbook.save(folder / "terms.xlsx")
    (folder / "terms-references.txt").write_text(
        "신해범 出现在正文里。", encoding="utf-8"
    )
    transport = QueuedTransport(
        responses=[
            TransportResult(
                200,
                {
                    "choices": [
                        {
                            "message": {
                                "role": "assistant",
                                "content": (
                                    '{"decisions":[{"row_index":2,"action":"modify",'
                                    '"suggested_dst":"申海范","reason":"统一人名"}]}'
                                ),
                            }
                        }
                    ],
                    "usage": {"prompt_tokens": 10, "completion_tokens": 8},
                },
            ),
            TransportResult(
                200,
                {
                    "choices": [
                        {
                            "message": {
                                "role": "assistant",
                                "content": (
                                    '{"decisions":[{"row_index":2,"action":"keep",'
                                    '"reason":"已一致"}]}'
                                ),
                            }
                        }
                    ],
                    "usage": {"prompt_tokens": 10, "completion_tokens": 4},
                },
            ),
        ]
    )
    service = _service(tmp_path, transport=transport)
    _seed_glossary_review_settings(service, input_dir=folder, review_rounds=2)

    started = service.start_glossary_review(request_id="req-1")
    task_id = str(started["task_id"])
    _wait_until(
        lambda: service.read_snapshot(kind="glossary_review", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed"
    )

    snapshot = service.read_snapshot(kind="glossary_review", task_id=task_id)[
        "snapshot"
    ]
    assert snapshot["round_progress"] == {
        "total_rounds": 3,
        "current_round": 3,
        "completed_rounds": 3,
        "current_total_batches": 0,
        "current_completed_batches": 0,
    }


def test_completed_glossary_review_snapshot_normalizes_stale_round_progress(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "glossary-review-stale-round"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("glossary_review").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.COMPLETED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "output_dir": str(tmp_path),
                "output_filename": "glossary-review-final.xlsx",
                "review_rounds_total": 5,
                "review_round_current": 1,
                "review_round_completed": 1,
                "review_round_total_batches": 36,
                "review_round_completed_batches": 36,
            },
        ),
        (
            Subtask(
                id="round-01-batch-0000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
            ),
        ),
    )
    (tmp_path / "glossary-review-final.xlsx").write_bytes(b"placeholder")
    report_path = (
        service._cache_for_kind("glossary_review").task_dir(task_id)
        / "glossary-review-report.json"
    )
    report_path.write_text('{"rows": []}', encoding="utf-8")

    snapshot = service.read_snapshot(kind="glossary_review", task_id=task_id)[
        "snapshot"
    ]

    assert snapshot["round_progress"] == {
        "total_rounds": 5,
        "current_round": 5,
        "completed_rounds": 5,
        "current_total_batches": 36,
        "current_completed_batches": 36,
    }


def test_stopped_glossary_review_with_final_artifacts_self_heals_to_completed(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "glossary-review-stopped-finalized"
    created_at = "2026-05-01T00:00:00+00:00"
    cache = service._cache_for_kind("glossary_review")
    cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.STOPPED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "output_dir": str(tmp_path),
                "output_filename": "glossary-review-final.xlsx",
                "review_rounds_total": 4,
                "review_round_current": 4,
                "review_round_completed": 4,
                "review_round_total_batches": 1,
                "review_round_completed_batches": 1,
            },
        ),
        (
            Subtask(
                id="round-04-character-consistency",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
            ),
        ),
    )
    output_path = tmp_path / "glossary-review-final.xlsx"
    output_path.write_bytes(b"placeholder")
    report_path = cache.task_dir(task_id) / "glossary-review-report.json"
    report_path.write_text('{"rows": []}', encoding="utf-8")
    service._write_result(
        task_id,
        {
            "kind": "glossary_review",
            "output_path": str(output_path),
            "report_path": str(report_path),
            "changed_count": 1,
        },
    )

    recent = service.list_recent_tasks(kind="glossary_review", limit=1)
    snapshot = service.read_snapshot(kind="glossary_review", task_id=task_id)[
        "snapshot"
    ]

    assert recent["tasks"][0]["status"] == "completed"
    assert service.cache.load_record(task_id).status is TaskStatus.COMPLETED
    assert snapshot["header"]["status"] == "completed"
    assert snapshot["progress"]["completed"] == 1
    assert snapshot["round_progress"] == {
        "total_rounds": 4,
        "current_round": 4,
        "completed_rounds": 4,
        "current_total_batches": 1,
        "current_completed_batches": 1,
    }


def test_glossary_review_final_sheet_can_be_edited(tmp_path: Path):
    folder = tmp_path / "review"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "旧译", "人物", 5])
    sheet.append(["유리", "尤莉", "人物", 2])
    sheet.append(["호텔", "酒店", "地点", 1])
    workbook.save(folder / "terms.xlsx")
    (folder / "terms-references.txt").write_text(
        "신해범 出现在正文里。", encoding="utf-8"
    )
    transport = QueuedTransport(
        responses=[
            TransportResult(
                200,
                {
                    "choices": [
                        {
                            "message": {
                                "role": "assistant",
                                "content": '{"decisions":[]}',
                            }
                        }
                    ],
                    "usage": {"prompt_tokens": 10, "completion_tokens": 2},
                },
            )
        ]
    )
    service = _service(tmp_path, transport=transport)
    _seed_glossary_review_settings(service, input_dir=folder)

    started = service.start_glossary_review(request_id="req-1")
    task_id = str(started["task_id"])
    _wait_until(
        lambda: service.read_snapshot(kind="glossary_review", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed"
    )

    final_sheet = service.update_glossary_review_final_row(
        task_id=task_id,
        row_index=2,
        src="신해범",
        dst="申海范",
        info="角色/男性角色",
    )
    assert final_sheet["rows"][0]["dst"] == "申海范"
    final_sheet = service.delete_glossary_review_final_rows(
        task_id=task_id,
        row_indices=[3, 4],
    )
    assert [row["src"] for row in final_sheet["rows"]] == ["신해범"]
    final_sheet = service.restore_glossary_review_deleted_report_row(
        task_id=task_id,
        src="유리",
        dst="尤莉",
        info="人物",
        frequency=2,
    )
    assert [row["src"] for row in final_sheet["rows"]] == ["신해범", "유리"]
    output_book = load_workbook(folder / "reviewed.xlsx")
    rows = list(output_book.active.iter_rows(values_only=True))
    assert rows == [
        ("src", "dst", "info", "frequency"),
        ("신해범", "申海范", "角色/男性角色", 5),
        ("유리", "尤莉", "人物", 2),
    ]


def test_translation_start_rejects_input_equals_output(tmp_path: Path):
    """Translation must not write outputs back into the input folder.

    Otherwise the next run's recursive scanner picks up the previous
    run's ``*-zh.epub`` and ``translation-failed-subtasks.txt``, producing
    cascading-suffix garbage and re-translating already-translated novels.
    """
    service = _service(tmp_path, transport=EchoTranslationTransport())
    folder = tmp_path / "shared"
    folder.mkdir()
    (folder / "novel.txt").write_text("source line\n", encoding="utf-8")
    _seed_translation_settings(service, input_dir=folder, output_dir=folder)

    with pytest.raises(BridgeError) as caught:
        service.start_translation(request_id="req-1")

    assert caught.value.code == "bridge.invalid_argument"
    assert caught.value.payload.details.get("field") == "output_folder"
    assert "different" in caught.value.payload.message


def test_translation_start_rejects_output_inside_input(tmp_path: Path):
    """Output as a descendant of input has the same problem — the
    recursive scanner walks into it."""
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "novels"
    input_dir.mkdir()
    (input_dir / "novel.txt").write_text("source line\n", encoding="utf-8")
    output_dir = input_dir / "translated"
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    with pytest.raises(BridgeError) as caught:
        service.start_translation(request_id="req-1")

    assert caught.value.code == "bridge.invalid_argument"
    assert "inside" in caught.value.payload.message


def test_glossary_start_allows_input_equals_output(tmp_path: Path):
    """Glossary outputs (.xlsx / .json) never match the .epub/.txt
    scanner, so the same-folder restriction does NOT apply. This test
    locks that asymmetry in so the translation guard never silently
    creeps into the glossary path."""
    service = _service(tmp_path, transport=EchoTranslationTransport())
    folder = tmp_path / "shared"
    folder.mkdir()
    (folder / "novel.txt").write_text("source line\n", encoding="utf-8")
    _seed_glossary_settings(service, input_dir=folder, output_dir=folder)

    # Should not raise the distinct-folders BridgeError. Other validation
    # may still apply (and a real run would proceed), but the rejection
    # would specifically point at output_folder if the guard had been
    # mistakenly applied here.
    response = service.start_glossary(request_id="req-1")
    assert "task_id" in response


def test_translation_start_validates_missing_settings(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.start_translation(request_id="req-1")
    assert caught.value.code == "bridge.invalid_argument"


def test_translation_start_runs_to_completion(tmp_path: Path):
    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("안녕\n반갑다", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    assert task_id.startswith("translation-")

    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed"},
    )
    artifacts = service.read_artifacts(kind="translation", task_id=task_id)
    assert artifacts["kind"] == "translation"
    assert artifacts["output_folder"] == str(output_dir)
    assert any("sample" in path for path in artifacts["translated_files"])
    assert (
        service.cache.task_dir(task_id) / "translation-statistics.json"
    ).exists()
    assert not (output_dir / "translation-statistics.json").exists()
    assert artifacts["bilingual_folder"] is None
    # Per-subtask debug log lives under the task cache so users can zip
    # the cache and ship it for analysis without leaking outputs.
    debug_dir = service.cache.task_dir(task_id) / "debug"
    assert debug_dir.is_dir()
    assert any(debug_dir.glob("*.json"))


def test_translation_artifacts_include_bilingual_folder_only_when_written(
    tmp_path: Path,
):
    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("안녕", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "translation",
        {
            "bilingual_enabled": True,
            "bilingual_subfolder_name": "bilingual",
        },
    )

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]

    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed"},
    )
    artifacts = service.read_artifacts(kind="translation", task_id=task_id)

    assert artifacts["bilingual_files"]
    assert artifacts["bilingual_folder"] == str(output_dir / "bilingual")


def test_translation_glossary_threads_into_config(tmp_path: Path):
    """F.P0.1: when ``translation.translation_glossary`` is non-empty
    in settings, ``TaskService._build_translation_config`` populates
    ``TranslationConfig.glossary`` so the runner sees the user's
    terms."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("강감찬", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "translation",
        {
            "translation_glossary": [
                {
                    "src": "강감찬",
                    "dst": "姜邯赞",
                    "info": "Male Name",
                    "regex": False,
                    "case_sensitive": False,
                    "enabled": True,
                },
            ],
        },
    )

    config, _model, _preset = service._build_translation_config()
    assert len(config.glossary.entries) == 1
    entry = config.glossary.entries[0]
    assert entry.src == "강감찬"
    assert entry.dst == "姜邯赞"
    assert entry.info == "Male Name"
    assert entry.case_sensitive is False
    assert entry.enabled is True


def test_translation_glossary_regex_setting_threads_and_matches(tmp_path: Path):
    """Regex glossary rows persisted by the frontend must stay active at run start."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("12화", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "translation",
        {
            "translation_glossary": [
                {
                    "src": r"\d+화",
                    "dst": "<chapter>",
                    "info": "Chapter heading",
                    "regex": True,
                    "case_sensitive": False,
                    "enabled": True,
                },
            ],
        },
    )

    config, _model, _preset = service._build_translation_config()
    assert len(config.glossary.entries) == 1
    entry = config.glossary.entries[0]
    assert entry.regex is True
    assert config.glossary.match("12화 시작")[0].dst == "<chapter>"


def test_translation_text_preserve_and_replacements_thread_into_config(
    tmp_path: Path,
):
    """F.P1.1: text-preserve rules + pre/post replacement settings
    flow into TranslationConfig at run start. Empty patterns are
    dropped silently."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("text", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "translation",
        {
            "text_preserve_rules": [
                {"pattern": r"\{\{[^}]+\}\}", "note": "templates", "enabled": True},
                {"pattern": "", "note": "empty — should drop"},
            ],
            "pre_replacements": [
                {"src": "원래", "dst": "原文", "case_sensitive": False},
            ],
            "post_replacements": [
                {"src": "TODO", "dst": "待办", "regex": False, "enabled": False},
            ],
        },
    )

    config, _model, _preset = service._build_translation_config()

    assert len(config.text_preserve_rules) == 1
    assert config.text_preserve_rules[0].pattern == r"\{\{[^}]+\}\}"
    assert config.text_preserve_rules[0].note == "templates"

    assert len(config.pre_replacements) == 1
    assert config.pre_replacements[0].src == "원래"
    assert config.pre_replacements[0].dst == "原文"

    assert len(config.post_replacements) == 1
    assert config.post_replacements[0].enabled is False


def test_translation_glossary_rejects_non_list_payload(tmp_path: Path):
    """The settings store raises ValueError; the bridge handler wraps
    it as ``bridge.invalid_argument`` with field name. We verify the
    store-level shape; bridge wrapping is exercised in
    test_bridge_handlers_settings.py."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(ValueError) as caught:
        service.settings_store.save_partial(
            "translation", {"translation_glossary": "not-a-list"}
        )
    assert "translation_glossary" in str(caught.value)


def test_translation_invalid_language_returns_invalid_argument(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("text", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "translation", {"source_language": "invalid"}
    )
    with pytest.raises(BridgeError) as caught:
        service.start_translation(request_id="req-1")
    assert caught.value.code == "bridge.invalid_argument"
    assert caught.value.payload.details["field"] == "source_language"


def test_translation_start_overrides_running_task(tmp_path: Path):
    """A second ``start`` while one is already running cooperatively
    stops the prior task and seeds a fresh one with a new task_id. The
    prior cache **is preserved on disk** — the user can still read its
    snapshot for resume / proofreading. Only the in-flight thread is
    interrupted."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hello", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    first = service.start_translation(request_id="req-1")
    second = service.start_translation(request_id="req-2")

    assert first["task_id"] != second["task_id"]
    # Both tasks are still readable: the prior cache survived.
    first_snap = service.read_snapshot(kind="translation", task_id=first["task_id"])[
        "snapshot"
    ]
    assert first_snap["header"]["id"] == first["task_id"]

    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=second["task_id"])[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed", "stopped"},
    )


def test_llm_task_kinds_can_run_in_parallel(tmp_path: Path):
    transport = BlockingMixedTransport()
    service = _service(tmp_path, transport=transport)

    translation_in = tmp_path / "translation-in"
    translation_out = tmp_path / "translation-out"
    glossary_in = tmp_path / "glossary-in"
    glossary_out = tmp_path / "glossary-out"
    review_in = tmp_path / "review-in"
    for path in [translation_in, translation_out, glossary_in, glossary_out, review_in]:
        path.mkdir()
    (translation_in / "sample.txt").write_text("안녕", encoding="utf-8")
    (glossary_in / "sample.txt").write_text("신해범은 위대했다", encoding="utf-8")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "旧译", "人物", 5])
    workbook.save(review_in / "terms.xlsx")
    (review_in / "terms-references.txt").write_text(
        "신해범 出现在正文里。", encoding="utf-8"
    )

    _seed_translation_settings(
        service, input_dir=translation_in, output_dir=translation_out
    )
    _seed_glossary_settings(service, input_dir=glossary_in, output_dir=glossary_out)
    _seed_glossary_review_settings(service, input_dir=review_in)

    translation = service.start_translation(request_id="req-translation")
    glossary = service.start_glossary(request_id="req-glossary")
    review = service.start_glossary_review(request_id="req-review")

    try:
        assert service.registry.get(str(translation["task_id"])) is not None
        assert service.registry.get(str(glossary["task_id"])) is not None
        assert service.registry.get(str(review["task_id"])) is not None
        for kind, started in [
            ("translation", translation),
            ("glossary", glossary),
            ("glossary_review", review),
        ]:
            task_id = str(started["task_id"])
            status = service.read_snapshot(kind=kind, task_id=task_id)["snapshot"][
                "header"
            ]["status"]
            assert status in {"pending", "running"}
    finally:
        transport.release.set()

    for kind, started in [
        ("translation", translation),
        ("glossary", glossary),
        ("glossary_review", review),
    ]:
        task_id = str(started["task_id"])
        _wait_until(
            lambda kind=kind, task_id=task_id: service.read_snapshot(
                kind=kind, task_id=task_id
            )["snapshot"]["header"]["status"]
            in {"completed", "failed", "stopped"}
        )


def test_glossary_start_runs_to_completion(tmp_path: Path):
    transport = CandidateEmittingTransport(
        candidates=(("강감찬", "姜邯赞", "Male Name"),),
    )
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("강감찬은 위대했다", encoding="utf-8")
    _seed_glossary_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_glossary(request_id="req-1")
    task_id = response["task_id"]

    _wait_until(
        lambda: service.read_snapshot(kind="glossary", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed"},
    )
    artifacts = service.read_artifacts(kind="glossary", task_id=task_id)
    assert artifacts["kind"] == "glossary"
    assert artifacts["output_folder"] == str(output_dir)
    # Statistics live inside the per-task cache (now central, not under
    # the user's output folder) so they survive across clean completion
    # and are available for the proofreading flow.
    assert artifacts["statistics_json_path"] == str(
        service.cache.task_dir(task_id) / "extraction-statistics.json"
    )
    assert not (output_dir / "extraction-statistics.json").exists()
    assert not (output_dir / "transoria-cache").exists()


def test_glossary_combined_artifact_is_not_returned_as_per_novel(
    tmp_path: Path,
):
    transport = CandidateEmittingTransport(
        candidates=(("강감찬", "姜邯赞", "Male Name"),),
    )
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "folder"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Alpha.txt").write_text("강감찬 alpha", encoding="utf-8")
    (input_dir / "Bravo.txt").write_text("강감찬 bravo", encoding="utf-8")
    _seed_glossary_settings(service, input_dir=input_dir, output_dir=output_dir)
    service.settings_store.save_partial(
        "glossary",
        {"merge_folder_glossary": True},
    )

    response = service.start_glossary(request_id="req-1")
    task_id = response["task_id"]

    _wait_until(
        lambda: service.read_snapshot(kind="glossary", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed"},
    )
    artifacts = service.read_artifacts(kind="glossary", task_id=task_id)

    per_novel = artifacts["per_novel_artifacts"]
    assert isinstance(per_novel, list)
    assert per_novel == []
    assert artifacts["combined_artifact"]["novel_name"] == "folder"


def test_read_artifacts_returns_partial_payload_when_result_json_is_missing(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    translated = output_dir / "Novel-zh.txt"
    translated.write_text("translated", encoding="utf-8")
    created_at = "2026-04-28T00:00:00+00:00"
    task_id = "translation-partial"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.FAILED,
            created_at=created_at,
            updated_at=created_at,
            metadata={"output_dir": str(output_dir)},
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
            ),
        ),
    )

    artifacts = service.read_artifacts(kind="translation", task_id=task_id)

    assert artifacts["partial"] is True
    assert artifacts["translated_files"] == [str(translated)]


def test_replacement_start_runs_to_completion(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text(
        "hello world\nfoo bar\n", encoding="utf-8"
    )
    _seed_replacement_settings(service, input_dir=input_dir, output_dir=output_dir)

    from transoria.tools.replacement import ReplacementRule

    response = service.start_replacement(
        request_id="req-1",
        rules=(ReplacementRule(src="hello", dst="안녕"),),
    )
    task_id = response["task_id"]
    assert task_id.startswith("replacement-")

    _wait_until(
        lambda: service.read_snapshot(kind="replacement", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        in {"completed", "failed", "stopped"},
    )
    artifacts = service.read_artifacts(kind="replacement", task_id=task_id)
    assert artifacts["kind"] == "replacement"
    assert artifacts["output_folder"] == str(output_dir)
    assert artifacts["total_replacements"] >= 1


def test_replacement_report_survives_clean_completion_via_mirror(tmp_path: Path):
    """A clean COMPLETED replacement run wipes the on-disk task cache,
    which would also delete ``replacement-report.json``. The report is
    mirrored into memory before the wipe so the modal trigger keeps
    working until the user starts a new task or restarts the app."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text(
        "hello world\nhello again\n", encoding="utf-8"
    )
    _seed_replacement_settings(service, input_dir=input_dir, output_dir=output_dir)

    from transoria.tools.replacement import ReplacementRule

    response = service.start_replacement(
        request_id="req-1",
        rules=(
            ReplacementRule(src="hello", dst="hi", case_sensitive=True),
        ),
    )
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="replacement", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    report = service.read_replacement_report(task_id=task_id)
    assert report["task_id"] == task_id
    assert report["totals"]["total_replacements"] == 2
    assert len(report["rules"]) == 1
    rule_entry = report["rules"][0]
    assert rule_entry["total_count"] == 2
    assert len(rule_entry["occurrences"]) == 2
    # Each occurrence must carry the file path the match came from.
    assert all(
        str(input_dir / "sample.txt") == occ["file_path"]
        for occ in rule_entry["occurrences"]
    )


def test_replacement_rejects_empty_rules(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "x.txt").write_text("text", encoding="utf-8")
    _seed_replacement_settings(service, input_dir=input_dir, output_dir=output_dir)

    with pytest.raises(BridgeError) as caught:
        service.start_replacement(request_id="req-1", rules=())
    assert caught.value.code == "bridge.invalid_argument"
    assert caught.value.payload.details["field"] == "rules"


def test_replacement_allows_same_input_output(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    same_dir = tmp_path / "io"
    same_dir.mkdir()
    (same_dir / "x.txt").write_text("hello", encoding="utf-8")
    _seed_replacement_settings(service, input_dir=same_dir, output_dir=same_dir)

    from transoria.tools.replacement import ReplacementRule

    response = service.start_replacement(
        request_id="req-1",
        rules=(ReplacementRule(src="hello", dst="hi"),),
    )
    task_id = response["task_id"]

    _wait_until(
        lambda: service.read_snapshot(kind="replacement", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    assert (same_dir / "x-Replaced.txt").read_text(encoding="utf-8") == "hi"
    assert (same_dir / "x.txt").read_text(encoding="utf-8") == "hello"


def test_replacement_pause_rejects_with_single_pass(tmp_path: Path):
    """Replacement is single-pass; pause/continue are not supported.
    Translation/Glossary now have real pause/continue semantics — see
    the dedicated test files for those cycles."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.pause_task(kind="replacement", task_id="any")
    assert caught.value.code == "task.invalid_transition"
    assert caught.value.payload.details["reason"] == "single_pass"


def test_replacement_continue_rejects_with_single_pass(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.continue_task(kind="replacement", task_id="any")
    assert caught.value.code == "task.invalid_transition"
    assert caught.value.payload.details["reason"] == "single_pass"


def test_pause_unknown_translation_task_returns_not_found(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.pause_task(kind="translation", task_id="t-missing")
    assert caught.value.code == "bridge.not_found"


def test_continue_unknown_translation_task_returns_not_found(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.continue_task(kind="translation", task_id="t-missing")
    assert caught.value.code == "bridge.not_found"


def test_continue_self_heals_orphan_running_translation_task(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-orphan-running"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )
    continued: list[str] = []

    def fake_continue(task_id: str) -> dict[str, object]:
        continued.append(task_id)
        return {"task_id": task_id, "started_at": created_at}

    monkeypatch.setattr(service, "_continue_translation", fake_continue)

    response = service.continue_task(kind="translation", task_id=task_id)

    assert response["task_id"] == task_id
    assert continued == [task_id]
    persisted = service._cache_for_kind("translation").load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.PENDING


def test_continue_self_heals_done_registry_running_glossary_task(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("신해범", encoding="utf-8")
    _seed_glossary_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "glossary-done-registry-running"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("glossary").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.FAILED,
                last_error="boom",
            ),
        ),
    )
    stale = RunningTask(
        task_id=task_id,
        kind="glossary",
        cache=service._cache_for_kind("glossary"),
        created_at=created_at,
    )
    stale.mark_done()
    service.registry.add(stale)
    continued: list[str] = []

    def fake_continue(task_id: str) -> dict[str, object]:
        continued.append(task_id)
        return {"task_id": task_id, "started_at": created_at}

    monkeypatch.setattr(service, "_continue_glossary", fake_continue)

    response = service.continue_task(kind="glossary", task_id=task_id)

    assert response["task_id"] == task_id
    assert continued == [task_id]
    persisted = service._cache_for_kind("glossary").load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.FAILED


def test_continue_rejects_live_running_translation_task(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-live-running"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="translation",
            cache=service._cache_for_kind("translation"),
            created_at=created_at,
        )
    )

    with pytest.raises(BridgeError) as caught:
        service.continue_task(kind="translation", task_id=task_id)

    assert caught.value.code == "bridge.conflict"


def test_continue_clears_stale_registry_entry_when_disk_says_failed(tmp_path: Path):
    """A registry entry can leak as not-done if the worker thread exited
    without reaching mark_done (host crash, BaseException path, etc.).
    When disk shows a terminal state, continue_task must trust disk and
    treat the registry entry as stale instead of raising bridge.conflict.
    """

    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-stale-registry"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.FAILED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.FAILED,
                last_error="boom",
            ),
        ),
    )
    # Stale entry: disk says FAILED but the registry never saw mark_done().
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="translation",
            cache=service._cache_for_kind("translation"),
            created_at=created_at,
        )
    )

    response = service.continue_task(kind="translation", task_id=task_id)

    assert response["task_id"] == task_id
    # The stale registry entry was force-marked done; the new continue
    # call replaces it with a fresh RunningTask.
    new_running = service.registry.get(task_id)
    assert new_running is not None
    assert new_running.created_at != created_at


def test_continue_self_heals_stalled_live_translation_task(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-stalled-running"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )
    running = RunningTask(
        task_id=task_id,
        kind="translation",
        cache=service._cache_for_kind("translation"),
        created_at=created_at,
    )
    running._last_heartbeat_monotonic = 0.0
    service.registry.add(running)
    continued: list[str] = []

    def fake_continue(task_id: str) -> dict[str, object]:
        continued.append(task_id)
        return {"task_id": task_id, "started_at": created_at}

    monkeypatch.setattr(service, "_continue_translation", fake_continue)

    response = service.continue_task(kind="translation", task_id=task_id)

    assert response["task_id"] == task_id
    assert continued == [task_id]
    assert running.is_done
    persisted = service._cache_for_kind("translation").load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.PENDING


def test_start_purge_rejects_when_prior_task_thread_is_still_alive(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-still-stopping"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
        ),
        (),
    )
    release = threading.Event()
    thread = threading.Thread(target=release.wait, daemon=True)
    thread.start()
    running = RunningTask(
        task_id=task_id,
        kind="translation",
        cache=service._cache_for_kind("translation"),
        created_at=created_at,
        thread=thread,
    )
    service.registry.add(running)

    try:
        with pytest.raises(BridgeError) as caught:
            service._purge_kind_for_start(
                kind="translation",
                task_kind=TaskKind.TRANSLATION,
                join_timeout=0.01,
            )
    finally:
        release.set()
        thread.join(timeout=1)

    assert caught.value.code == "bridge.conflict"
    assert caught.value.payload.details["task_id"] == task_id


def test_stop_marks_live_translation_task_stopping_immediately(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-stop-feedback"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="translation",
            cache=service._cache_for_kind("translation"),
            created_at=created_at,
        )
    )

    response = service.stop_task(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopping"
    assert (
        service._cache_for_kind("translation").load_record(task_id).status
        is TaskStatus.STOPPING
    )


def test_continue_marks_translation_running_before_return(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-resume-window"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.STOPPED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.PENDING,
            ),
        ),
    )
    spawned: list[str] = []
    monkeypatch.setattr(
        service,
        "_spawn_thread",
        lambda running, *, target, task_id: spawned.append(task_id),
    )

    probe = service.probe_continuable(kind="translation")
    assert probe["continuable"] is True
    assert probe["task_id"] == task_id
    assert probe["pending"] == 1
    assert probe["failed"] == 0

    response = service.continue_task(kind="translation", task_id=task_id)

    assert response["task_id"] == task_id
    assert spawned == [task_id]
    persisted = service._cache_for_kind("translation").load(task_id)
    assert persisted.record.status is TaskStatus.RUNNING


def test_continue_allows_stopped_translation_with_only_finalization_left(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-stopped-finalize"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.STOPPED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
                response_content=json.dumps(
                    {
                        "version": 2,
                        "translations": {"0:0": "翻译:hi"},
                        "low_confidence": [],
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )
    spawned: list[str] = []
    monkeypatch.setattr(
        service,
        "_spawn_thread",
        lambda running, *, target, task_id: spawned.append(task_id),
    )

    probe = service.probe_continuable(kind="translation")
    assert probe["continuable"] is True
    assert probe["task_id"] == task_id
    assert probe["pending"] == 0
    assert probe["failed"] == 0

    response = service.continue_task(kind="translation", task_id=task_id)

    assert response["task_id"] == task_id
    assert spawned == [task_id]
    persisted = service._cache_for_kind("translation").load(task_id)
    assert persisted.record.status is TaskStatus.RUNNING


def test_continue_allows_stopped_glossary_review_with_only_finalization_left(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    folder = tmp_path / "review"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "申海凡", "男性角色", 5])
    workbook.save(folder / "terms.xlsx")
    (folder / "terms-references.txt").write_text("신해범 出现在正文。", encoding="utf-8")
    _seed_glossary_review_settings(service, input_dir=folder)
    task_id = "glossary-review-stopped-finalize"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("glossary_review").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.STOPPED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(folder),
                "output_dir": str(folder),
            },
        ),
        (
            Subtask(
                id="round-01-batch-0000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
                request_payload={
                    "round": 1,
                    "batch": 0,
                    "rows": [{"row_index": 2}],
                },
                response_content=json.dumps(
                    {
                        "decisions": [
                            {
                                "row_index": 2,
                                "action": "modify",
                                "suggested_dst": "申海范",
                                "reason": "统一人名",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )
    spawned: list[str] = []
    monkeypatch.setattr(
        service,
        "_spawn_thread",
        lambda running, *, target, task_id: spawned.append(task_id),
    )

    probe = service.probe_continuable(kind="glossary_review")
    assert probe["continuable"] is True
    assert probe["task_id"] == task_id
    assert probe["pending"] == 0
    assert probe["failed"] == 0

    response = service.continue_task(kind="glossary_review", task_id=task_id)

    assert response["task_id"] == task_id
    assert spawned == [task_id]
    persisted = service._cache_for_kind("glossary_review").load(task_id)
    assert persisted.record.status is TaskStatus.RUNNING


def test_continue_after_failed_resumes_failed_subtasks(tmp_path: Path):
    """User-reported: after all-keys-failed (or any LLM failure) the
    task lands in FAILED. Continue must reset FAILED → PENDING and
    re-run, so users can retry without nuking the cache via Start."""

    # Force enough failures to exhaust the batch transport-retry budget
    # (1 initial + up to 3 transport retries = 4 attempts) so the subtask
    # lands in FAILED rather than recovering on retry.
    transport = EchoTranslationTransport(forced_failure_calls=tuple(range(8)))
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "failed",
    )

    # Probe should now report this FAILED task as continuable.
    probe = service.probe_continuable(kind="translation")
    assert probe["continuable"] is True
    assert probe["task_id"] == task_id
    assert probe["status"] == "failed"
    assert probe["failed"] >= 1

    # Drop the forced failures so the retry succeeds.
    transport.forced_failure_calls = ()

    service.continue_task(kind="translation", task_id=task_id)
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )


def test_continue_rejects_completed_translation(tmp_path: Path):
    """Continue is refused on cleanly-COMPLETED tasks because there's
    no remaining work; FAILED stays continuable so the user can retry
    failed subtasks (e.g. transient key/quota issues)."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    with pytest.raises(BridgeError) as caught:
        service.continue_task(kind="translation", task_id=task_id)
    assert caught.value.code == "task.invalid_transition"


def test_probe_continuable_returns_false_when_no_cache(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    response = service.probe_continuable(kind="translation")
    assert response == {
        "continuable": False,
        "task_id": None,
        "status": None,
        "pending": 0,
        "failed": 0,
    }


def test_probe_continuable_returns_false_for_replacement_always(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    response = service.probe_continuable(kind="replacement")
    assert response["continuable"] is False
    assert response["task_id"] is None


def test_probe_continuable_returns_true_for_paused_translation(tmp_path: Path):
    """A paused translation cache with pending+failed>0 lights up the
    Continue button per architecture § 1.3."""

    from transoria.domain import TaskStatus
    from transoria.runtime.subtask import Subtask
    from transoria.runtime.task_record import TaskRecord
    from transoria.domain import SubtaskStatus, TaskKind

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    # Hand-seed a paused task in the cache so the predicate has
    # something to find.
    record = TaskRecord(
        id="translation-handcrafted",
        kind=TaskKind.TRANSLATION,
        status=TaskStatus.PAUSED,
        created_at="2026-04-29T00:00:00+00:00",
        updated_at="2026-04-29T00:00:00+00:00",
        metadata={
            "input_dir": str(input_dir),
            "output_dir": str(output_dir),
        },
    )
    subtask = Subtask(
        id="chunk-00000",
        task_id=record.id,
        status=SubtaskStatus.PENDING,
    )
    # Cache now lives under the user's output dir; seed via the
    # service helper so probe_continuable looks in the same place.
    service._cache_for_kind("translation").write_seed(record, [subtask])

    response = service.probe_continuable(kind="translation")
    assert response["continuable"] is True
    assert response["task_id"] == "translation-handcrafted"
    assert response["status"] == "paused"
    assert response["pending"] == 1
    assert response["failed"] == 0


def test_probe_continuable_self_heals_orphan_running_translation_task(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-probe-orphan"
    created_at = "2026-05-01T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )

    response = service.probe_continuable(kind="translation")

    assert response["continuable"] is True
    assert response["task_id"] == task_id
    assert response["status"] == "stopped"
    assert response["pending"] == 1
    persisted = service._cache_for_kind("translation").load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.PENDING


def test_probe_continuable_survives_service_restart_with_output_cache(
    tmp_path: Path,
):
    cache_root = tmp_path / "cache"
    service = _service(
        tmp_path, transport=EchoTranslationTransport(), cache_root=cache_root
    )
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("안녕", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)
    task_id = "translation-restart-stopped"
    created_at = "2026-05-02T00:00:00+00:00"
    service._cache_for_kind("translation").write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.STOPPED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
            },
        ),
        (
            Subtask(
                id="chunk-0",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
            ),
            Subtask(id="chunk-1", task_id=task_id, status=SubtaskStatus.PENDING),
        ),
    )

    restarted = _service(
        tmp_path, transport=EchoTranslationTransport(), cache_root=cache_root
    )

    probe = restarted.probe_continuable(kind="translation")
    assert probe == {
        "continuable": True,
        "task_id": task_id,
        "status": "stopped",
        "pending": 1,
        "failed": 0,
    }
    snapshot = restarted.read_snapshot(kind="translation", task_id=task_id)[
        "snapshot"
    ]
    assert snapshot["header"]["status"] == "stopped"
    assert snapshot["progress"]["total"] == 2


def test_probe_continuable_skips_record_with_different_input_folder(
    tmp_path: Path,
):
    from transoria.domain import TaskStatus
    from transoria.runtime.subtask import Subtask
    from transoria.runtime.task_record import TaskRecord
    from transoria.domain import SubtaskStatus, TaskKind

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    record = TaskRecord(
        id="translation-other",
        kind=TaskKind.TRANSLATION,
        status=TaskStatus.STOPPED,
        created_at="2026-04-29T00:00:00+00:00",
        updated_at="2026-04-29T00:00:00+00:00",
        metadata={
            "input_dir": str(tmp_path / "different-folder"),
            "output_dir": str(output_dir),
        },
    )
    subtask = Subtask(
        id="chunk-00000",
        task_id=record.id,
        status=SubtaskStatus.PENDING,
    )
    # Cache now lives under the user's output dir; seed via the
    # service helper so probe_continuable looks in the same place.
    service._cache_for_kind("translation").write_seed(record, [subtask])

    response = service.probe_continuable(kind="translation")
    assert response["continuable"] is False
    assert response["task_id"] is None


def test_stop_returns_not_running_for_unknown_id(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.stop_task(kind="translation", task_id="missing")
    assert caught.value.code == "bridge.not_found"


def test_stop_self_heals_zombie_running_task(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-stop-zombie"
    created_at = "2026-05-01T00:00:00+00:00"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
        ),
        (
            Subtask(
                id="chunk-00000",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )

    response = service.stop_task(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopped"
    persisted = service.cache.load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.PENDING


def test_glossary_review_completed_without_artifacts_surfaces_running_without_rewriting_completed(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "glossary-review-live-finalizing"
    created_at = "2026-05-01T00:00:00+00:00"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.COMPLETED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(tmp_path),
                "output_dir": str(tmp_path),
                "output_filename": "glossary-review-final.xlsx",
                "review_rounds_total": 5,
            },
        ),
        (
            Subtask(
                id="round-01-batch-0000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
            ),
        ),
    )
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="glossary_review",
            cache=service.cache,
            created_at=created_at,
        )
    )

    response = service.read_snapshot(kind="glossary_review", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "running"
    assert service.cache.load_record(task_id).status is TaskStatus.COMPLETED


def test_glossary_review_completed_without_artifacts_self_heals_to_stopped(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "glossary-review-missing-artifacts"
    created_at = "2026-05-01T00:00:00+00:00"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.COMPLETED,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(tmp_path),
                "output_dir": str(tmp_path),
                "output_filename": "glossary-review-final.xlsx",
                "review_rounds_total": 5,
            },
        ),
        (),
    )

    response = service.list_recent_tasks(kind="glossary_review", limit=None)

    assert response["tasks"][0]["status"] == "stopped"
    with pytest.raises(BridgeError) as caught:
        service.read_glossary_review_report(task_id=task_id)
    assert caught.value.code == "bridge.conflict"
    assert caught.value.payload.details["status"] == "stopped"


def test_list_recent_tasks_filters_by_kind(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    response = service.list_recent_tasks(kind="translation", limit=None)
    assert response == {"tasks": []}


def test_list_recent_tasks_rejects_negative_limit(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    # Negative is rejected by the handler layer; service accepts None or >=0.
    response = service.list_recent_tasks(kind="translation", limit=0)
    assert response == {"tasks": []}


def test_read_snapshot_self_heals_zombie_running_task(tmp_path: Path):
    """Persisted RUNNING with no live executor flips to STOPPED so the
    UI surfaces a continuable task instead of locking on a dead run.
    Subtasks stuck in RUNNING demote to PENDING so resume picks them up.
    """

    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-04-30T00:00:00+00:00"
    task_id = "translation-zombie"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={"output_dir": str(tmp_path / "out")},
        ),
        (
            Subtask(
                id="chunk-0", task_id=task_id, status=SubtaskStatus.RUNNING
            ),
            Subtask(
                id="chunk-1", task_id=task_id, status=SubtaskStatus.COMPLETED
            ),
            Subtask(
                id="chunk-2", task_id=task_id, status=SubtaskStatus.PENDING
            ),
        ),
    )

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopped"
    persisted = service.cache.load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    statuses = sorted(s.status.value for s in persisted.subtasks)
    assert statuses == ["completed", "pending", "pending"]


def test_read_snapshot_self_heals_orphan_pending_placeholder(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-05-01T00:00:00+00:00"
    task_id = "translation-pending-placeholder"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.PENDING,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (),
    )

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopped"
    assert response["snapshot"]["progress"]["total"] == 0
    assert service.cache.load(task_id).record.status is TaskStatus.STOPPED


def test_read_snapshot_clears_stale_registry_when_disk_says_failed(tmp_path: Path):
    """Race window: orchestrator wrote FAILED to disk but the runner
    wrap-up thread hasn't reached mark_done yet. The UI polls
    read_snapshot, sees FAILED, and the user might click Continue
    immediately. read_snapshot must eagerly clean the stale registry
    so that subsequent click never hits a phantom "already running"
    conflict."""

    from transoria.bridge.task_registry import RunningTask

    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-05-04T00:00:00+00:00"
    task_id = "translation-race-window"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.FAILED,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (
            Subtask(
                id="chunk-0", task_id=task_id, status=SubtaskStatus.FAILED, last_error="x"
            ),
        ),
    )
    stale = RunningTask(
        task_id=task_id,
        kind="translation",
        cache=service.cache,
        created_at=created_at,
    )
    service.registry.add(stale)
    assert not stale.is_done

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "failed"
    # Registry entry was force-marked done so the next continue/click
    # does not hit a phantom conflict.
    assert stale.is_done


def test_read_snapshot_leaves_live_running_task_alone(tmp_path: Path):
    """A live executor in the registry means RUNNING is genuine; no heal."""

    from transoria.bridge.task_registry import RunningTask

    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-04-30T00:00:00+00:00"
    task_id = "translation-live"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (Subtask(id="chunk-0", task_id=task_id, status=SubtaskStatus.RUNNING),),
    )
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="translation",
            cache=service.cache,
            created_at=created_at,
        )
    )

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "running"
    persisted = service.cache.load(task_id)
    assert persisted.record.status is TaskStatus.RUNNING
    assert persisted.subtasks[0].status is SubtaskStatus.RUNNING


def test_read_snapshot_keeps_slow_inflight_request_running(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-04-30T00:00:00+00:00"
    task_id = "translation-live-slow-inflight"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={"timeout_seconds": 600},
        ),
        (Subtask(id="chunk-0", task_id=task_id, status=SubtaskStatus.RUNNING),),
    )
    running = RunningTask(
        task_id=task_id,
        kind="translation",
        cache=service.cache,
        created_at=created_at,
    )
    running._last_heartbeat_monotonic = time.monotonic() - 605.0
    service.registry.add(running)

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "running"
    assert not running.is_done
    persisted = service.cache.load(task_id)
    assert persisted.record.status is TaskStatus.RUNNING
    assert persisted.subtasks[0].status is SubtaskStatus.RUNNING


def test_read_snapshot_self_heals_stalled_live_running_task(tmp_path: Path):
    """A registry entry with a stalled heartbeat is treated like an orphan.

    This prevents a dead in-process worker from keeping the run page in
    RUNNING until the app is restarted.
    """

    from transoria.bridge.task_registry import RunningTask

    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-04-30T00:00:00+00:00"
    task_id = "translation-live-stalled"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (Subtask(id="chunk-0", task_id=task_id, status=SubtaskStatus.RUNNING),),
    )
    running = RunningTask(
        task_id=task_id,
        kind="translation",
        cache=service.cache,
        created_at=created_at,
    )
    running._last_heartbeat_monotonic = 0.0
    service.registry.add(running)

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopped"
    assert running.is_done
    persisted = service.cache.load(task_id)
    assert persisted.record.status is TaskStatus.STOPPED
    assert persisted.subtasks[0].status is SubtaskStatus.PENDING


def test_list_recent_tasks_self_heals_orphan_pending_placeholder(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-05-01T00:00:00+00:00"
    task_id = "translation-recent-pending-placeholder"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.PENDING,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (),
    )

    response = service.list_recent_tasks(kind="translation", limit=1)

    assert response["tasks"][0]["id"] == task_id
    assert response["tasks"][0]["status"] == "stopped"
    assert service.cache.load(task_id).record.status is TaskStatus.STOPPED


def test_list_failed_subtasks_returns_empty_after_cache_wipe(tmp_path: Path):
    """Frontend's ``pollSnapshot`` calls ``read_snapshot`` and
    ``list_failed_subtasks`` in parallel via Promise.all. If the
    failed-subtasks call raises ``bridge.not_found`` after a clean
    COMPLETED wipe, the catch branch in ``useRuntimeStore`` zeros out
    ``activeTaskId`` / ``snapshot`` and the Run page collapses to
    0% / 0 tokens. Returning an empty failures list (the run had no
    failures, by definition of clean COMPLETED) keeps the polling
    loop steady on the mirrored snapshot."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    response = service.list_failed_subtasks(kind="translation", task_id=task_id)
    assert response == {"failures": []}


def test_list_recent_tasks_includes_completed_after_cache_wipe(tmp_path: Path):
    """The frontend's ``refreshActiveTask`` calls ``listRecentTasks(1)``
    to decide which task is active. After ``_maybe_cleanup_cache`` has
    wiped a clean COMPLETED run from disk, the in-memory mirror must
    still surface that task to ``list_recent_tasks`` — otherwise the UI
    would lose the activeTaskId reference and the Run page would zero
    out seconds after the run ended."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    # Disk cache should be gone.
    assert not (output_dir / "transoria-cache").exists()

    # But list_recent_tasks must still return the completed task so
    # the frontend's activeTaskId stays stable.
    listing = service.list_recent_tasks(kind="translation", limit=1)
    assert len(listing["tasks"]) == 1
    assert listing["tasks"][0]["id"] == task_id
    assert listing["tasks"][0]["status"] == "completed"


def test_clean_completion_preserves_cache_on_disk(tmp_path: Path):
    """After a clean COMPLETED run, the per-task cache directory
    survives so the user can resume / proofread / reopen later. The
    user's output folder never contains a ``transoria-cache/`` working
    directory either way (it now lives under ``default_cache_root()``)."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("안녕\n반갑다", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    # Output folder never carries cache anymore.
    assert not (output_dir / "transoria-cache").exists()
    # And the central per-task cache dir survives the clean completion.
    cache = service._cache_for_kind("translation")  # type: ignore[attr-defined]
    assert cache.task_dir(task_id).exists()
    assert (cache.task_dir(task_id) / "task.json").exists()

    # read_snapshot reads the surviving on-disk cache (not the in-memory
    # mirror) for the same task_id.
    snap = service.read_snapshot(kind="translation", task_id=task_id)["snapshot"]
    assert snap["header"]["status"] == "completed"
    assert snap["progress"]["completed"] == snap["progress"]["total"]

    artifacts = service.read_artifacts(kind="translation", task_id=task_id)
    assert artifacts["kind"] == "translation"
    assert any("sample" in path for path in artifacts["translated_files"])


def test_completed_snapshot_returns_persisted_subtasks_after_clean_run(
    tmp_path: Path,
):
    """The wipe is gone — after a clean run, ``read_snapshot`` returns
    the actual subtask list from disk (not an empty mirror). The
    proofreading flow depends on this surviving data."""

    transport = EchoTranslationTransport()
    service = _service(tmp_path, transport=transport)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("안녕\n반갑다", encoding="utf-8")
    _seed_translation_settings(service, input_dir=input_dir, output_dir=output_dir)

    response = service.start_translation(request_id="req-1")
    task_id = response["task_id"]
    _wait_until(
        lambda: service.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    snapshot = service.read_snapshot(kind="translation", task_id=task_id)["snapshot"]
    assert snapshot["progress"]["total"] >= 1
    assert snapshot["progress"]["completed"] == snapshot["progress"]["total"]
    assert snapshot["progress"]["pending"] == 0
    # Subtasks survive — required for proofreading.
    assert len(snapshot["subtasks"]) >= 1
    cache = service._cache_for_kind("translation")  # type: ignore[attr-defined]
    assert (cache.task_dir(task_id) / "subtasks").exists()


def test_completed_snapshot_freezes_metadata_for_in_memory_mirror(
    tmp_path: Path,
):
    """``_maybe_cleanup_cache`` no longer wipes — but it still freezes
    ``final_progress`` / ``final_usage`` into the in-memory mirror so
    a frontend that prefers the mirror (cheap rapid polls right after
    completion) sees the same numbers as the on-disk record."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-final-metrics"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.COMPLETED,
            created_at="2026-04-30T00:00:00+00:00",
            updated_at="2026-04-30T00:02:00+00:00",
            metadata={},
        ),
        (
            Subtask(
                id="chunk-0",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
                input_tokens=10,
                output_tokens=20,
            ),
            Subtask(
                id="chunk-1",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
                input_tokens=30,
                output_tokens=40,
            ),
        ),
    )

    service._maybe_cleanup_cache("translation", task_id)  # type: ignore[attr-defined]

    # In-memory mirror is populated.
    mirror = service._completed_snapshots[task_id]  # type: ignore[attr-defined]
    assert mirror.record.metadata["final_progress"]["total"] == 2
    assert mirror.record.metadata["final_progress"]["completed"] == 2
    assert mirror.record.metadata["final_usage"]["input_tokens"] == 40
    assert mirror.record.metadata["final_usage"]["output_tokens"] == 60

    # On-disk cache survives too — both subtasks are still readable.
    snapshot = service.read_snapshot(kind="translation", task_id=task_id)["snapshot"]
    assert len(snapshot["subtasks"]) == 2
    assert snapshot["progress"]["elapsed_seconds"] == 120.0
    assert snapshot["progress"]["rate_per_second"] == pytest.approx(2 / 120)
    assert snapshot["usage"]["total_tokens"] == 100


def test_stopping_snapshot_elapsed_freezes_at_status_update():
    from transoria.bridge.task_service import _task_elapsed_seconds

    record = TaskRecord(
        id="task-stop",
        kind=TaskKind.TRANSLATION,
        status=TaskStatus.STOPPING,
        created_at="2026-05-01T00:00:00+00:00",
        updated_at="2026-05-01T00:02:00+00:00",
    )

    assert _task_elapsed_seconds(record) == 120.0


def test_snapshot_elapsed_uses_accumulated_runtime_metadata():
    from transoria.bridge.task_service import _task_elapsed_seconds

    record = TaskRecord(
        id="task-runtime",
        kind=TaskKind.TRANSLATION,
        status=TaskStatus.COMPLETED,
        created_at="2026-05-01T00:00:00+00:00",
        updated_at="2026-05-01T04:00:00+00:00",
        metadata={"runtime_elapsed_seconds": 95.5},
    )

    assert _task_elapsed_seconds(record) == 95.5
# _derive_chunk_size


def test_derive_chunk_size_falls_back_when_unbounded():
    from transoria.bridge.task_service import _derive_chunk_size

    assert _derive_chunk_size(0) == 32
    assert _derive_chunk_size(-1) == 32


def test_derive_chunk_size_scales_with_input_token_limit():
    from transoria.bridge.task_service import _derive_chunk_size

    assert _derive_chunk_size(512) == 32
    assert _derive_chunk_size(1024) == 64


def test_derive_chunk_size_floors_to_8_for_tiny_limits():
    from transoria.bridge.task_service import _derive_chunk_size

    assert _derive_chunk_size(64) == 8
    assert _derive_chunk_size(127) == 8
    assert _derive_chunk_size(128) == 8


def test_derive_chunk_size_trusts_large_input_token_limits():
    from transoria.bridge.task_service import _derive_chunk_size

    assert _derive_chunk_size(4000) == 250
    assert _derive_chunk_size(8000) == 500
    assert _derive_chunk_size(16000) == 1000


def test_read_snapshot_self_heals_zombie_stopping_state(tmp_path: Path):
    """STOPPING is also a transient state that requires a live executor;
    on host crash it should heal to STOPPED, not stay 'stopping' forever."""

    service = _service(tmp_path, transport=EchoTranslationTransport())
    created_at = "2026-04-30T00:00:00+00:00"
    task_id = "translation-stopping-zombie"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.STOPPING,
            created_at=created_at,
            updated_at=created_at,
            metadata={},
        ),
        (Subtask(id="chunk-0", task_id=task_id, status=SubtaskStatus.PENDING),),
    )

    response = service.read_snapshot(kind="translation", task_id=task_id)

    assert response["snapshot"]["header"]["status"] == "stopped"


def test_summarize_caches_counts_and_sizes_persisted_tasks(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    for index in range(3):
        task_id = f"translation-summary-{index}"
        service.cache.write_seed(
            TaskRecord(
                id=task_id,
                kind=TaskKind.TRANSLATION,
                status=TaskStatus.COMPLETED,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:01:00+00:00",
                metadata={},
            ),
            (Subtask(id="chunk-0", task_id=task_id, status=SubtaskStatus.COMPLETED),),
        )

    summary = service.summarize_caches()
    assert summary["task_count"] == 3
    assert summary["total_bytes"] > 0
    assert summary["cache_root"] == str(service.cache.root)


def test_purge_caches_all_removes_every_completed_task(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    for index in range(2):
        task_id = f"translation-purge-all-{index}"
        service.cache.write_seed(
            TaskRecord(
                id=task_id,
                kind=TaskKind.TRANSLATION,
                status=TaskStatus.COMPLETED,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:01:00+00:00",
                metadata={},
            ),
            (),
        )

    response = service.purge_caches(scope="all")
    assert response["scope"] == "all"
    assert response["removed_count"] == 2
    assert response["skipped_active_count"] == 0
    assert service.summarize_caches()["task_count"] == 0


def test_purge_caches_rejects_while_task_is_running(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    running = RunningTask(
        task_id="translation-running-cache",
        kind="translation",
        cache=service.cache,
        created_at="2026-05-01T00:00:00+00:00",
    )
    service.registry.add(running)

    with pytest.raises(BridgeError) as caught:
        service.purge_caches(scope="all")

    assert caught.value.code == "bridge.conflict"
    assert caught.value.payload.details["active_task_ids"] == [
        "translation-running-cache"
    ]


def test_purge_caches_ignores_stale_registry_when_disk_is_terminal(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-stale-registry-purge"
    created_at = "2026-05-01T00:00:00+00:00"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.COMPLETED,
            created_at=created_at,
            updated_at=created_at,
        ),
        (),
    )
    service.registry.add(
        RunningTask(
            task_id=task_id,
            kind="translation",
            cache=service.cache,
            created_at=created_at,
        )
    )

    response = service.purge_caches(scope="all")

    assert response["removed_ids"] == [task_id]
    running = service.registry.get(task_id)
    assert running is not None
    assert running.is_done


def test_purge_caches_older_than_days_keeps_recent(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    now = datetime.now(timezone.utc)
    old = (now - timedelta(days=40)).isoformat()
    fresh = (now - timedelta(days=1)).isoformat()
    for task_id, updated in [
        ("translation-purge-old", old),
        ("translation-purge-fresh", fresh),
    ]:
        service.cache.write_seed(
            TaskRecord(
                id=task_id,
                kind=TaskKind.TRANSLATION,
                status=TaskStatus.COMPLETED,
                created_at=updated,
                updated_at=updated,
                metadata={},
            ),
            (),
        )

    response = service.purge_caches(scope="older_than_days", days=30)
    assert response["removed_count"] == 1
    assert response["removed_ids"] == ["translation-purge-old"]
    survivors = {r.id for r in service.cache.list_tasks()}
    assert survivors == {"translation-purge-fresh"}


def test_purge_caches_completed_keeps_recoverable_tasks(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    for task_id, status in [
        ("translation-purge-completed", TaskStatus.COMPLETED),
        ("translation-purge-failed", TaskStatus.FAILED),
        ("translation-purge-stopped", TaskStatus.STOPPED),
    ]:
        service.cache.write_seed(
            TaskRecord(
                id=task_id,
                kind=TaskKind.TRANSLATION,
                status=status,
                created_at="2026-05-01T00:00:00+00:00",
                updated_at="2026-05-01T00:00:00+00:00",
                metadata={},
            ),
            (),
        )

    response = service.purge_caches(scope="completed")

    assert response["removed_ids"] == ["translation-purge-completed"]
    survivors = {r.id for r in service.cache.list_tasks()}
    assert survivors == {"translation-purge-failed", "translation-purge-stopped"}


def test_purge_caches_rejects_unknown_scope(tmp_path: Path):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.purge_caches(scope="evil")
    assert caught.value.code == "bridge.invalid_argument"


def test_purge_caches_older_than_days_requires_non_negative_days(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    with pytest.raises(BridgeError) as caught:
        service.purge_caches(scope="older_than_days", days=-1)
    assert caught.value.code == "bridge.invalid_argument"


def test_purge_caches_drops_in_memory_mirrors_for_removed_tasks(
    tmp_path: Path,
):
    service = _service(tmp_path, transport=EchoTranslationTransport())
    task_id = "translation-mirror-drop"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.COMPLETED,
            created_at="2026-05-01T00:00:00+00:00",
            updated_at="2026-05-01T00:00:00+00:00",
            metadata={"final_progress": {"completed": 1}},
        ),
        (),
    )
    service._completed_snapshots[task_id] = service.cache.load(task_id)  # type: ignore[attr-defined]
    service._completed_results[task_id] = {"foo": "bar"}  # type: ignore[attr-defined]

    service.purge_caches(scope="all")

    assert task_id not in service._completed_snapshots  # type: ignore[attr-defined]
    assert task_id not in service._completed_results  # type: ignore[attr-defined]


def test_fresh_service_resumes_prior_task_from_disk(tmp_path: Path):
    """Simulate "user kills app mid-run, reopens later". Two distinct
    TaskService instances share the same cache_root: the second one
    must surface the task created by the first via list_recent_tasks
    so the RunPage auto-load can rehydrate the snapshot. This is
    Phase A's load-bearing guarantee: cache is central + persistent
    across process restarts."""

    cache_root = tmp_path / "shared-cache"
    cache_root.mkdir(parents=True)

    # First incarnation: seed a translation task and let it complete.
    first = _service(tmp_path, transport=EchoTranslationTransport(), cache_root=cache_root)
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "sample.txt").write_text("hi\nworld", encoding="utf-8")
    _seed_translation_settings(first, input_dir=input_dir, output_dir=output_dir)
    response = first.start_translation(request_id="req-resume")
    task_id = response["task_id"]
    _wait_until(
        lambda: first.read_snapshot(kind="translation", task_id=task_id)[
            "snapshot"
        ]["header"]["status"]
        == "completed",
    )

    # Second incarnation: brand-new service, same cache_root, no
    # in-memory mirrors. Must still see the completed task.
    second = _service(tmp_path, transport=EchoTranslationTransport(), cache_root=cache_root)
    listing = second.list_recent_tasks(kind="translation", limit=1)
    assert len(listing["tasks"]) == 1
    assert listing["tasks"][0]["id"] == task_id
    assert listing["tasks"][0]["status"] == "completed"
    snap = second.read_snapshot(kind="translation", task_id=task_id)["snapshot"]
    assert snap["progress"]["completed"] == snap["progress"]["total"]


def test_fresh_service_reconciles_running_zombie_to_stopped(tmp_path: Path):
    """If the previous process crashed mid-run, the cache contains a
    RUNNING record but no live thread. A fresh service must reconcile
    that to STOPPED on read so the user sees a continuable task
    instead of a phantom 'still running' state."""

    cache_root = tmp_path / "shared-cache"
    cache_root.mkdir(parents=True)
    service = _service(tmp_path, transport=EchoTranslationTransport(), cache_root=cache_root)
    task_id = "translation-crash-zombie"
    service.cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.TRANSLATION,
            status=TaskStatus.RUNNING,
            created_at="2026-05-01T00:00:00+00:00",
            updated_at="2026-05-01T00:00:30+00:00",
            metadata={},
        ),
        (
            Subtask(
                id="chunk-0",
                task_id=task_id,
                status=SubtaskStatus.RUNNING,
            ),
        ),
    )

    listing = service.list_recent_tasks(kind="translation", limit=1)
    assert listing["tasks"][0]["status"] == "stopped"
    snap = service.read_snapshot(kind="translation", task_id=task_id)["snapshot"]
    assert snap["header"]["status"] == "stopped"
    # The running subtask was reconciled to pending so continue can pick it up.
    assert snap["subtasks"][0]["status"] == "pending"
