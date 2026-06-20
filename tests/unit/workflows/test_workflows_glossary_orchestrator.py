from __future__ import annotations

import asyncio
import json
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Mapping

import pytest
from openpyxl import load_workbook

from transoria.domain import Language, SubtaskStatus, TaskStatus
from transoria.llm import LlmClient, ModelConfig, ProviderFormat, ThinkingLevel
from transoria.llm.client import TransportResult
from transoria.prompts import PromptKind, default_preset
from transoria.runtime import Subtask, SubtaskResult, TaskCache
from transoria.workflows.glossary import (
    GLOSSARY_FILENAME_JSON,
    GLOSSARY_FILENAME_REFERENCES,
    GLOSSARY_FILENAME_XLSX,
    GLOSSARY_STATISTICS_FILENAME_JSON,
    GlossaryConfig,
    GlossaryOrchestrator,
)
from transoria.workflows.glossary.orchestrator import _default_runner_factory


@dataclass
class CandidateEmittingTransport:
    """Each request returns a fixed list of glossary candidates as JSONL."""

    candidates: tuple[tuple[str, str, str], ...] = (
        ("신해범", "申海范", "Male Name"),
        ("공이", "孔二", "Author"),
        ("흑룡", "黑龙", "Creature"),
    )
    forced_failures: tuple[int, ...] = ()
    requests: list[dict[str, object]] = field(default_factory=list)

    async def execute(
        self,
        url: str,
        headers: Mapping[str, str],
        payload: Mapping[str, object],
        timeout: float,
    ) -> TransportResult:
        index = len(self.requests)
        self.requests.append(dict(payload))
        if index in self.forced_failures:
            return TransportResult(500, {"error": "boom"})
        lines = [
            json.dumps({"src": src, "dst": dst, "type": info}, ensure_ascii=False)
            for src, dst, info in self.candidates
        ]
        body = {
            "choices": [
                {"message": {"role": "assistant", "content": "\n".join(lines)}}
            ],
            "usage": {"prompt_tokens": 25, "completion_tokens": 55},
        }
        return TransportResult(200, body)


def _model() -> ModelConfig:
    return ModelConfig(
        id="m",
        display_name="m",
        provider_format=ProviderFormat.OPENAI,
        base_url="https://example/api/v1/",
        model_id="model-x",
        api_keys=("key",),
        thinking_level=ThinkingLevel.OFF,
        concurrency_limit=2,
        rpm_limit=0,
    )


_TIMES = iter(range(120))


def _frozen_clock() -> str:
    return f"2026-04-27T00:00:{next(_TIMES, 119):02d}+00:00"


def _config(input_dir: Path, output_dir: Path) -> GlossaryConfig:
    return GlossaryConfig(
        input_dir=input_dir,
        output_dir=output_dir,
        source_language=Language.KOREAN,
        target_language=Language.CHINESE_SIMPLIFIED,
        model=_model(),
        prompt_preset=default_preset(PromptKind.GLOSSARY),
        chunk_char_limit=200,
        request_retry_attempts=0,
    )


def _new_orchestrator(transport, cache_root: Path) -> GlossaryOrchestrator:
    counter = iter(range(1000))

    def id_factory() -> str:
        return f"task-{next(counter):04d}"

    return GlossaryOrchestrator(
        cache=TaskCache(root=cache_root),
        client=LlmClient(transport=transport),
        clock=_frozen_clock,
        id_factory=id_factory,
    )


def test_default_glossary_runner_uses_config_request_retry_attempts(tmp_path: Path) -> None:
    config = replace(
        _config(input_dir=tmp_path / "in", output_dir=tmp_path / "out"),
        request_retry_attempts=4,
    )

    runner = _default_runner_factory(None, config)  # type: ignore[arg-type]

    assert runner.transport_retry_attempts == 4  # type: ignore[attr-defined]


def test_orchestrator_extracts_glossary_end_to_end_for_txt(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()

    (input_dir / "Novel.txt").write_text(
        "\n".join(
            [
                "신해범 entered the room.",
                "공이 followed silently.",
                "Then 흑룡 spoke.",
                "신해범 sat down.",
                "흑룡 nodded.",
                "신해범 was thinking.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    assert result.final_status is TaskStatus.COMPLETED
    assert len(result.glossary_outputs_per_file) == 1
    xlsx_path, json_path, references_path = result.glossary_outputs_per_file[0]
    assert xlsx_path.name == "Novel" + GLOSSARY_FILENAME_XLSX
    assert json_path.name == "Novel" + GLOSSARY_FILENAME_JSON
    assert references_path.name == "Novel" + GLOSSARY_FILENAME_REFERENCES

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    by_src = {row["src"]: row for row in payload}
    assert "신해범" in by_src
    assert by_src["신해범"]["dst"] == "申海范"
    # Frequency is segment-line count: 신해범 appears in 3 lines.
    assert by_src["신해범"]["frequency"] == 3
    assert by_src["흑룡"]["frequency"] == 2
    assert by_src["공이"]["frequency"] == 1

    workbook = load_workbook(xlsx_path)
    rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    assert rows[0] == ("src", "dst", "info", "regex", "frequency")
    assert {row[0] for row in rows[1:]} == {"신해범", "공이", "흑룡"}

    references_text = references_path.read_text(encoding="utf-8")
    assert "原文: 신해범" in references_text
    assert "참조" not in references_text  # ensure we use the design's labels
    assert "신해범 entered the room." in references_text


def test_orchestrator_writes_extraction_statistics_file(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Novel.txt").write_text(
        "\n".join(["신해범 line 1.", "신해범 line 2.", "공이 here."]) + "\n",
        encoding="utf-8",
    )

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    stats_path = result.statistics_path
    assert stats_path == tmp_path / "cache" / result.task_id / GLOSSARY_STATISTICS_FILENAME_JSON
    assert not (output_dir / GLOSSARY_STATISTICS_FILENAME_JSON).exists()
    stats = json.loads(stats_path.read_text(encoding="utf-8"))
    assert stats["candidate_count"] >= 3
    assert stats["final_entry_count"] >= 1
    assert stats["failed_subtasks"] == 0
    assert stats["input_tokens"] >= 25


def test_orchestrator_records_failed_subtasks_in_statistics(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Novel.txt").write_text("신해범\n", encoding="utf-8")

    transport = CandidateEmittingTransport(forced_failures=(0,))
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(
        orchestrator.run(
            GlossaryConfig(
                input_dir=input_dir,
                output_dir=output_dir,
                source_language=Language.KOREAN,
                target_language=Language.CHINESE_SIMPLIFIED,
                model=_model(),
                prompt_preset=default_preset(PromptKind.GLOSSARY),
                chunk_char_limit=200,
                request_retry_attempts=0,
            )
        )
    )

    assert result.final_status is TaskStatus.FAILED
    assert result.statistics.failed_subtasks == 1


def test_orchestrator_leaves_failed_glossary_chunk_for_manual_continue(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    source_text = "신해범 " + "x" * 800
    (input_dir / "Novel.txt").write_text(source_text, encoding="utf-8")

    transport = CandidateEmittingTransport(forced_failures=(0,))
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    result = asyncio.run(
        orchestrator.run(
            GlossaryConfig(
                input_dir=input_dir,
                output_dir=output_dir,
                source_language=Language.KOREAN,
                target_language=Language.CHINESE_SIMPLIFIED,
                model=_model(),
                prompt_preset=default_preset(PromptKind.GLOSSARY),
                chunk_char_limit=1000,
                request_retry_attempts=0,
            )
        )
    )

    snapshot = orchestrator.cache.load(result.task_id)
    subtasks = {subtask.id: subtask for subtask in snapshot.subtasks}

    assert result.final_status is TaskStatus.FAILED
    assert subtasks["chunk-00000"].status is SubtaskStatus.FAILED
    assert result.statistics.failed_subtasks == 1
    assert len(transport.requests) == 1


def test_orchestrator_finalizes_outputs_when_stop_hits_after_all_glossary_chunks(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Novel.txt").write_text("신해범 line\n신해범 again\n", encoding="utf-8")
    holder: dict[str, object] = {}

    class CompleteRunner:
        async def run(self, subtask: Subtask) -> SubtaskResult:
            return SubtaskResult(
                response_content=json.dumps(
                    {
                        "entries": [
                            {"src": "신해범", "dst": "申海范", "info": "男性角色"}
                        ],
                        "issues": [],
                    },
                    ensure_ascii=False,
                )
            )

    def request_stop(_event: object) -> None:
        executor = holder.get("executor")
        if executor is not None:
            executor.request_stop()  # type: ignore[attr-defined]

    orchestrator = GlossaryOrchestrator(
        cache=TaskCache(root=tmp_path / "cache"),
        client=LlmClient(transport=CandidateEmittingTransport()),
        runner_factory=lambda _client, _config: CompleteRunner(),
        progress=request_stop,
        on_executor_created=lambda executor: holder.__setitem__("executor", executor),
        id_factory=lambda: "glossary-stop-after-complete",
    )

    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    assert result.final_status is TaskStatus.COMPLETED
    assert len(result.glossary_outputs_per_file) == 1
    assert result.glossary_outputs_per_file[0].xlsx_path.exists()
    assert orchestrator.cache.load(result.task_id).record.status is TaskStatus.COMPLETED
    assert getattr(holder["executor"], "subtask_timeout_seconds") == 0.0


def test_orchestrator_emits_only_combined_artifacts_when_enabled(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()
    (input_dir / "Alpha.txt").write_text("신해범 line\n", encoding="utf-8")
    (input_dir / "Bravo.txt").write_text("공이 line\n", encoding="utf-8")
    stale = output_dir / f"Alpha{GLOSSARY_FILENAME_XLSX}"
    stale.write_bytes(b"old")

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    config = replace(_config(input_dir, output_dir), combine_folder_glossary=True)
    result = asyncio.run(orchestrator.run(config))

    assert result.glossary_outputs_per_file == ()
    assert result.combined_output is not None
    assert result.combined_output.novel_name == "in"
    assert result.combined_output.xlsx_path.name == f"in{GLOSSARY_FILENAME_XLSX}"
    assert not stale.exists()


def test_orchestrator_raises_typed_error_for_empty_input(tmp_path: Path) -> None:
    """Empty input no longer silently completes — it raises
    ``GlossaryEmptyInputError`` so ``_on_task_failure`` can record a
    specific reason in ``record.metadata.last_error`` instead of the
    user seeing "completed 0/0" with no explanation."""

    from transoria.workflows.glossary.orchestrator import (
        GlossaryEmptyInputError,
    )

    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    with pytest.raises(GlossaryEmptyInputError):
        asyncio.run(orchestrator.run(_config(input_dir, output_dir)))


def test_orchestrator_reports_source_language_mismatch(tmp_path: Path) -> None:
    from transoria.workflows.glossary.orchestrator import (
        GlossaryEmptyInputError,
    )

    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Novel.txt").write_text(
        "これは日本語の本文です。\n東京で事件が起きた。", encoding="utf-8"
    )

    transport = CandidateEmittingTransport()
    orchestrator = _new_orchestrator(transport, tmp_path / "cache")

    with pytest.raises(GlossaryEmptyInputError, match="未检测到配置源语言"):
        asyncio.run(orchestrator.run(_config(input_dir, output_dir)))


def test_orchestrator_writes_decode_issues_even_when_no_candidates(
    tmp_path: Path,
) -> None:
    """User-reported: when LLM responses parse to 0 candidates, the
    decode-issues file must still be emitted so users can see why the
    output is empty."""

    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    (input_dir / "Sample.txt").write_text(
        "신해범 walked in.\n공이 said hello.\n", encoding="utf-8"
    )

    @dataclass
    class GibberishTransport:
        requests: list[dict[str, object]] = field(default_factory=list)

        async def execute(
            self, url, headers, payload, timeout
        ) -> TransportResult:
            self.requests.append(dict(payload))
            body = {
                "choices": [
                    {
                        "message": {
                            "role": "assistant",
                            # Three malformed lines — the JSONLINE
                            # decoder rejects each with a decode issue.
                            "content": "not json\nstill not json\n{broken",
                        }
                    }
                ],
                "usage": {"prompt_tokens": 5, "completion_tokens": 3},
            }
            return TransportResult(200, body)

    orchestrator = _new_orchestrator(GibberishTransport(), tmp_path / "cache")
    result = asyncio.run(orchestrator.run(_config(input_dir, output_dir)))

    assert result.statistics.candidate_count == 0
    assert result.statistics.decode_issue_count > 0
    decode_files = list(output_dir.glob("*-decode-issues.txt"))
    assert len(decode_files) == 1
    body = decode_files[0].read_text(encoding="utf-8")
    assert "reason:" in body
