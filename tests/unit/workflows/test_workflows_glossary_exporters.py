from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from transoria.workflows.glossary import (
    GLOSSARY_FILENAME_JSON,
    GLOSSARY_FILENAME_REFERENCES,
    GLOSSARY_FILENAME_XLSX,
    GlossaryRecord,
    glossary_basename,
    write_glossary_artifacts,
)


def _record(
    src: str,
    dst: str,
    *,
    info: str = "",
    frequency: int = 1,
    references: tuple[str, ...] = (),
) -> GlossaryRecord:
    return GlossaryRecord(
        src=src,
        dst=dst,
        info=info,
        frequency=frequency,
        references=references,
    )


def test_glossary_basename_uses_source_stem() -> None:
    assert glossary_basename(Path("/in/Novel Name.epub")) == "Novel Name"
    assert glossary_basename(Path("/in/Lord of Mysteries.txt")) == "Lord of Mysteries"


def test_write_artifacts_emits_three_files_with_hyphenated_suffixes(tmp_path: Path) -> None:
    records = (
        _record("신해범", "申海范", info="Male Name", frequency=12),
        _record("공이", "孔二", info="Author", frequency=3),
    )

    xlsx_path, json_path, references_path = write_glossary_artifacts(
        records, tmp_path, source_path=Path("/in/Novel Name.txt")
    )

    assert xlsx_path == tmp_path / f"Novel Name{GLOSSARY_FILENAME_XLSX}"
    assert json_path == tmp_path / f"Novel Name{GLOSSARY_FILENAME_JSON}"
    assert references_path == tmp_path / f"Novel Name{GLOSSARY_FILENAME_REFERENCES}"
    for path in (xlsx_path, json_path, references_path):
        assert path.exists()


def test_xlsx_columns_match_design_doc(tmp_path: Path) -> None:
    records = (_record("신해범", "申海范", info="Male Name", frequency=8),)

    xlsx_path, *_ = write_glossary_artifacts(
        records, tmp_path, source_path=Path("/in/Sample.txt")
    )

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("src", "dst", "info", "regex", "frequency")
    assert rows[1] == ("신해범", "申海范", "Male Name", False, 8)


def test_json_content_matches_xlsx_content(tmp_path: Path) -> None:
    records = (
        _record("신해범", "申海范", info="Male Name", frequency=8),
        _record("공이", "孔二", frequency=3),
    )

    xlsx_path, json_path, _ = write_glossary_artifacts(
        records, tmp_path, source_path=Path("/in/Sample.txt")
    )

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload == [
        {
            "src": "신해범",
            "dst": "申海范",
            "info": "Male Name",
            "regex": False,
            "frequency": 8,
        },
        {
            "src": "공이",
            "dst": "孔二",
            "info": "",
            "regex": False,
            "frequency": 3,
        },
    ]

    workbook = load_workbook(xlsx_path)
    # openpyxl reports empty string cells as None on read-back; normalize for
    # the cross-format equality check.
    rows = [
        tuple(value if value is not None else "" for value in row)
        for row in workbook.active.iter_rows(values_only=True)
    ][1:]
    assert rows[0] == ("신해범", "申海范", "Male Name", False, 8)
    assert rows[1] == ("공이", "孔二", "", False, 3)


def test_references_text_uses_kg_chinese_labeled_blocks(tmp_path: Path) -> None:
    records = (
        _record(
            "신해범",
            "申海范",
            info="Male Name",
            frequency=2,
            references=(
                "신해범 walked in.",
                "Then 신해범 sat down.",
            ),
        ),
    )

    _, _, references_path = write_glossary_artifacts(
        records, tmp_path, source_path=Path("/in/Sample.txt")
    )

    content = references_path.read_text(encoding="utf-8")
    assert "原文: 신해범" in content
    assert "译文: 申海范" in content
    assert "备注: Male Name" in content
    assert "出现次数: 2" in content
    assert "参考文本: " in content
    assert "신해범 walked in." in content
    assert "Then 신해범 sat down." in content


def test_references_text_separates_blocks_with_blank_line(tmp_path: Path) -> None:
    records = (
        _record("신해범", "申海范", frequency=1, references=("신해범 line",)),
        _record("공이", "孔二", frequency=1, references=("공이 line",)),
    )

    _, _, references_path = write_glossary_artifacts(
        records, tmp_path, source_path=Path("/in/Sample.txt")
    )

    content = references_path.read_text(encoding="utf-8")
    assert "신해범 line\n\n原文: 공이" in content


def test_purge_glossary_artifacts_removes_files_for_input_basenames(
    tmp_path: Path,
) -> None:
    from transoria.workflows.glossary import purge_glossary_artifacts

    src = Path("/in/Sample.txt")
    write_glossary_artifacts(
        (_record("a", "b"),),
        tmp_path,
        source_path=src,
    )
    # Unrelated artifact must survive (different basename).
    (tmp_path / "Other-Glossary.xlsx").write_bytes(b"x")

    purge_glossary_artifacts(tmp_path, source_paths=(src,))

    base = glossary_basename(src)
    assert not (tmp_path / f"{base}{GLOSSARY_FILENAME_XLSX}").exists()
    assert not (tmp_path / f"{base}{GLOSSARY_FILENAME_JSON}").exists()
    assert not (tmp_path / f"{base}{GLOSSARY_FILENAME_REFERENCES}").exists()
    assert (tmp_path / "Other-Glossary.xlsx").exists()

