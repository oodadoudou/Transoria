from __future__ import annotations

import asyncio
import json
from dataclasses import replace

from openpyxl import Workbook, load_workbook

from transoria.llm.config import ModelConfig
from transoria.prompts import PromptPreset
from transoria.domain import Language, SubtaskStatus, TaskKind, TaskStatus
from transoria.runtime import Subtask, SubtaskResult, TaskCache
from transoria.runtime.task_record import TaskRecord
from transoria.workflows.glossary_review.config import GlossaryReviewConfig
from transoria.workflows.glossary_review.loader import GlossaryReviewRow
from transoria.workflows.glossary_review.orchestrator import GlossaryReviewOrchestrator
from transoria.workflows.glossary_review.orchestrator import _default_runner_factory
from transoria.workflows.glossary_review.runner import decode_review_response


def test_decode_review_response_ignores_keep_rows() -> None:
    decoded = decode_review_response(
        '{"decisions":[{"row_index":2,"action":"keep"},{"row_index":3,"action":"modify","suggested_dst":"甲","reason":"better"}]}'
    )

    assert len(decoded) == 2
    assert decoded[0].action == "keep"
    assert decoded[1].row_index == 3
    assert decoded[1].suggested_dst == "甲"


def test_round_payload_includes_tier_and_history(tmp_path) -> None:
    config = _config(tmp_path, novel_background="신해범 是核心人物")
    row = GlossaryReviewRow(
        row_index=2,
        src="신해범",
        dst="申海凡",
        info="男性角色",
        frequency=8,
        context="신해범 出现在正文。",
    )
    orchestrator = GlossaryReviewOrchestrator(cache=None, client=None)  # type: ignore[arg-type]

    subtasks = orchestrator._build_round_subtasks(
        "task",
        round_index=2,
        config=config,
        rows=(row,),
        history={row.src: ()},
    )

    payload_row = subtasks[0].request_payload["rows"][0]
    assert payload_row["tier"] == "S"
    assert "核心设定词" in payload_row["instruction"]
    assert payload_row["is_character"] is True
    assert payload_row["current_category"] == "男性角色"


def test_consensus_terms_are_skipped_from_third_round(tmp_path) -> None:
    config = _config(tmp_path)
    row = GlossaryReviewRow(
        row_index=2,
        src="응",
        dst="嗯",
        info="其它",
        frequency=20,
    )
    orchestrator = GlossaryReviewOrchestrator(cache=None, client=None)  # type: ignore[arg-type]

    subtasks = orchestrator._build_round_subtasks(
        "task",
        round_index=3,
        config=config,
        rows=(row,),
        history={
            row.src: (
                type("History", (), {"dst": "嗯", "info": "其它", "deleted": True})(),
                type("History", (), {"dst": "嗯", "info": "其它", "deleted": True})(),
            )
        },
    )

    assert subtasks == []


def test_review_finalizes_output_when_stop_hits_after_all_batches(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "申海凡", "男性角色", 5])
    workbook.save(tmp_path / "terms.xlsx")
    (tmp_path / "terms-references.txt").write_text("신해범 出现在正文。", encoding="utf-8")
    holder: dict[str, object] = {}

    class CompleteRunner:
        async def run(self, subtask: Subtask) -> SubtaskResult:
            row = subtask.request_payload["rows"][0]
            return SubtaskResult(
                response_content=json.dumps(
                    {
                        "decisions": [
                            {
                                "row_index": row["row_index"],
                                "action": "modify",
                                "suggested_dst": "申海范",
                                "suggested_info": "男性角色",
                                "reason": "统一人名",
                            }
                        ]
                    },
                    ensure_ascii=False,
                )
            )

    def request_stop(_event: object) -> None:
        executor = holder.get("executor")
        if executor is not None:
            executor.request_stop()  # type: ignore[attr-defined]

    orchestrator = GlossaryReviewOrchestrator(
        cache=TaskCache(root=tmp_path / "cache"),
        client=None,  # type: ignore[arg-type]
        runner_factory=lambda _client, _config: CompleteRunner(),
        progress=request_stop,
        on_executor_created=lambda executor: holder.__setitem__("executor", executor),
        id_factory=lambda: "review-stop-after-complete",
    )

    result = asyncio.run(
        orchestrator.run(
            GlossaryReviewConfig(
                **{
                    **_config(tmp_path).__dict__,
                    "review_rounds": 1,
                }
            )
        )
    )

    assert result.final_status is TaskStatus.COMPLETED
    assert result.output_path is not None
    assert result.report_path is not None
    assert result.output_path.exists()
    assert result.report_path.exists()
    rows = [tuple(row) for row in load_workbook(result.output_path).active.iter_rows(values_only=True)]
    assert rows[1][1] == "申海范"
    assert orchestrator.cache.load(result.task_id).record.status is TaskStatus.COMPLETED
    assert getattr(holder["executor"], "subtask_timeout_seconds") == 0.0


def test_review_marks_completed_when_all_rounds_were_already_done(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["신해범", "申海凡", "男性角色", 5])
    workbook.save(tmp_path / "terms.xlsx")
    (tmp_path / "terms-references.txt").write_text("신해범 出现在正文。", encoding="utf-8")
    cache = TaskCache(root=tmp_path / "cache")
    task_id = "review-already-done"
    created_at = "2026-05-01T00:00:00+00:00"
    cache.write_seed(
        TaskRecord(
            id=task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            status=TaskStatus.RUNNING,
            created_at=created_at,
            updated_at=created_at,
            metadata={
                "input_dir": str(tmp_path),
                "output_dir": str(tmp_path),
                "output_filename": "reviewed.xlsx",
                "input_xlsx": str(tmp_path / "terms.xlsx"),
                "reference_files": [str(tmp_path / "terms-references.txt")],
                "review_rounds_total": 1,
                "review_round_current": 1,
                "review_round_completed": 1,
                "review_round_total_batches": 1,
                "review_round_completed_batches": 1,
            },
        ),
        (
            Subtask(
                id="round-01-batch-0000",
                task_id=task_id,
                status=SubtaskStatus.COMPLETED,
                request_payload={
                    "round": 1,
                    "batch": 0,
                    "rows": [
                        {
                            "row_index": 2,
                            "src": "신해범",
                            "dst": "申海凡",
                            "info": "男性角色",
                            "frequency": 5,
                        }
                    ],
                },
                response_content=json.dumps(
                    {
                        "decisions": [
                            {
                                "row_index": 2,
                                "action": "modify",
                                "suggested_dst": "申海范",
                                "suggested_info": "男性角色",
                                "reason": "统一人名",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
            ),
        ),
    )

    class KeepRunner:
        async def run(self, subtask: Subtask) -> SubtaskResult:
            return SubtaskResult(
                response_content=json.dumps(
                    {
                        "decisions": [
                            {"row_index": row["row_index"], "action": "keep"}
                            for row in subtask.request_payload["rows"]
                        ]
                    },
                    ensure_ascii=False,
                )
            )

    orchestrator = GlossaryReviewOrchestrator(
        cache=cache,
        client=None,  # type: ignore[arg-type]
        runner_factory=lambda _client, _config: KeepRunner(),
        id_factory=lambda: task_id,
    )

    result = asyncio.run(
        orchestrator.run(
            GlossaryReviewConfig(
                **{
                    **_config(tmp_path).__dict__,
                    "review_rounds": 1,
                }
            )
        )
    )

    assert result.final_status is TaskStatus.COMPLETED
    assert result.output_path is not None
    assert result.report_path is not None
    assert result.output_path.exists()
    assert result.report_path.exists()
    assert cache.load(task_id).record.status is TaskStatus.COMPLETED


def test_final_character_consistency_round_updates_name_and_gender_category(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["하얀이", "李夏彦", "女性角色", 12])
    sheet.append(["하얀", "李夏言", "男性角色", 9])
    sheet.append(["센터", "中心", "命名组织", 3])
    workbook.save(tmp_path / "terms.xlsx")
    (tmp_path / "terms-references.txt").write_text(
        "하얀이 和 하얀 是同一角色的称呼。", encoding="utf-8"
    )

    class ConsistencyRunner:
        async def run(self, subtask: Subtask) -> SubtaskResult:
            if subtask.request_payload.get("mode") == "character_consistency":
                return SubtaskResult(
                    response_content=json.dumps(
                        {
                            "decisions": [
                                {
                                    "row_index": 2,
                                    "action": "modify_category",
                                    "suggested_dst": "李夏言",
                                    "suggested_info": "男性角色",
                                    "reason": "同一角色译名和性别分类统一",
                                }
                            ]
                        },
                        ensure_ascii=False,
                    )
                )
            return SubtaskResult(
                response_content=json.dumps(
                    {
                        "decisions": [
                            {"row_index": row["row_index"], "action": "keep"}
                            for row in subtask.request_payload["rows"]
                        ]
                    },
                    ensure_ascii=False,
                )
            )

    cache = TaskCache(root=tmp_path / "cache")
    orchestrator = GlossaryReviewOrchestrator(
        cache=cache,
        client=None,  # type: ignore[arg-type]
        runner_factory=lambda _client, _config: ConsistencyRunner(),
        id_factory=lambda: "review-character-consistency",
    )

    result = asyncio.run(
        orchestrator.run(
            GlossaryReviewConfig(
                **{
                    **_config(tmp_path).__dict__,
                    "review_rounds": 1,
                }
            )
        )
    )

    assert result.final_status is TaskStatus.COMPLETED
    assert result.output_path is not None
    assert result.report_path is not None
    rows = [
        tuple(row)
        for row in load_workbook(result.output_path).active.iter_rows(values_only=True)
    ]
    assert rows[1][1] == "李夏言"
    assert rows[1][2] == "男性角色"
    report = json.loads(result.report_path.read_text(encoding="utf-8"))
    assert report["rows"][0]["action"] == "name_consistency"
    assert report["rows"][0]["suggested_info"] == "男性角色"
    metadata = cache.load(result.task_id).record.metadata
    assert metadata["review_rounds_total"] == 2
    consistency_subtasks = [
        subtask
        for subtask in cache.load_subtasks(result.task_id)
        if subtask.request_payload.get("mode") == "character_consistency"
    ]
    assert len(consistency_subtasks) == 1
    assert [
        row["src"] for row in consistency_subtasks[0].request_payload["rows"]
    ] == ["하얀이", "하얀"]


def test_final_character_consistency_round_accepts_category_only_fix(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["src", "dst", "info", "frequency"])
    sheet.append(["선우", "善宇", "女性角色", 155])
    sheet.append(["남선우", "南善宇", "男性角色", 5])
    workbook.save(tmp_path / "terms.xlsx")
    (tmp_path / "terms-references.txt").write_text(
        "선우 是 남선우 的简称。", encoding="utf-8"
    )

    class ConsistencyRunner:
        async def run(self, subtask: Subtask) -> SubtaskResult:
            if subtask.request_payload.get("mode") == "character_consistency":
                return SubtaskResult(
                    response_content=json.dumps(
                        {
                            "decisions": [
                                {
                                    "row_index": 2,
                                    "action": "category",
                                    "suggested_info": "男性角色",
                                    "reason": "简称与全名指向同一男性角色",
                                }
                            ]
                        },
                        ensure_ascii=False,
                    )
                )
            return SubtaskResult(
                response_content=json.dumps(
                    {
                        "decisions": [
                            {"row_index": row["row_index"], "action": "keep"}
                            for row in subtask.request_payload["rows"]
                        ]
                    },
                    ensure_ascii=False,
                )
            )

    cache = TaskCache(root=tmp_path / "cache")
    orchestrator = GlossaryReviewOrchestrator(
        cache=cache,
        client=None,  # type: ignore[arg-type]
        runner_factory=lambda _client, _config: ConsistencyRunner(),
        id_factory=lambda: "review-character-gender-consistency",
    )

    result = asyncio.run(
        orchestrator.run(
            GlossaryReviewConfig(
                **{
                    **_config(tmp_path).__dict__,
                    "review_rounds": 1,
                }
            )
        )
    )

    assert result.final_status is TaskStatus.COMPLETED
    assert result.output_path is not None
    rows = [
        tuple(row)
        for row in load_workbook(result.output_path).active.iter_rows(values_only=True)
    ]
    assert rows[1][1] == "善宇"
    assert rows[1][2] == "男性角色"
    assert result.report_path is not None
    report = json.loads(result.report_path.read_text(encoding="utf-8"))
    assert report["rows"][0]["action"] == "name_consistency"
    assert report["rows"][0]["suggested_info"] == "男性角色"


def test_default_review_runner_uses_config_retry_attempts(tmp_path) -> None:
    config = replace(_config(tmp_path), retry_attempts=4)

    runner = _default_runner_factory(None, config)  # type: ignore[arg-type]

    assert runner.transport_retry_attempts == 4  # type: ignore[attr-defined]


def _config(tmp_path, *, novel_background: str = "") -> GlossaryReviewConfig:
    return GlossaryReviewConfig(
        input_dir=tmp_path,
        selected_xlsx_path=None,
        selected_reference_paths=(),
        source_language=Language.KOREAN,
        target_language=Language.CHINESE_SIMPLIFIED,
        output_filename="reviewed.xlsx",
        novel_background=novel_background,
        review_rounds=3,
        batch_size=20,
        retry_attempts=3,
        model=ModelConfig(
            id="m",
            display_name="m",
            provider_format="custom",
            base_url="http://localhost/v1",
            model_id="model",
            api_keys=("key",),
        ),
        prompt_preset=PromptPreset(
            id="p",
            kind="glossary_review",
            name="p",
            description="",
            enabled=True,
            is_system=False,
            system_prompt="review",
        ),
    )
