"""Bridge-facing orchestration of long-running tasks.

This module is the single owner of:

- ``TaskRegistry`` — in-memory map of in-flight tasks.
- ``TaskCache`` — file-backed task records under ``<cache_root>/tasks/``.
- Translation / Glossary / Replacement launches: validate settings, build the
  workflow config, run the async orchestrator (or sync replacement loop) on a
  background thread, and persist artifact metadata for ``read_artifacts``.

The bridge handlers are thin shells around :class:`TaskService` so the bridge
contract surface stays declarative and the runtime wiring lives in one place.
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import os
import re
import threading
import time
import traceback
import unicodedata
from collections import Counter
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Callable, Mapping, Sequence
from uuid import uuid4

from openpyxl import load_workbook

from transoria.bridge.errors import BridgeError
from transoria.bridge.task_registry import RunningTask, TaskRegistry
from transoria.domain import (
    DocumentFormat,
    Language,
    SubtaskStatus,
    TaskKind,
    TaskStatus,
)
from transoria.formats.scanner import scan_input_directory
from transoria.llm.client import ChatRequest, HttpxChatTransport, LlmClient
from transoria.llm.config import ModelConfig, effective_concurrency_limit
from transoria.model_profiles import ModelProfileStore
from transoria.prompts import (
    PromptKind,
    PromptPreset,
    PromptPresetStore,
)
from transoria.runtime.cache import TaskCache, TaskNotFoundError
from transoria.runtime.executor import TaskExecutor
from transoria.runtime.request_log import request_log_scope
from transoria.runtime.subtask import Subtask
from transoria.runtime.task_record import TaskRecord, TaskSnapshot
from transoria.runtime.task_timing import (
    elapsed_seconds_for_record,
    runtime_metadata_for_status,
)
from transoria.settings import SettingsStore
from transoria.tools.replacement import (
    REPLACED_SUFFIX,
    ReplacementRule,
    replace_epub_file,
    replace_txt_file,
)
from transoria.tools.epub_compressor import (
    EpubCompressAction,
    EpubCompressOptions,
    build_epub_compress_plan,
    build_epub_compress_report,
    compress_epub_file,
)
from transoria.tools.epub_merger import (
    EpubMergeAction,
    EpubMergeOptions,
    EpubMergeResult,
    build_epub_merge_plan,
    build_epub_merge_report,
    merge_epub_files,
)
from transoria.tools.epub_converter import (
    EpubConvertAction,
    EpubConvertOptions,
    build_epub_convert_plan,
    build_epub_convert_report,
    convert_epub_to_txt,
)
from transoria.tools.txt_to_epub import (
    TxtToEpubAction,
    TxtToEpubOptions,
    build_txt_to_epub_plan,
    build_txt_to_epub_report,
    convert_txt_to_epub,
)
from transoria.workflows.glossary.config import GlossaryConfig
from transoria.workflows.glossary_review.config import GlossaryReviewConfig
from transoria.workflows.glossary_review.exporters import REPORT_FILENAME
from transoria.workflows.glossary_review.loader import (
    discover_review_input_candidates,
    load_glossary_xlsx,
    normalize_output_filename,
)
from transoria.workflows.glossary_review.orchestrator import (
    GlossaryReviewOrchestrator,
    GlossaryReviewResult,
)
from transoria.workflows.translation.confidence import evaluate_segment_confidence
from transoria.workflows.translation.rules import (
    Glossary,
    ReplacementRule as TranslationReplacementRule,
    TextPreserveRule,
)
from transoria.workflows.glossary.exporters import (
    GLOSSARY_FILENAME_DECODE_ISSUES,
    GLOSSARY_FILENAME_JSON,
    GLOSSARY_FILENAME_REFERENCES,
    GLOSSARY_FILENAME_XLSX,
)
from transoria.workflows.glossary.orchestrator import (
    GlossaryArtifactSet,
    GlossaryExtractionResult,
    GlossaryOrchestrator,
)
from transoria.workflows.glossary.statistics import GLOSSARY_STATISTICS_FILENAME_JSON
from transoria.workflows.translation.config import TranslationConfig
from transoria.workflows.translation.glossary_report import (
    GLOSSARY_REPORT_FILENAME_JSON,
    GLOSSARY_REPORT_FILENAME_MD,
)
from transoria.workflows.translation.orchestrator import (
    TranslationOrchestrator,
    TranslationRunResult,
)
from transoria.workflows.translation.segment_state import (
    collect_segment_state_from_authoritative_subtasks,
    low_confidence_by_segment,
    mark_accepted_override,
)
from transoria.workflows.translation.statistics import STATISTICS_FILENAME_JSON
from transoria.utils.paths import describe_os_error, normalize_path_key

LlmClientFactory = Callable[[], LlmClient]


_RESULT_FILENAME = "result.json"
_REPLACEMENT_REPORT_FILENAME = "replacement-report.json"
_EPUB_COMPRESS_REPORT_FILENAME = "epub-compress-report.json"
_EPUB_MERGE_REPORT_FILENAME = "epub-merge-report.json"
_EPUB_CONVERT_REPORT_FILENAME = "epub-convert-report.json"
_TXT_TO_EPUB_REPORT_FILENAME = "txt-to-epub-report.json"
_DEBUG_LOG_FILENAME_SAFE = re.compile(r"[^A-Za-z0-9._-]")
_REQUEST_LOG_RESPONSE_PREVIEW_LIMIT = 20000
_REQUEST_LOG_ERROR_PREVIEW_LIMIT = 2000
# Hard cap on occurrences captured per rule across the whole task —
# the per-file cap inside ``apply_rules`` already prevents pathological
# files; this guards the aggregated report so a 100k-match rule cannot
# blow up the JSON we ship to the frontend.
_REPORT_MAX_OCCURRENCES_PER_RULE = 200
_KIND_TO_TASKKIND: dict[str, TaskKind] = {
    "translation": TaskKind.TRANSLATION,
    "glossary": TaskKind.GLOSSARY,
    "glossary_review": TaskKind.GLOSSARY_REVIEW,
    "replacement": TaskKind.REPLACEMENT,
    "epub_compress": TaskKind.EPUB_COMPRESS,
    "epub_merge": TaskKind.EPUB_MERGE,
    "epub_convert": TaskKind.EPUB_CONVERT,
    "txt_to_epub": TaskKind.TXT_TO_EPUB,
}

# Persisted statuses that imply a live executor must exist; if the
# in-memory registry is empty for one of these, the host crashed during
# start or mid-run.
_ZOMBIE_TASK_STATES: frozenset[TaskStatus] = frozenset(
    {
        TaskStatus.PENDING,
        TaskStatus.RUNNING,
        TaskStatus.STOPPING,
        TaskStatus.PAUSING,
    }
)
# Disk states that imply the executor finished (success or failure).
# A registry entry that's still not-done while disk shows one of these
# is stale — usually a thread that exited without reaching mark_done.
_TERMINAL_TASK_STATES: frozenset[TaskStatus] = frozenset(
    {
        TaskStatus.COMPLETED,
        TaskStatus.FAILED,
        TaskStatus.STOPPED,
    }
)
_STALE_RUNNING_SUBTASK_TASK_STATES: frozenset[TaskStatus] = frozenset(
    {
        TaskStatus.STOPPED,
        TaskStatus.PAUSED,
        TaskStatus.FAILED,
    }
)
_LIVE_TASK_STALL_SECONDS = 900.0
_STOP_REQUEST_STALL_SECONDS = 900.0
_LIVE_TASK_STALL_TIMEOUT_HEADROOM_SECONDS = 120.0
_MAX_RETRANSLATE_JOBS = 50
_RETRANSLATE_TERMINAL_TTL_SECONDS = 300.0
_RETRANSLATE_TERMINAL_STATUSES = frozenset(
    {"completed", "failed", "stale", "skipped", "unresolved"}
)
_RETRANSLATE_QUALITY_BATCH_MAX_ITEMS = 5
_RETRANSLATE_QUALITY_BATCH_WAIT_SECONDS = 0.2


def _metadata_timeout_seconds(metadata: Mapping[str, object]) -> float | None:
    raw = metadata.get("timeout_seconds")
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    return value


@dataclass
class RetranslateJob:
    request_id: str
    task_id: str
    segment_id: str
    original_dst: str
    model_id: str | None = None
    prompt_preset_id: str | None = None
    status: str = "pending"
    result_dst: str = ""
    error: str = ""
    created_at: float = 0.0
    source_hash: str = ""
    original_dst_hash: str = ""
    attempts: int = 0
    last_error: str = ""
    last_translation: str = ""
    seg_data: dict[str, object] | None = None
    metadata: dict[str, object] | None = None
    model_snapshot: dict[str, object] | None = None
    prompt_snapshot: dict[str, object] | None = None
    batch_items: list[dict[str, object]] | None = None
    batch_results: list[dict[str, object]] | None = None
    cache_applied: bool = False
    created_at_wall: str = ""
    updated_at_wall: str = ""
    schema_version: int = 1

    def to_dict(self) -> dict[str, object]:
        return {
            "schema_version": self.schema_version,
            "request_id": self.request_id,
            "task_id": self.task_id,
            "segment_id": self.segment_id,
            "original_dst": self.original_dst,
            "model_id": self.model_id,
            "prompt_preset_id": self.prompt_preset_id,
            "status": self.status,
            "result_dst": self.result_dst,
            "error": self.error,
            "created_at": self.created_at,
            "source_hash": self.source_hash,
            "original_dst_hash": self.original_dst_hash,
            "attempts": self.attempts,
            "last_error": self.last_error,
            "last_translation": self.last_translation,
            "seg_data": self.seg_data or {},
            "metadata": self.metadata or {},
            "model_snapshot": self.model_snapshot or {},
            "prompt_snapshot": self.prompt_snapshot or {},
            "batch_items": self.batch_items or [],
            "batch_results": self.batch_results or [],
            "cache_applied": self.cache_applied,
            "created_at_wall": self.created_at_wall,
            "updated_at_wall": self.updated_at_wall,
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, object]) -> "RetranslateJob":
        def optional_str(key: str) -> str | None:
            value = data.get(key)
            return None if value is None else str(value)

        def mapping_or_none(key: str) -> dict[str, object] | None:
            value = data.get(key)
            return dict(value) if isinstance(value, Mapping) else None

        def mappings(key: str) -> list[dict[str, object]] | None:
            value = data.get(key)
            if not isinstance(value, list):
                return None
            return [dict(item) for item in value if isinstance(item, Mapping)]

        created_at = data.get("created_at")
        try:
            created_at_float = float(created_at)
        except (TypeError, ValueError):
            created_at_float = time.monotonic()
        attempts = data.get("attempts")
        try:
            attempts_int = int(attempts)
        except (TypeError, ValueError):
            attempts_int = 0
        schema_version = data.get("schema_version", 1)
        try:
            schema_version_int = int(schema_version)
        except (TypeError, ValueError):
            schema_version_int = 1
        return cls(
            request_id=str(data["request_id"]),
            task_id=str(data["task_id"]),
            segment_id=str(data["segment_id"]),
            original_dst=str(data.get("original_dst", "")),
            model_id=optional_str("model_id"),
            prompt_preset_id=optional_str("prompt_preset_id"),
            status=str(data.get("status", "pending")),
            result_dst=str(data.get("result_dst", "")),
            error=str(data.get("error", "")),
            created_at=created_at_float,
            source_hash=str(data.get("source_hash", "")),
            original_dst_hash=str(data.get("original_dst_hash", "")),
            attempts=attempts_int,
            last_error=str(data.get("last_error", "")),
            last_translation=str(data.get("last_translation", "")),
            seg_data=mapping_or_none("seg_data"),
            metadata=mapping_or_none("metadata"),
            model_snapshot=mapping_or_none("model_snapshot"),
            prompt_snapshot=mapping_or_none("prompt_snapshot"),
            batch_items=mappings("batch_items"),
            batch_results=mappings("batch_results"),
            cache_applied=bool(data.get("cache_applied", False)),
            created_at_wall=str(data.get("created_at_wall", "")),
            updated_at_wall=str(data.get("updated_at_wall", "")),
            schema_version=schema_version_int,
        )


def _retranslate_job_segment_ids(job: RetranslateJob) -> set[str]:
    if job.batch_items:
        return {
            str(item.get("segment_id", ""))
            for item in job.batch_items
            if item.get("segment_id") not in (None, "")
        }
    return {job.segment_id}


def _retranslate_log_subtask_id(request_id: str) -> str:
    return f"proofreading-{request_id}"


@dataclass(frozen=True)
class _RetranslateRunnerResult:
    translations: dict[str, str]
    low_confidence: dict[str, dict[str, list[str]]]
    quality_review_segments: frozenset[str] = frozenset()


@dataclass(frozen=True)
class _RetranslateTaskIndex:
    task_updated_at: str
    owner_subtask_ids: dict[str, tuple[str, ...]]


@dataclass(frozen=True)
class _RetranslateQualityDecision:
    decision: str
    reason: str


@dataclass
class _RetranslateQualityReviewItem:
    source_text: str
    existing_dst: str
    candidate_dst: str
    source_language: str
    target_language: str
    model_id: str
    model_snapshot: dict[str, object] | None
    estimated_tokens: int
    result: _RetranslateQualityDecision | None = None
    error: Exception | None = None
    done: threading.Event = field(default_factory=threading.Event)


@dataclass
class _RetranslateQualityReviewGroup:
    items: list[_RetranslateQualityReviewItem] = field(default_factory=list)


_RETRANSLATE_QUALITY_DECISIONS = {"accept_new", "keep_existing", "uncertain"}
_RETRANSLATE_QUALITY_COMPARATOR_PROMPT = (
    "You are a conservative translation quality comparator. "
    "Do not translate, rewrite, or improve either candidate. Compare the "
    "existing translation and the new candidate against the source and "
    "choose whether the new candidate is safe to store. Exact preservation "
    "can be correct for symbols, emoticons, sound effects, names, codes, "
    "and other non-prose content. Do not prefer a candidate merely because "
    "it contains more target-language characters. Reject unrelated prose, "
    "hallucination, omission, or a semantic/function mismatch. "
    "A source-script phonetic spelling may become a canonical Latin proper "
    "title or name only when that identity is clear; reject unrelated Latin "
    "prose or uncertain phonetic matches. Choose "
    "accept_new only when the new candidate is clearly faithful and at "
    "least as suitable as the existing translation. Choose keep_existing "
    "when the existing translation is safer. Choose uncertain when the "
    "evidence is insufficient."
)


def _decode_retranslate_quality_decision(raw: str) -> _RetranslateQualityDecision:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end < start:
        raise ValueError("quality review returned no JSON object")
    payload = json.loads(text[start : end + 1])
    if not isinstance(payload, Mapping):
        raise ValueError("quality review JSON must be an object")
    decision = str(payload.get("decision", "")).strip()
    if decision not in _RETRANSLATE_QUALITY_DECISIONS:
        raise ValueError(f"quality review returned invalid decision {decision!r}")
    reason = str(payload.get("reason", "")).strip()
    if not reason:
        reason = "no reason provided"
    return _RetranslateQualityDecision(decision, reason[:500])


def _decode_retranslate_quality_decisions(
    raw: str,
    expected_ids: Sequence[str],
) -> dict[str, _RetranslateQualityDecision]:
    text = raw.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end < start:
        raise ValueError("quality review returned no JSON object")
    payload = json.loads(text[start : end + 1])
    if not isinstance(payload, Mapping):
        raise ValueError("quality review JSON must be an object")
    decisions = payload.get("decisions")
    if not isinstance(decisions, Mapping):
        raise ValueError("quality review JSON is missing decisions")
    decoded: dict[str, _RetranslateQualityDecision] = {}
    for item_id in expected_ids:
        entry = decisions.get(item_id)
        if not isinstance(entry, Mapping):
            raise ValueError(f"quality review is missing item {item_id!r}")
        decision = str(entry.get("decision", "")).strip()
        if decision not in _RETRANSLATE_QUALITY_DECISIONS:
            raise ValueError(
                f"quality review returned invalid decision {decision!r} "
                f"for item {item_id!r}"
            )
        reason = str(entry.get("reason", "")).strip() or "no reason provided"
        decoded[item_id] = _RetranslateQualityDecision(decision, reason[:500])
    return decoded


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds")


_CHUNK_SIZE_FLOOR = 8
_CHUNK_SIZE_FALLBACK_WHEN_UNBOUNDED = 32
_GLOSSARY_LEGACY_DEFAULT_CHUNK_TOKEN_LIMIT = 4_000
_GLOSSARY_DEFAULT_CHUNK_TOKEN_LIMIT = 2_000


def _derive_chunk_size(input_token_limit: int) -> int:
    if input_token_limit <= 0:
        return _CHUNK_SIZE_FALLBACK_WHEN_UNBOUNDED
    return max(_CHUNK_SIZE_FLOOR, input_token_limit // 16)


def _estimate_dynamic_input_tokens(text: str) -> int:
    if not text:
        return 0
    ascii_chars = sum(1 for char in text if ord(char) < 128)
    non_ascii_chars = len(text) - ascii_chars
    return max(1, (ascii_chars + 3) // 4 + non_ascii_chars)


def _effective_glossary_chunk_token_limit(value: int) -> int:
    if value == _GLOSSARY_LEGACY_DEFAULT_CHUNK_TOKEN_LIMIT:
        return _GLOSSARY_DEFAULT_CHUNK_TOKEN_LIMIT
    return max(0, value)


def _new_task_id(kind: str) -> str:
    return f"{kind}-{uuid4().hex[:12]}"


def _hash_text(text: str) -> str:
    return "sha256:" + hashlib.sha256(text.encode("utf-8")).hexdigest()


def _retranslate_source_text(seg_data: Mapping[str, object]) -> str:
    value = seg_data.get("original_text")
    if value is None:
        value = seg_data.get("prompt_text", "")
    return str(value)


def _find_segment_payload(
    snapshot: TaskSnapshot, segment_id: str
) -> Mapping[str, object] | None:
    for subtask in snapshot.subtasks:
        segments = subtask.request_payload.get("segments")
        if not isinstance(segments, list):
            continue
        for seg in segments:
            if isinstance(seg, Mapping) and seg.get("segment_id") == segment_id:
                return seg
    return None


def _build_retranslate_task_index(snapshot: TaskSnapshot) -> _RetranslateTaskIndex:
    owner_subtask_ids: dict[str, list[str]] = {}
    for subtask in snapshot.subtasks:
        segments = subtask.request_payload.get("segments")
        if not isinstance(segments, list):
            continue
        for segment in segments:
            if not isinstance(segment, Mapping):
                continue
            segment_id = segment.get("segment_id")
            if segment_id in (None, ""):
                continue
            key = str(segment_id)
            owners = owner_subtask_ids.setdefault(key, [])
            if subtask.id not in owners:
                owners.append(subtask.id)
    return _RetranslateTaskIndex(
        task_updated_at=snapshot.record.updated_at,
        owner_subtask_ids={
            segment_id: tuple(owner_ids)
            for segment_id, owner_ids in owner_subtask_ids.items()
        },
    )


def _read_segment_dst(snapshot: TaskSnapshot, segment_id: str) -> str:
    translations, _low_confidence = collect_segment_state_from_authoritative_subtasks(
        snapshot.subtasks
    )
    if segment_id in translations:
        return translations[segment_id]
    return ""


def _patch_segment_dst(
    cache: TaskCache,
    snapshot: TaskSnapshot,
    segment_id: str,
    new_dst: str,
) -> None:
    _patch_segment_dsts(cache, snapshot, {segment_id: new_dst})


def _patch_segment_dsts(
    cache: TaskCache,
    snapshot: TaskSnapshot,
    updates: Mapping[str, str],
) -> None:
    confidence_entries = {
        segment_id: _confidence_entry_for_segment(snapshot, segment_id, new_dst)
        for segment_id, new_dst in updates.items()
    }
    for subtask in snapshot.subtasks:
        segments = subtask.request_payload.get("segments")
        if not isinstance(segments, list):
            continue
        matched_ids = {
            str(segment.get("segment_id"))
            for segment in segments
            if isinstance(segment, Mapping)
            and segment.get("segment_id") in updates
        }
        if not matched_ids:
            continue
        try:
            payload = json.loads(subtask.response_content) if subtask.response_content else {}
        except json.JSONDecodeError:
            payload = {}
        if not isinstance(payload, dict):
            payload = {}
        payload.setdefault("version", 2)
        translations = payload.setdefault("translations", {})
        if not isinstance(translations, dict):
            translations = {}
            payload["translations"] = translations
        for segment_id in matched_ids:
            translations[segment_id] = updates[segment_id]
            mark_accepted_override(payload, segment_id)
            _replace_low_confidence_entry(
                payload,
                segment_id,
                confidence_entries[segment_id],
            )
        cache.save_subtask(
            replace(subtask, response_content=json.dumps(payload, ensure_ascii=False))
        )


def _confidence_entry_for_segment(
    snapshot: TaskSnapshot, segment_id: str, dst: str
) -> dict[str, object] | None:
    source = ""
    for subtask in snapshot.subtasks:
        segments = subtask.request_payload.get("segments")
        if not isinstance(segments, list):
            continue
        for segment in segments:
            if (
                not isinstance(segment, Mapping)
                or segment.get("segment_id") != segment_id
            ):
                continue
            source = str(
                segment.get("original_text")
                or segment.get("prompt_text")
                or ""
            )
            break
        if source:
            break
    if not source:
        return None
    try:
        source_language = Language(snapshot.record.metadata.get("source_language", ""))
    except ValueError:
        source_language = None
    try:
        target_language = Language(snapshot.record.metadata.get("target_language", ""))
    except ValueError:
        target_language = None
    verdict = evaluate_segment_confidence(
        source,
        dst,
        min_length_ratio=0.25,
        max_length_ratio=4.0,
        max_punctuation_delta=12,
        source_language=source_language,
        target_language=target_language,
    )
    if not verdict.is_low_confidence:
        return None
    entry: dict[str, object] = {
        "segment_id": segment_id,
        "reasons": list(verdict.reasons),
    }
    if verdict.tags:
        entry["tags"] = list(verdict.tags)
    return entry


def _replace_low_confidence_entry(
    payload: dict[str, object],
    segment_id: str,
    entry: dict[str, object] | None,
) -> None:
    existing = payload.get("low_confidence")
    if isinstance(existing, list):
        next_entries = [
            item
            for item in existing
            if isinstance(item, dict) and item.get("segment_id") != segment_id
        ]
    else:
        next_entries = []
    if entry is not None:
        next_entries.append(entry)
    payload["low_confidence"] = next_entries


def _should_keep_existing_retranslation(
    snapshot: TaskSnapshot,
    segment_id: str,
    existing_dst: str,
    candidate_dst: str,
    candidate_entry: Mapping[str, object] | None,
) -> bool:
    source_segment = _find_segment_payload(snapshot, segment_id)
    if source_segment is not None and _is_preserved_nonprose_retranslation(
        _retranslate_source_text(source_segment), candidate_dst
    ):
        return False
    candidate_tags = set(_string_values((candidate_entry or {}).get("tags")))
    if "source_residue" not in candidate_tags:
        return False
    if candidate_dst == existing_dst:
        return True
    existing_entry = _confidence_entry_for_segment(
        snapshot, segment_id, existing_dst
    )
    return _retranslation_quality_rank(candidate_entry) >= _retranslation_quality_rank(
        existing_entry
    )


def _is_preserved_nonprose_retranslation(
    source_text: str, candidate_text: str
) -> bool:
    source = source_text.strip()
    candidate = candidate_text.strip()
    if not source or not candidate or not _looks_like_compact_nonprose(source):
        return False
    if source == candidate:
        return True
    if not _looks_like_compact_nonprose(candidate):
        return False
    if len(candidate) > max(80, len(source) * 3):
        return False
    source_structure = _nonprose_structure(source)
    candidate_structure = Counter(_nonprose_structure(candidate))
    if not source_structure:
        return False
    remaining = Counter(source_structure)
    overlap = sum(
        min(count, candidate_structure.get(char, 0))
        for char, count in remaining.items()
    )
    return overlap / len(source_structure) >= 0.6


def _looks_like_compact_nonprose(text: str) -> bool:
    compact = "".join(text.split())
    if not compact or len(compact) > 80:
        return False
    jamo_count = sum(_is_korean_jamo(char) for char in compact)
    letter_count = sum(unicodedata.category(char).startswith("L") for char in compact)
    lexical_letter_count = letter_count - jamo_count
    expressive_count = sum(
        unicodedata.category(char)[0] in {"P", "S"}
        and char not in "[](){}<>【】（）"
        for char in compact
    )
    if jamo_count and (
        lexical_letter_count == 0
        or (
            compact[0] in "[({<【（"
            and expressive_count > 0
            and lexical_letter_count <= 4
        )
    ):
        return True
    return (
        expressive_count >= 2
        and expressive_count / len(compact) >= 0.15
        and lexical_letter_count <= max(4, len(compact) // 3)
    )


def _nonprose_structure(text: str) -> list[str]:
    return [
        char
        for char in text
        if _is_korean_jamo(char) or unicodedata.category(char)[0] in {"P", "S"}
    ]


def _is_korean_jamo(char: str) -> bool:
    codepoint = ord(char)
    return (
        0x1100 <= codepoint <= 0x11FF
        or 0x3130 <= codepoint <= 0x318F
        or 0xA960 <= codepoint <= 0xA97F
        or 0xD7B0 <= codepoint <= 0xD7FF
        or 0xFFA0 <= codepoint <= 0xFFDC
    )


def _retranslation_quality_rank(
    entry: Mapping[str, object] | None,
) -> tuple[int, int, int, int, int]:
    tags = set(_string_values((entry or {}).get("tags")))
    return (
        int("source_residue" in tags),
        int("verbatim_echo" in tags),
        int("target_language_weak" in tags),
        int("model_chatter" in tags),
        len(tags),
    )


def _string_values(value: object) -> tuple[str, ...]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes)):
        return ()
    return tuple(str(item) for item in value if item not in (None, ""))


def _coerce_text_preserve_rules(
    raw: Sequence[Mapping[str, object]],
) -> tuple[TextPreserveRule, ...]:
    """Build immutable ``TextPreserveRule`` tuple from settings.

    Bad rows (empty pattern) are dropped silently — the UI's
    pattern editor enforces non-empty values, so anything reaching
    this path is either a hand-edited JSON or a stale entry.
    """

    rules: list[TextPreserveRule] = []
    for entry in raw:
        if not isinstance(entry, Mapping):
            continue
        pattern = str(entry.get("pattern", "")).strip()
        if not pattern:
            continue
        rules.append(
            TextPreserveRule(
                pattern=pattern,
                note=str(entry.get("note", "")),
                enabled=bool(entry.get("enabled", True)),
            )
        )
    return tuple(rules)


def _coerce_translation_replacements(
    raw: Sequence[Mapping[str, object]],
) -> tuple[TranslationReplacementRule, ...]:
    """Build immutable translation ``ReplacementRule`` tuple from
    settings entries. Distinct from ``transoria.tools.replacement.
    ReplacementRule`` — translation rules carry a ``note`` field for
    user annotation."""

    rules: list[TranslationReplacementRule] = []
    for entry in raw:
        if not isinstance(entry, Mapping):
            continue
        src = str(entry.get("src", "")).strip()
        if not src:
            continue
        rules.append(
            TranslationReplacementRule(
                src=src,
                dst=str(entry.get("dst", "")),
                regex=bool(entry.get("regex", False)),
                case_sensitive=bool(entry.get("case_sensitive", False)),
                note=str(entry.get("note", "")),
                enabled=bool(entry.get("enabled", True)),
            )
        )
    return tuple(rules)


def _coerce_language(value: str, *, field: str) -> Language:
    try:
        return Language(value)
    except ValueError as exc:
        raise BridgeError.invalid_argument(
            f"unsupported language {value!r} for {field}.",
            field=field,
        ) from exc


def _normalized_dir(value: object) -> str:
    """Compare-friendly form of a directory string. Empty / non-string
    inputs collapse to ``""`` so two missing values match each other,
    not a real path.
    """

    if not isinstance(value, str) or not value:
        return ""
    try:
        return normalize_path_key(Path(value))
    except Exception:
        return value


def _require_directory(value: str, *, field: str) -> Path:
    if not value:
        raise BridgeError.invalid_argument(
            f"{field} is required.",
            field=field,
        )
    path = Path(value)
    if not path.exists():
        raise BridgeError.invalid_argument(
            f"{field} does not exist: {value!r}",
            field=field,
        )
    if not path.is_dir():
        raise BridgeError.invalid_argument(
            f"{field} is not a directory: {value!r}",
            field=field,
        )
    return path


def _require_input_with_supported_files(path: Path, *, field: str) -> None:
    """Reject empty / unsupported-only input folders before a task starts.

    Translation and glossary cannot do useful work on a folder that
    contains no ``.epub`` / ``.txt`` files; without this check the task
    seeds, runs zero subtasks, and reports COMPLETED with no output —
    which looks like a silent failure to the user. Raise a typed bridge
    error so the frontend's existing error banner surfaces a clear
    message instead.
    """

    if not scan_input_directory(path):
        raise BridgeError.invalid_argument(
            f"{field} contains no supported files (.epub, .txt): {path!s}",
            field=field,
        )


def _require_distinct_translation_folders(
    input_dir: Path, output_dir: Path
) -> None:
    """Reject translation configs where output lives inside input.

    The scanner is recursive (``rglob("*")``), so any ``.epub`` / ``.txt``
    written under the input tree gets picked up on the next run. The
    pathological case is ``input == output``: ``translation-failed-
    subtasks.txt`` and ``*-zh.epub`` from one run become "input files"
    for the next, producing cascading suffix files like
    ``translation-failed-subtasks-zh-kr-zh-kr-zh.txt`` and re-translating
    already-translated novels.

    Glossary tasks intentionally allow ``input == output`` — their outputs
    are ``.xlsx`` / ``.json`` and never match the scanner's
    ``.epub``/``.txt`` filter.
    """

    in_key = normalize_path_key(input_dir)
    out_key = normalize_path_key(output_dir)
    if in_key == out_key:
        raise BridgeError.invalid_argument(
            "Translation input_folder and output_folder must be different "
            "directories — writing translated files into the input would "
            "cause them to be re-scanned and re-translated on the next run.",
            field="output_folder",
            message_key="translation.input_equals_output",
        )
    # Resolve to handle ``..`` / symlink edge cases before the descendant
    # check; bail silently if either side refuses to resolve so the
    # downstream IO error gives a clearer message than this guard would.
    try:
        resolved_in = input_dir.resolve()
        resolved_out = output_dir.resolve()
    except OSError:
        return
    try:
        resolved_out.relative_to(resolved_in)
    except ValueError:
        return  # output is not inside input — OK.
    raise BridgeError.invalid_argument(
        f"Translation output_folder must not live inside input_folder "
        f"({resolved_out!s} is under {resolved_in!s}); the recursive "
        "scanner would pick up generated files on the next run.",
        field="output_folder",
        message_key="translation.output_inside_input",
    )


def _ensure_output_dir(value: str, *, field: str) -> Path:
    if not value:
        raise BridgeError.invalid_argument(
            f"{field} is required.",
            field=field,
        )
    path = Path(value)
    try:
        path.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise BridgeError(
            "bridge.io_error",
            describe_os_error(exc, action=f"create {field}"),
            retryable=False,
            details={"field": field, "path": value},
        ) from exc
    return path

def _format_header(record: TaskRecord) -> dict[str, object]:
    return {
        "id": record.id,
        "kind": record.kind.value,
        "status": record.status.value,
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


def _format_request_events(
    events: Sequence[Mapping[str, object]],
    *,
    limit: int | None,
    offset: int = 0,
    status: str = "",
    task_status: TaskStatus | None = None,
    subtask_statuses: Mapping[str, SubtaskStatus] | None = None,
    live_subtask_ids: set[str] | None = None,
) -> tuple[list[dict[str, object]], int]:
    latest_by_request: dict[str, dict[str, object]] = {}
    order: dict[str, int] = {}
    for index, event in enumerate(events):
        request_id = str(event.get("request_id", ""))
        if not request_id:
            continue
        previous = latest_by_request.get(request_id, {})
        merged = {**previous, **dict(event)}
        latest_by_request[request_id] = merged
        order[request_id] = index
    rows = sorted(
        latest_by_request.values(),
        key=lambda row: order.get(str(row.get("request_id", "")), 0),
        reverse=True,
    )
    rows = [
        _cancel_stale_running_request_event(
            row,
            task_status=task_status,
            subtask_statuses=subtask_statuses or {},
            live_subtask_ids=live_subtask_ids or set(),
        )
        for row in rows
    ]
    if status:
        rows = [row for row in rows if str(row.get("status", "")) == status]
    total = len(rows)
    if offset > 0:
        rows = rows[offset:]
    if limit is not None:
        rows = rows[:limit]
    return rows, total


def _translation_request_segment_refs(
    events: Sequence[Mapping[str, object]],
    snapshot: TaskSnapshot,
) -> list[dict[str, object]]:
    index_maps: dict[str, dict[str, str]] = {}
    for subtask in snapshot.subtasks:
        segments = subtask.request_payload.get("segments", [])
        if not isinstance(segments, list):
            continue
        index_map: dict[str, str] = {}
        for segment in segments:
            if not isinstance(segment, Mapping):
                continue
            segment_id = segment.get("segment_id")
            chunk_index = segment.get("chunk_index")
            if segment_id in (None, "") or chunk_index in (None, ""):
                continue
            index_map[str(chunk_index)] = str(segment_id)
        if index_map:
            index_maps[subtask.id] = index_map

    translations, _low_confidence = collect_segment_state_from_authoritative_subtasks(
        snapshot.subtasks
    )
    enriched: list[dict[str, object]] = []
    for raw_event in events:
        event = dict(raw_event)
        index_map = index_maps.get(str(event.get("subtask_id", "")))
        response_text = event.get("response_text") or event.get(
            "partial_response_text"
        )
        if not index_map or not isinstance(response_text, str):
            enriched.append(event)
            continue

        response_rows: dict[str, str] = {}
        for line in response_text.splitlines():
            candidate = line.strip()
            if not candidate.startswith("{") or not candidate.endswith("}"):
                continue
            try:
                decoded = json.loads(candidate)
            except json.JSONDecodeError:
                continue
            if not isinstance(decoded, Mapping):
                continue
            for request_index, model_text in decoded.items():
                key = str(request_index)
                if key in index_map:
                    response_rows[key] = str(model_text)

        segment_refs: list[dict[str, str]] = []
        for request_index, model_text in response_rows.items():
            segment_id = index_map[request_index]
            current_text = translations.get(segment_id)
            cache_status = (
                "missing"
                if current_text is None
                else "matched"
                if current_text == model_text
                else "different"
            )
            segment_refs.append(
                {
                    "request_index": request_index,
                    "segment_id": segment_id,
                    "cache_status": cache_status,
                }
            )
        if segment_refs:
            event["segment_refs"] = segment_refs
        enriched.append(event)
    return enriched


def _cancel_stale_running_request_event(
    row: Mapping[str, object],
    *,
    task_status: TaskStatus | None,
    subtask_statuses: Mapping[str, SubtaskStatus],
    live_subtask_ids: set[str] | None = None,
) -> dict[str, object]:
    current = dict(row)
    if current.get("status") != "running":
        return current
    subtask_id = str(current.get("subtask_id", ""))
    if subtask_id in (live_subtask_ids or set()):
        return current
    subtask_status = subtask_statuses.get(subtask_id)
    if subtask_status is SubtaskStatus.RUNNING:
        return current
    if subtask_status is None and task_status in _ZOMBIE_TASK_STATES:
        return current
    current["status"] = "cancelled"
    current["phase"] = "cancelled"
    current.setdefault("error", "Request was cancelled by a previous app session.")
    return current


def _truncate_request_log_text(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def _load_subtask_debug_log(
    cache: TaskCache, *, task_id: str, subtask_id: str
) -> Mapping[str, object]:
    safe_id = _DEBUG_LOG_FILENAME_SAFE.sub("_", subtask_id)
    path = cache.task_dir(task_id) / "debug" / f"{safe_id}.json"
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, Mapping):
        return {}
    return payload


def _last_debug_response(debug_payload: Mapping[str, object]) -> str:
    raw = debug_payload.get("raw_response")
    if isinstance(raw, str) and raw:
        return raw
    attempts = debug_payload.get("attempts")
    if isinstance(attempts, Sequence) and not isinstance(attempts, (str, bytes)):
        for attempt in reversed(attempts):
            if not isinstance(attempt, Mapping):
                continue
            raw = attempt.get("raw_response")
            if isinstance(raw, str) and raw:
                return raw
    restored = debug_payload.get("restored_response")
    if isinstance(restored, str) and restored:
        return restored
    return ""


def _failed_subtask_request_events(
    cache: TaskCache,
    *,
    record: TaskRecord,
    snapshot: TaskSnapshot,
    request_events: Sequence[Mapping[str, object]],
) -> tuple[dict[str, object], ...]:
    metadata_by_subtask: dict[str, dict[str, object]] = {}
    inherited_keys = (
        "model_profile_id",
        "model_id",
        "provider_format",
        "provider_attempt",
        "prompt_chars",
        "timeout_seconds",
        "subtask_attempt",
    )
    copied_keys = tuple(key for key in inherited_keys if key != "subtask_attempt")
    local_failure_attempts_by_subtask: dict[str, set[int]] = {}
    for event in request_events:
        subtask_id = str(event.get("subtask_id", ""))
        if not subtask_id:
            continue
        metadata = metadata_by_subtask.setdefault(subtask_id, {})
        for key in inherited_keys:
            if key in event:
                metadata[key] = event[key]
        if event.get("local_failure") is True:
            try:
                attempt = int(event.get("subtask_attempt", 0) or 0)
            except (TypeError, ValueError):
                attempt = 0
            if attempt > 0:
                local_failure_attempts_by_subtask.setdefault(
                    subtask_id, set()
                ).add(attempt)

    synthetic: list[dict[str, object]] = []
    for subtask in snapshot.subtasks:
        if subtask.status is not SubtaskStatus.FAILED:
            continue
        base = metadata_by_subtask.get(subtask.id, {})
        subtask_attempt = subtask.attempt_count or int(
            base.get("subtask_attempt", 1) or 1
        )
        if subtask_attempt in local_failure_attempts_by_subtask.get(
            subtask.id, set()
        ):
            continue
        debug_payload = _load_subtask_debug_log(
            cache, task_id=record.id, subtask_id=subtask.id
        )
        error = str(
            debug_payload.get("terminal_error") or subtask.last_error or "Subtask failed."
        )
        raw_response = _last_debug_response(debug_payload)
        response_text = error
        if raw_response:
            response_text = f"{error}\n\n--- Last model response ---\n{raw_response}"

        event: dict[str, object] = {
            "schema_version": 1,
            "request_id": f"{record.id}:{subtask.id}:local-failure",
            "timestamp": subtask.last_error_at or record.updated_at or record.created_at,
            "task_id": record.id,
            "subtask_id": subtask.id,
            "subtask_attempt": subtask_attempt,
            "status": "failed",
            "phase": "validation",
            "last_activity_at": subtask.last_error_at
            or record.updated_at
            or record.created_at,
            "label": f"{subtask.id} local validation",
            "error": _truncate_request_log_text(
                error, _REQUEST_LOG_ERROR_PREVIEW_LIMIT
            ),
            "response_text": _truncate_request_log_text(
                response_text, _REQUEST_LOG_RESPONSE_PREVIEW_LIMIT
            ),
            "local_failure": True,
        }
        for key in copied_keys:
            if key in base:
                event[key] = base[key]
        synthetic.append(event)
    return tuple(synthetic)


def _progress_to_block(
    progress, *, elapsed_seconds: float, longest_running_seconds: float = 0.0
) -> dict[str, object]:
    return {
        "total": progress.total,
        "pending": progress.pending,
        "running": progress.running,
        "completed": progress.completed,
        "failed": progress.failed,
        "skipped": progress.skipped,
        "elapsed_seconds": elapsed_seconds,
        "rate_per_second": progress.rate_per_second,
        "longest_running_seconds": longest_running_seconds,
    }


def _usage_to_block(usage) -> dict[str, object]:
    return {
        "input_tokens": usage.input_tokens,
        "output_tokens": usage.output_tokens,
        "cached_input_tokens": usage.cached_input_tokens,
        "total_tokens": usage.input_tokens + usage.output_tokens,
    }


def _glossary_review_round_progress(
    record: TaskRecord, metadata: Mapping[str, object]
) -> dict[str, int] | None:
    if record.kind is not TaskKind.GLOSSARY_REVIEW:
        return None
    total_rounds = int(metadata.get("review_rounds_total", 0) or 0)
    if total_rounds <= 0:
        return None
    if record.status is TaskStatus.COMPLETED:
        return {
            "total_rounds": total_rounds,
            "current_round": total_rounds,
            "completed_rounds": total_rounds,
            "current_total_batches": int(
                metadata.get("review_round_total_batches", 0) or 0
            ),
            "current_completed_batches": int(
                metadata.get("review_round_completed_batches", 0) or 0
            ),
        }
    return {
        "total_rounds": total_rounds,
        "current_round": int(metadata.get("review_round_current", 0) or 0),
        "completed_rounds": int(metadata.get("review_round_completed", 0) or 0),
        "current_total_batches": int(
            metadata.get("review_round_total_batches", 0) or 0
        ),
        "current_completed_batches": int(
            metadata.get("review_round_completed_batches", 0) or 0
        ),
    }


def _format_snapshot(snapshot: TaskSnapshot) -> dict[str, object]:
    elapsed_seconds = _task_elapsed_seconds(snapshot.record)
    progress = snapshot.progress(elapsed_seconds=elapsed_seconds)
    usage = snapshot.usage()
    metadata = dict(snapshot.record.metadata)
    longest_running_seconds = _longest_running_seconds(snapshot)
    # When subtasks/ has been pruned post-completion (see
    # ``_maybe_cleanup_cache``), the live progress recomputes to 0/0;
    # surface the frozen totals stashed in metadata instead so the Run
    # page keeps showing the completed run's final stats until the user
    # starts a fresh task.
    if len(snapshot.subtasks) == 0:
        frozen_progress = metadata.get("final_progress")
        if isinstance(frozen_progress, dict):
            frozen_elapsed = _coerce_float(
                frozen_progress.get("elapsed_seconds"), elapsed_seconds
            )
            progress_block = {
                "total": int(frozen_progress.get("total", 0)),
                "pending": int(frozen_progress.get("pending", 0)),
                "running": int(frozen_progress.get("running", 0)),
                "completed": int(frozen_progress.get("completed", 0)),
                "failed": int(frozen_progress.get("failed", 0)),
                "skipped": int(frozen_progress.get("skipped", 0)),
                "elapsed_seconds": frozen_elapsed,
                "rate_per_second": _coerce_float(
                    frozen_progress.get("rate_per_second"), 0.0
                ),
                "longest_running_seconds": 0.0,
            }
        else:
            progress_block = _progress_to_block(progress, elapsed_seconds=elapsed_seconds)
        frozen_usage = metadata.get("final_usage")
        if isinstance(frozen_usage, dict):
            input_t = int(frozen_usage.get("input_tokens", 0))
            output_t = int(frozen_usage.get("output_tokens", 0))
            cached_input_t = int(frozen_usage.get("cached_input_tokens", 0))
            usage_block = {
                "input_tokens": input_t,
                "output_tokens": output_t,
                "cached_input_tokens": cached_input_t,
                "total_tokens": input_t + output_t,
            }
        else:
            usage_block = _usage_to_block(usage)
    else:
        progress_block = _progress_to_block(
            progress,
            elapsed_seconds=elapsed_seconds,
            longest_running_seconds=longest_running_seconds,
        )
        usage_block = _usage_to_block(usage)
    low_conf_block = _low_confidence_summary(snapshot, metadata)
    return {
        "header": _format_header(snapshot.record),
        "progress": progress_block,
        "usage": usage_block,
        "low_confidence": low_conf_block,
        "round_progress": _glossary_review_round_progress(
            snapshot.record, metadata
        ),
        # Per-chunk status drives the chunk-grid UX. Tuple-of-objects
        # is preserved in the order the orchestrator seeded them, so
        # the grid renders chunk-0 leftmost. ``last_error`` is included
        # only for FAILED chunks so the chunk-grid tooltip can tell the
        # user *why* a red square is red without forcing a separate
        # API call per chunk.
        "subtasks": [
            {
                "id": s.id,
                "status": s.status.value,
                "attempts": s.attempt_count,
                "started_at": s.started_at if s.status is SubtaskStatus.RUNNING else "",
                "last_error": s.last_error if s.status is SubtaskStatus.FAILED else "",
            }
            for s in snapshot.subtasks
        ],
        "active_model_id": metadata.get("active_model_id")
        or metadata.get("model_id"),
        "active_prompt_id": metadata.get("active_prompt_id")
        or metadata.get("prompt_preset_id"),
        "metadata": metadata,
    }


def _low_confidence_summary(
    snapshot: TaskSnapshot, metadata: Mapping[str, object]
) -> dict[str, int]:
    # When subtasks are cleaned post-completion, fall back to the frozen
    # snapshot stashed in metadata at task finalize.
    if not snapshot.subtasks:
        frozen = metadata.get("final_low_confidence")
        if isinstance(frozen, Mapping):
            return {
                "total": int(frozen.get("total", 0)),
                "source_residue": int(frozen.get("source_residue", 0)),
            }
    _translations, low_confidence = collect_segment_state_from_authoritative_subtasks(
        snapshot.subtasks
    )
    total = len(low_confidence)
    residue = sum(
        1
        for entry in low_confidence.values()
        if "source_residue" in entry.get("tags", [])
    )
    return {"total": total, "source_residue": residue}


def _task_elapsed_seconds(record: TaskRecord) -> float:
    return elapsed_seconds_for_record(record)


def _longest_running_seconds(snapshot: TaskSnapshot) -> float:
    now = datetime.now(timezone.utc)
    longest = 0.0
    for subtask in snapshot.subtasks:
        if subtask.status is not SubtaskStatus.RUNNING:
            continue
        started = _parse_iso_timestamp(subtask.started_at)
        if started is None:
            continue
        longest = max(longest, (now - started).total_seconds())
    return max(0.0, longest)


def _parse_iso_timestamp(value: str) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _coerce_float(value: object, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _format_failures(snapshot: TaskSnapshot) -> list[dict[str, object]]:
    failures: list[dict[str, object]] = []
    for subtask in snapshot.subtasks:
        if subtask.status is not SubtaskStatus.FAILED:
            continue
        last_error = subtask.last_error or ""
        code = ""
        message = last_error
        if last_error.startswith("[") and "] " in last_error:
            close = last_error.index("] ")
            code = last_error[1:close]
            message = last_error[close + 2 :]
        source_file = ""
        payload = subtask.request_payload
        if isinstance(payload, Mapping):
            for key in ("source_file", "file_path", "path"):
                value = payload.get(key)
                if isinstance(value, str) and value:
                    source_file = value
                    break
        failures.append(
            {
                "subtask_id": subtask.id,
                "source_file": source_file,
                "message": message,
                "attempts": subtask.attempt_count,
                "last_error_code": code,
                "last_error_at": subtask.last_error_at,
            }
        )
    return failures


def _translation_result_payload(
    result: TranslationRunResult, *, config: TranslationConfig
) -> dict[str, object]:
    statistics = result.statistics
    bilingual_folder: str | None = None
    if result.bilingual_outputs:
        bilingual_folder = str(config.output_dir / config.bilingual_subfolder)
    return {
        "kind": "translation",
        "output_folder": str(config.output_dir),
        "bilingual_folder": bilingual_folder,
        "translated_files": [str(p) for p in result.translated_outputs],
        "bilingual_files": [str(p) for p in result.bilingual_outputs],
        "statistics_json_path": str(result.statistics_path)
        if result.statistics_path
        else None,
        "glossary_report_path": str(result.glossary_report_path)
        if result.glossary_report_path
        else None,
        "glossary_report_json_path": str(result.glossary_report_json_path)
        if result.glossary_report_json_path
        else None,
        "processed_files": list(statistics.processed_files),
        "completed_segments": statistics.completed_segments,
        "total_segments": statistics.total_segments,
    }


def _glossary_result_payload(
    result: GlossaryExtractionResult, *, config: GlossaryConfig
) -> dict[str, object]:
    per_novel = [
        _glossary_artifact_payload(item)
        for item in result.glossary_outputs_per_file
    ]
    combined = (
        _glossary_artifact_payload(result.combined_output)
        if result.combined_output is not None
        else None
    )
    decode_paths = [
        str(item.decode_issue_path)
        for item in result.glossary_outputs_per_file
        if item.decode_issue_path is not None
    ]
    return {
        "kind": "glossary",
        "output_folder": str(config.output_dir),
        "per_novel_artifacts": per_novel,
        "combined_artifact": combined,
        "statistics_json_path": str(result.statistics_path)
        if result.statistics_path
        else None,
        "decode_issue_path": decode_paths[0] if decode_paths else None,
    }


def _glossary_review_result_payload(
    result: GlossaryReviewResult, *, config: GlossaryReviewConfig
) -> dict[str, object]:
    return {
        "kind": "glossary_review",
        "output_folder": str(config.input_dir),
        "output_path": str(result.output_path) if result.output_path else None,
        "report_path": str(result.report_path) if result.report_path else None,
        "changed_count": result.changed_count,
    }


def _glossary_artifact_payload(artifact: GlossaryArtifactSet) -> dict[str, object]:
    payload: dict[str, object] = {
        "novel_name": artifact.novel_name,
        "xlsx_path": str(artifact.xlsx_path),
        "json_path": str(artifact.json_path),
        "references_path": str(artifact.references_path),
    }
    if artifact.decode_issue_path is not None:
        payload["decode_issue_path"] = str(artifact.decode_issue_path)
    return payload


def _partial_translation_payload(
    snapshot: TaskSnapshot, *, output_dir: Path, statistics_dir: Path
) -> dict[str, object]:
    stats_path = statistics_dir / STATISTICS_FILENAME_JSON
    glossary_report_path = statistics_dir / GLOSSARY_REPORT_FILENAME_MD
    glossary_report_json_path = statistics_dir / GLOSSARY_REPORT_FILENAME_JSON
    stats = _read_json_file(stats_path)
    translated = _string_list(stats.get("translated_outputs")) if stats else []
    bilingual = _string_list(stats.get("bilingual_outputs")) if stats else []
    if not translated:
        translated = _scan_files(output_dir, exclude_dirs={"bilingual outputs"})
    if not bilingual:
        bilingual = _scan_bilingual_files(output_dir)
    progress = snapshot.progress()
    bilingual_folder = str(Path(bilingual[0]).parent) if bilingual else None
    return {
        "kind": "translation",
        "partial": True,
        "output_folder": str(output_dir),
        "bilingual_folder": bilingual_folder,
        "translated_files": translated,
        "bilingual_files": bilingual,
        "statistics_json_path": str(stats_path) if stats_path.exists() else None,
        "glossary_report_path": (
            str(glossary_report_path) if glossary_report_path.exists() else None
        ),
        "glossary_report_json_path": (
            str(glossary_report_json_path)
            if glossary_report_json_path.exists()
            else None
        ),
        "processed_files": _string_list(stats.get("processed_files"))
        if stats
        else [],
        "completed_segments": (
            int(stats.get("completed_segments", progress.completed))
            if stats
            else progress.completed
        ),
        "total_segments": (
            int(stats.get("total_segments", progress.total))
            if stats
            else progress.total
        ),
    }


def _partial_glossary_payload(
    *, output_dir: Path, input_folder_name: str, statistics_dir: Path
) -> dict[str, object]:
    artifacts: list[dict[str, object]] = []
    combined: dict[str, object] | None = None
    for xlsx_path in sorted(output_dir.glob(f"*{GLOSSARY_FILENAME_XLSX}")):
        basename = xlsx_path.name.removesuffix(GLOSSARY_FILENAME_XLSX)
        json_path = output_dir / f"{basename}{GLOSSARY_FILENAME_JSON}"
        references_path = output_dir / f"{basename}{GLOSSARY_FILENAME_REFERENCES}"
        decode_path = output_dir / f"{basename}{GLOSSARY_FILENAME_DECODE_ISSUES}"
        item: dict[str, object] = {
            "novel_name": basename,
            "xlsx_path": str(xlsx_path),
            "json_path": str(json_path) if json_path.exists() else None,
            "references_path": str(references_path)
            if references_path.exists()
            else None,
        }
        if decode_path.exists():
            item["decode_issue_path"] = str(decode_path)
        if input_folder_name and basename == input_folder_name:
            combined = item
        else:
            artifacts.append(item)
    decode_paths = [
        str(path)
        for path in sorted(output_dir.glob(f"*{GLOSSARY_FILENAME_DECODE_ISSUES}"))
    ]
    stats_path = statistics_dir / GLOSSARY_STATISTICS_FILENAME_JSON
    return {
        "kind": "glossary",
        "partial": True,
        "output_folder": str(output_dir),
        "per_novel_artifacts": artifacts,
        "combined_artifact": combined,
        "statistics_json_path": str(stats_path) if stats_path.exists() else None,
        "decode_issue_path": decode_paths[0] if decode_paths else None,
    }


def _partial_glossary_review_payload(
    *, output_dir: Path, output_filename: str
) -> dict[str, object]:
    path = output_dir / normalize_output_filename(output_filename)
    return {
        "kind": "glossary_review",
        "partial": True,
        "output_folder": str(output_dir),
        "output_path": str(path) if path.exists() else None,
        "report_path": None,
        "changed_count": 0,
    }


def _partial_replacement_payload(
    snapshot: TaskSnapshot, *, output_dir: Path
) -> dict[str, object]:
    output_files: list[str] = []
    total_replacements = 0
    for subtask in snapshot.subtasks:
        if subtask.status is not SubtaskStatus.COMPLETED:
            continue
        payload = _loads_json_object(subtask.response_content)
        output_path = payload.get("output_path")
        if isinstance(output_path, str) and output_path:
            output_files.append(output_path)
        count = payload.get("replacement_count")
        if isinstance(count, int):
            total_replacements += count
    return {
        "kind": "replacement",
        "partial": True,
        "output_folder": str(output_dir),
        "output_files": output_files,
        "statistics_json_path": None,
        "total_replacements": total_replacements,
    }


def _partial_epub_compress_payload(
    snapshot: TaskSnapshot, *, output_folder: Path, report_path: Path
) -> dict[str, object]:
    compressed = 0
    failed = 0
    output_files: list[str] = []
    for subtask in snapshot.subtasks:
        if subtask.status is SubtaskStatus.COMPLETED:
            compressed += 1
            payload = _loads_json_object(subtask.response_content)
            output_path = payload.get("output_path")
            if isinstance(output_path, str) and output_path:
                output_files.append(output_path)
        elif subtask.status is SubtaskStatus.FAILED:
            failed += 1
    return {
        "kind": "epub_compress",
        "partial": True,
        "output_folder": str(output_folder),
        "report_path": str(report_path) if report_path.exists() else None,
        "output_files": output_files,
        "compressed_count": compressed,
        "failed_count": failed,
    }


def _partial_epub_merge_payload(
    snapshot: TaskSnapshot, *, output_folder: Path, report_path: Path
) -> dict[str, object]:
    merged = 0
    failed = 0
    output_files: list[str] = []
    for subtask in snapshot.subtasks:
        if subtask.status is SubtaskStatus.COMPLETED:
            merged += 1
            payload = _loads_json_object(subtask.response_content)
            output_path = payload.get("output_path")
            if isinstance(output_path, str) and output_path:
                output_files.append(output_path)
        elif subtask.status is SubtaskStatus.FAILED:
            failed += 1
    return {
        "kind": "epub_merge",
        "partial": True,
        "output_folder": str(output_folder),
        "report_path": str(report_path) if report_path.exists() else None,
        "output_files": output_files,
        "merged_count": merged,
        "failed_count": failed,
    }


def _partial_epub_convert_payload(
    snapshot: TaskSnapshot, *, output_folder: Path, report_path: Path
) -> dict[str, object]:
    converted = 0
    failed = 0
    output_files: list[str] = []
    for subtask in snapshot.subtasks:
        if subtask.status is SubtaskStatus.COMPLETED:
            converted += 1
            payload = _loads_json_object(subtask.response_content)
            output_path = payload.get("output_path")
            if isinstance(output_path, str) and output_path:
                output_files.append(output_path)
        elif subtask.status is SubtaskStatus.FAILED:
            failed += 1
    return {
        "kind": "epub_convert",
        "partial": True,
        "output_folder": str(output_folder),
        "report_path": str(report_path) if report_path.exists() else None,
        "output_files": output_files,
        "converted_count": converted,
        "failed_count": failed,
    }


def _partial_txt_to_epub_payload(
    snapshot: TaskSnapshot, *, output_folder: Path, report_path: Path
) -> dict[str, object]:
    converted = 0
    failed = 0
    output_files: list[str] = []
    for subtask in snapshot.subtasks:
        if subtask.status is SubtaskStatus.COMPLETED:
            converted += 1
            payload = _loads_json_object(subtask.response_content)
            output_path = payload.get("output_path")
            if isinstance(output_path, str) and output_path:
                output_files.append(output_path)
        elif subtask.status is SubtaskStatus.FAILED:
            failed += 1
    return {
        "kind": "txt_to_epub",
        "partial": True,
        "output_folder": str(output_folder),
        "report_path": str(report_path) if report_path.exists() else None,
        "output_files": output_files,
        "converted_count": converted,
        "failed_count": failed,
    }


def _read_json_file(path: Path) -> dict[str, object] | None:
    if not path.exists():
        return None
    try:
        return _loads_json_object(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def _loads_json_object(raw: str) -> dict[str, object]:
    if not raw:
        return {}
    payload = json.loads(raw)
    return payload if isinstance(payload, dict) else {}


def _string_list(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, str)]


def _scan_files(output_dir: Path, *, exclude_dirs: set[str]) -> list[str]:
    if not output_dir.exists():
        return []
    files: list[str] = []
    for path in sorted(output_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".txt", ".epub"}:
            continue
        if any(part in exclude_dirs for part in path.relative_to(output_dir).parts[:-1]):
            continue
        if path.name.startswith(("translation-statistics", "extraction-statistics")):
            continue
        files.append(str(path))
    return files


def _scan_bilingual_files(output_dir: Path) -> list[str]:
    if not output_dir.exists():
        return []
    files: list[str] = []
    for path in sorted(output_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".txt", ".epub"}:
            continue
        if path.parent == output_dir:
            continue
        files.append(str(path))
    return files


@dataclass
class TaskService:
    """Bridge-facing facade over the runtime + workflow orchestrators."""

    cache: TaskCache  # central task cache — <cache_root>/tasks/<task_id>/
    registry: TaskRegistry
    settings_store: SettingsStore
    profile_store: ModelProfileStore
    prompts_cache_root: Path
    llm_client_factory: LlmClientFactory

    def __post_init__(self) -> None:
        # Per-task in-memory mirrors. Populated by
        # ``_maybe_cleanup_cache`` right before it wipes a successfully
        # completed task's disk cache, so ``read_snapshot`` and
        # ``read_artifacts`` can still surface the run's final stats +
        # output paths until the user starts a new task or restarts
        # the app. The disk wipe is what the user sees on Finder; the
        # mirrors keep the Run page coherent without leaving cache
        # artifacts in the output folder.
        self._completed_snapshots: dict[str, TaskSnapshot] = {}
        self._completed_results: dict[str, dict[str, object]] = {}
        # Replacement runs that finish cleanly have their cache wiped,
        # which would also delete ``replacement-report.json``. Mirror
        # the report into memory right before the wipe so the modal
        # trigger keeps working until the user starts a new task or
        # restarts the app — same lifecycle as the snapshot mirror.
        self._completed_replacement_reports: dict[str, dict[str, object]] = {}
        # Per-kind start lock. Without this, two near-simultaneous
        # ``start_*`` calls (rapid double-click, re-run dialog firing
        # twice, frontend retry of a stuck request) race through
        # ``_purge_kind_for_start`` and end up with two threads
        # writing to the same output folder, corrupting the cache and
        # producing mixed-task artifacts. Held only for the seed +
        # spawn window, not for the full run, so single-threaded
        # callers see no extra latency.
        self._start_locks: dict[str, threading.Lock] = {
            "translation": threading.Lock(),
            "glossary": threading.Lock(),
            "glossary_review": threading.Lock(),
            "replacement": threading.Lock(),
            "epub_compress": threading.Lock(),
            "epub_merge": threading.Lock(),
            "epub_convert": threading.Lock(),
            "txt_to_epub": threading.Lock(),
        }
        self._retranslate_jobs: dict[str, RetranslateJob] = {}
        self._retranslate_lock = threading.Lock()
        self._retranslate_task_locks: dict[str, threading.Lock] = {}
        self._retranslate_task_indices: dict[str, _RetranslateTaskIndex] = {}
        self._retranslate_gate = threading.Condition()
        self._active_retranslates_by_model: dict[str, int] = {}
        self._retranslate_quality_gate = threading.Condition()
        self._retranslate_quality_groups: dict[
            str, _RetranslateQualityReviewGroup
        ] = {}

    def _retranslate_task_lock(self, task_id: str) -> threading.Lock:
        with self._retranslate_lock:
            return self._retranslate_task_locks.setdefault(task_id, threading.Lock())

    def _load_retranslate_snapshot(
        self, task_id: str, segment_ids: Sequence[str]
    ) -> TaskSnapshot:
        record = self.cache.load_record(task_id)
        index = self._retranslate_task_indices.get(task_id)
        if index is None or index.task_updated_at != record.updated_at:
            full_snapshot = self.cache.load(task_id)
            index = _build_retranslate_task_index(full_snapshot)
            self._retranslate_task_indices[task_id] = index
            available = {subtask.id: subtask for subtask in full_snapshot.subtasks}
        else:
            available = {}

        owner_ids: set[str] = set()
        for segment_id in segment_ids:
            owner_ids.update(index.owner_subtask_ids.get(segment_id, ()))
        if not owner_ids:
            return TaskSnapshot(record=record, subtasks=())

        try:
            subtasks = tuple(
                available.get(owner_id)
                or self.cache.load_subtask(task_id, owner_id)
                for owner_id in sorted(owner_ids)
            )
        except (TaskNotFoundError, OSError, ValueError):
            full_snapshot = self.cache.load(task_id)
            index = _build_retranslate_task_index(full_snapshot)
            self._retranslate_task_indices[task_id] = index
            owner_ids = {
                owner_id
                for segment_id in segment_ids
                for owner_id in index.owner_subtask_ids.get(segment_id, ())
            }
            subtasks = tuple(
                subtask
                for subtask in full_snapshot.subtasks
                if subtask.id in owner_ids
            )
            record = full_snapshot.record
        return TaskSnapshot(record=record, subtasks=subtasks)

    # All task records live under a single central root
    # (``<cache_root>/tasks/<task_id>/``). Records carry their kind
    # internally so list_tasks can filter; storing them flat means
    # output folder changes do not orphan caches and tasks survive
    # regardless of the user's current output_folder setting. Required
    # for "open app → resume last task" and the upcoming proofreading
    # feature, both of which need cache that outlives a clean run.

    def _cache_for_kind(self, kind: str) -> TaskCache:  # noqa: ARG002
        return self.cache

    def _cache_for_task(self, task_id: str) -> TaskCache:  # noqa: ARG002
        return self.cache

    def summarize_caches(self) -> dict[str, object]:
        """Aggregate stats for the cache cleanup UI."""

        records = self.cache.list_tasks()
        total_bytes = 0
        for record in records:
            task_dir = self.cache.task_dir(record.id)
            if not task_dir.exists():
                continue
            for path in task_dir.rglob("*"):
                if path.is_file():
                    try:
                        total_bytes += path.stat().st_size
                    except OSError:
                        continue
        return {
            "task_count": len(records),
            "total_bytes": total_bytes,
            "cache_root": str(self.cache.root),
        }

    def purge_caches(
        self, *, scope: str, days: int | None = None
    ) -> dict[str, object]:
        """Delete cache entries by scope.

        ``scope`` values:
          - ``"all"``                   — every entry (with active threads skipped)
          - ``"older_than_days"``       — entries with ``updated_at`` older than
                                          ``days`` days from now (active skipped)
          - ``"completed"``             — entries whose task status is completed

        Active in-flight tasks are never deleted regardless of scope —
        their thread would keep writing to a now-gone directory and
        corrupt the next start. The in-memory mirrors for any deleted
        task are dropped too so the Run page does not surface stale
        100% stats from the wiped run.
        """

        if scope not in {"all", "older_than_days", "completed"}:
            raise BridgeError.invalid_argument(
                f"unsupported purge scope: {scope!r}",
                field="scope",
                details={
                    "scope": scope,
                    "allowed": ["all", "older_than_days", "completed"],
                },
            )
        cutoff: datetime | None = None
        if scope == "older_than_days":
            if days is None or days < 0:
                raise BridgeError.invalid_argument(
                    "days must be a non-negative integer for older_than_days.",
                    field="days",
                )
            cutoff = datetime.now(timezone.utc) - timedelta(days=days)
        active_ids = self._active_task_ids()
        if active_ids:
            raise BridgeError.conflict(
                "cache cleanup is disabled while tasks are running.",
                details={"active_task_ids": active_ids},
            )

        removed: list[str] = []
        skipped_active: list[str] = []
        for record in self.cache.list_tasks():
            running = self._resolve_live_running(record.id, record.status)
            if running is not None and not running.is_done:
                skipped_active.append(record.id)
                continue
            if scope == "completed" and record.status is not TaskStatus.COMPLETED:
                continue
            if cutoff is not None:
                try:
                    updated = datetime.fromisoformat(record.updated_at)
                except ValueError:
                    continue
                if updated.tzinfo is None:
                    updated = updated.replace(tzinfo=timezone.utc)
                if updated > cutoff:
                    continue
            try:
                self.cache.delete(record.id)
            except TaskNotFoundError:
                continue
            removed.append(record.id)
            self._completed_snapshots.pop(record.id, None)
            self._completed_results.pop(record.id, None)
            self._completed_replacement_reports.pop(record.id, None)
        return {
            "scope": scope,
            "days": days,
            "removed_count": len(removed),
            "removed_ids": removed,
            "skipped_active_count": len(skipped_active),
        }

    def _active_task_ids(self) -> list[str]:
        active: list[str] = []
        for kind in _KIND_TO_TASKKIND:
            for running in self.registry.list_by_kind(kind):
                if running.is_done:
                    continue
                try:
                    record = self._cache_for_task(running.task_id).load_record(
                        running.task_id
                    )
                except (TaskNotFoundError, OSError, ValueError):
                    active.append(running.task_id)
                    continue
                if self._resolve_live_running(running.task_id, record.status) is None:
                    continue
                active.append(running.task_id)
        return sorted(set(active))

    def _maybe_cleanup_cache(self, kind: str, task_id: str) -> None:
        """Freeze final stats into in-memory mirrors for a clean run.

        ``Clean success`` = task status COMPLETED with zero failed
        subtasks. The on-disk cache is **no longer wiped** here — every
        task survives in ``<cache_root>/tasks/`` until the user clears
        it via the Settings cache cleanup UI. The in-memory mirrors
        below remain populated so ``read_snapshot`` and
        ``read_artifacts`` keep returning the run's final state without
        re-loading from disk on every poll.
        """

        cache = self._cache_for_kind(kind)
        try:
            snapshot = cache.load(task_id)
        except (TaskNotFoundError, OSError, ValueError):
            return
        if snapshot.record.status is not TaskStatus.COMPLETED:
            return
        elapsed_seconds = _task_elapsed_seconds(snapshot.record)
        progress = snapshot.progress(elapsed_seconds=elapsed_seconds)
        if progress.failed > 0:
            return

        usage = snapshot.usage()
        frozen = dict(snapshot.record.metadata)
        frozen["final_progress"] = {
            "total": progress.total,
            "pending": progress.pending,
            "running": progress.running,
            "completed": progress.completed,
            "failed": progress.failed,
            "skipped": progress.skipped,
            "elapsed_seconds": elapsed_seconds,
            "rate_per_second": progress.rate_per_second,
        }
        frozen["final_usage"] = {
            "input_tokens": usage.input_tokens,
            "output_tokens": usage.output_tokens,
            "cached_input_tokens": usage.cached_input_tokens,
        }
        frozen["final_low_confidence"] = _low_confidence_summary(snapshot, frozen)
        frozen_record = replace(
            snapshot.record,
            metadata=frozen,
            updated_at=_utc_now_iso(),
        )
        self._completed_snapshots[task_id] = TaskSnapshot(
            record=frozen_record, subtasks=()
        )
        result_payload = self._read_result(task_id)
        if result_payload is not None:
            self._completed_results[task_id] = result_payload
        report_path = self._replacement_report_path(task_id)
        if report_path.exists():
            try:
                self._completed_replacement_reports[task_id] = json.loads(
                    report_path.read_text(encoding="utf-8")
                )
            except (OSError, json.JSONDecodeError):
                pass

    def start_retranslate(
        self,
        *,
        task_id: str,
        segment_id: str,
        segment_ids: Sequence[str] | None = None,
        model_id: str | None = None,
        prompt_preset_id: str | None = None,
    ) -> dict[str, object]:
        requested_ids = list(dict.fromkeys(segment_ids or (segment_id,)))
        if not requested_ids or len(requested_ids) > 5:
            raise BridgeError.invalid_argument(
                "retranslate accepts between 1 and 5 segment ids.",
                field="segment_ids",
            )
        try:
            with self._retranslate_task_lock(task_id):
                snapshot = self._load_retranslate_snapshot(task_id, requested_ids)
        except TaskNotFoundError as exc:
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if snapshot.record.kind is not TaskKind.TRANSLATION:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} is not a translation task.",
                details={"task_id": task_id},
            )
        if snapshot.record.status in _ZOMBIE_TASK_STATES:
            running = self._resolve_live_running(task_id, snapshot.record.status)
            if running is not None and not running.is_done:
                raise BridgeError.conflict(
                    "cannot retranslate while the task is running.",
                    details={"task_id": task_id, "status": snapshot.record.status.value},
                )
            with self._retranslate_task_lock(task_id):
                snapshot = self._reconcile_zombie(
                    self.cache.load(task_id), self.cache
                )
                self._retranslate_task_indices[task_id] = (
                    _build_retranslate_task_index(snapshot)
                )
        if snapshot.record.status not in {
            TaskStatus.COMPLETED,
            TaskStatus.FAILED,
            TaskStatus.STOPPED,
        }:
            raise BridgeError.conflict(
                "retranslate is only available after translation stops or completes.",
                details={"task_id": task_id, "status": snapshot.record.status.value},
            )

        batch_items: list[dict[str, object]] = []
        for requested_id in requested_ids:
            seg_data = _find_segment_payload(snapshot, requested_id)
            if seg_data is None:
                raise BridgeError.not_found(
                    f"segment {requested_id!r} not found in task {task_id!r}.",
                    details={"task_id": task_id, "segment_id": requested_id},
                )
            original_dst = _read_segment_dst(snapshot, requested_id)
            batch_items.append(
                {
                    "segment_id": requested_id,
                    "original_dst": original_dst,
                    "source_hash": _hash_text(_retranslate_source_text(seg_data)),
                    "seg_data": dict(seg_data),
                }
            )
        primary = batch_items[0]
        segment_id = str(primary["segment_id"])
        original_dst = str(primary["original_dst"])
        seg_data = dict(primary["seg_data"])
        source_text = _retranslate_source_text(seg_data)
        settings = self.settings_store.load_all()
        effective_model_id = model_id or settings.app.active_translation_model_id
        effective_prompt_id = prompt_preset_id or settings.app.active_translation_prompt_id
        metadata = dict(snapshot.record.metadata)

        with self._retranslate_lock:
            for job in self._retranslate_jobs.values():
                if (
                    job.task_id == task_id
                    and set(requested_ids).intersection(
                        _retranslate_job_segment_ids(job)
                    )
                    and job.status in {"pending", "running"}
                ):
                    raise BridgeError.conflict(
                        "a retranslate job is already running for this segment.",
                        details={"request_id": job.request_id},
                    )
            request_id = f"retranslate-{uuid4().hex[:12]}"
            job = RetranslateJob(
                request_id=request_id,
                task_id=task_id,
                segment_id=segment_id,
                original_dst=original_dst,
                model_id=effective_model_id,
                prompt_preset_id=effective_prompt_id,
                status="pending",
                created_at=time.monotonic(),
                source_hash=_hash_text(source_text),
                original_dst_hash=_hash_text(original_dst),
                seg_data=dict(seg_data),
                metadata=metadata,
                model_snapshot=self._model_snapshot_for_retranslate(
                    effective_model_id
                ),
                prompt_snapshot=self._prompt_snapshot_for_retranslate(
                    effective_prompt_id, metadata
                ),
                batch_items=batch_items,
                batch_results=[],
                schema_version=3,
                created_at_wall=_utc_now_iso(),
                updated_at_wall=_utc_now_iso(),
            )
            self._retranslate_jobs[request_id] = job
            self._save_retranslate_job(job)
            self._gc_retranslate_jobs()

        self._start_retranslate_thread(request_id)
        return self._retranslate_status_payload(job)

    def read_retranslate_status(
        self, *, request_id: str
    ) -> dict[str, object]:
        with self._retranslate_lock:
            self._gc_retranslate_jobs()
            job = self._retranslate_jobs.get(request_id)
        if job is None:
            job = self._load_retranslate_job(request_id)
        if job is None:
            raise BridgeError.not_found(
                f"retranslate request {request_id!r} not found or expired.",
                details={"request_id": request_id},
            )
        self._sync_completed_retranslate_result(job)
        return self._retranslate_status_payload(job)

    def read_retranslate_statuses(
        self, *, request_ids: Sequence[str]
    ) -> dict[str, object]:
        return {
            "statuses": [
                self.read_retranslate_status(request_id=request_id)
                for request_id in dict.fromkeys(request_ids)
            ]
        }

    def resume_retranslate(self, *, request_id: str) -> dict[str, object]:
        with self._retranslate_lock:
            self._gc_retranslate_jobs()
            job = self._retranslate_jobs.get(request_id)
            live_job = job is not None
        if job is None:
            job = self._load_retranslate_job(request_id)
        if job is None:
            raise BridgeError.not_found(
                f"retranslate request {request_id!r} not found or expired.",
                details={"request_id": request_id},
            )
        if live_job and job.status in {"pending", "running"}:
            with self._retranslate_lock:
                self._retranslate_jobs[job.request_id] = job
            return self._retranslate_status_payload(job)
        if job.status in _RETRANSLATE_TERMINAL_STATUSES and job.status != "failed":
            return self._retranslate_status_payload(job)
        job.status = "pending"
        job.error = ""
        job.updated_at_wall = _utc_now_iso()
        with self._retranslate_lock:
            self._retranslate_jobs[job.request_id] = job
            self._save_retranslate_job(job)
            self._gc_retranslate_jobs()
        self._start_retranslate_thread(job.request_id)
        return self._retranslate_status_payload(job)

    def _validate_retranslate_request_id(self, request_id: str) -> str:
        if (
            not request_id.startswith("retranslate-")
            or "/" in request_id
            or "\\" in request_id
            or request_id in {".", ".."}
        ):
            raise BridgeError.invalid_argument(
                "invalid retranslate request id.",
                details={"request_id": request_id},
            )
        return request_id

    def _retranslate_job_path(self, task_id: str, request_id: str) -> Path:
        safe_request_id = self._validate_retranslate_request_id(request_id)
        return self.cache.task_dir(task_id) / "retranslate" / f"{safe_request_id}.json"

    def _find_retranslate_job_path(self, request_id: str) -> Path | None:
        safe_request_id = self._validate_retranslate_request_id(request_id)
        for path in self.cache.root.glob(f"*/retranslate/{safe_request_id}.json"):
            return path
        return None

    def _save_retranslate_job(self, job: RetranslateJob) -> None:
        path = self._retranslate_job_path(job.task_id, job.request_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".json.tmp")
        tmp.write_text(
            json.dumps(job.to_dict(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        os.replace(tmp, path)

    def _load_retranslate_job(self, request_id: str) -> RetranslateJob | None:
        path = self._find_retranslate_job_path(request_id)
        if path is None:
            return None
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError, ValueError):
            return None
        if not isinstance(data, Mapping):
            return None
        return RetranslateJob.from_dict(data)

    def _retranslate_status_payload(self, job: RetranslateJob) -> dict[str, object]:
        results = list(job.batch_results or [])
        if not results and job.status in _RETRANSLATE_TERMINAL_STATUSES:
            result: dict[str, object] = {
                "segment_id": job.segment_id,
                "status": job.status,
            }
            if job.result_dst:
                result["result_dst"] = job.result_dst
            if job.error:
                result["error"] = job.error
            results.append(result)
        created_at = _parse_iso_timestamp(job.created_at_wall)
        finished_at = (
            _parse_iso_timestamp(job.updated_at_wall)
            if job.status in _RETRANSLATE_TERMINAL_STATUSES
            else datetime.now(timezone.utc)
        )
        elapsed_seconds = (
            max(0.0, (finished_at - created_at).total_seconds())
            if created_at is not None and finished_at is not None
            else 0.0
        )
        return {
            "request_id": job.request_id,
            "task_id": job.task_id,
            "segment_id": job.segment_id,
            "status": job.status,
            "result_dst": job.result_dst,
            "error": job.error,
            "attempts": job.attempts,
            "last_error": job.last_error,
            "last_translation": job.last_translation,
            "results": results,
            "model_id": job.model_id or "",
            "segment_count": len(job.batch_items or ()) or 1,
            "created_at": job.created_at_wall,
            "updated_at": job.updated_at_wall,
            "elapsed_seconds": round(elapsed_seconds, 1),
        }

    def _sync_completed_retranslate_result(self, job: RetranslateJob) -> None:
        if job.status != "completed" or job.cache_applied:
            return
        completed_results = [
            result
            for result in job.batch_results or []
            if result.get("status") == "completed"
        ]
        segment_ids = tuple(
            str(result.get("segment_id", "")) for result in completed_results
        ) or (job.segment_id,)
        with self._retranslate_task_lock(job.task_id):
            try:
                snapshot = self._load_retranslate_snapshot(
                    job.task_id, segment_ids
                )
            except (TaskNotFoundError, OSError, ValueError):
                return
            updates: dict[str, str] = {}
            if job.batch_results:
                for result in completed_results:
                    segment_id = str(result.get("segment_id", ""))
                    original_dst = str(result.get("original_dst", ""))
                    if _read_segment_dst(snapshot, segment_id) == original_dst:
                        updates[segment_id] = str(result.get("result_dst", ""))
            elif job.result_dst:
                current_seg = _find_segment_payload(snapshot, job.segment_id)
                if current_seg is not None and not (
                    job.source_hash
                    and _hash_text(_retranslate_source_text(current_seg))
                    != job.source_hash
                ):
                    current_dst = _read_segment_dst(snapshot, job.segment_id)
                    if current_dst == job.original_dst:
                        updates[job.segment_id] = job.result_dst
            try:
                _patch_segment_dsts(self.cache, snapshot, updates)
            except (TaskNotFoundError, OSError):
                return
        job.cache_applied = True
        self._save_retranslate_job(job)

    def _model_snapshot_for_retranslate(
        self, model_id: str | None
    ) -> dict[str, object] | None:
        if not model_id:
            return None
        profile = self.profile_store.get(model_id)
        if profile is None:
            return None
        data = profile.to_dict()
        data["api_keys"] = []
        return data

    def _prompt_snapshot_for_retranslate(
        self,
        prompt_preset_id: str | None,
        metadata: Mapping[str, object],
    ) -> dict[str, object] | None:
        if prompt_preset_id:
            try:
                return self._resolve_prompt_preset(
                    prompt_preset_id,
                    kind=PromptKind.TRANSLATION,
                ).to_dict()
            except BridgeError:
                return None
        preset_data = metadata.get("prompt_preset")
        return dict(preset_data) if isinstance(preset_data, Mapping) else None

    def _start_retranslate_thread(self, request_id: str) -> None:
        thread = threading.Thread(
            target=self._run_retranslate,
            args=(request_id,),
            daemon=True,
        )
        thread.start()

    def _gc_retranslate_jobs(self) -> None:
        now = time.monotonic()
        expired = [
            rid
            for rid, job in self._retranslate_jobs.items()
            if job.status in _RETRANSLATE_TERMINAL_STATUSES
            and now - job.created_at > _RETRANSLATE_TERMINAL_TTL_SECONDS
        ]
        for rid in expired:
            self._retranslate_jobs.pop(rid, None)
        overflow = len(self._retranslate_jobs) - _MAX_RETRANSLATE_JOBS
        if overflow > 0:
            ordered = sorted(
                self._retranslate_jobs.items(), key=lambda kv: kv[1].created_at
            )
            for rid, _ in ordered[:overflow]:
                self._retranslate_jobs.pop(rid, None)

    def _run_retranslate(self, request_id: str) -> None:
        with self._retranslate_lock:
            job = self._retranslate_jobs.get(request_id)
        if job is None:
            job = self._load_retranslate_job(request_id)
        if job is None:
            return
        model_key, concurrency_limit = self._retranslate_concurrency(job)
        self._acquire_retranslate_slot(model_key, concurrency_limit)
        try:
            self._execute_retranslate(job)
        finally:
            self._release_retranslate_slot(model_key)

    def _execute_retranslate(self, job: RetranslateJob) -> None:
        if job.batch_items and len(job.batch_items) > 1:
            self._run_retranslate_batch(job)
            return
        if job.seg_data is None or job.metadata is None:
            job.error = "retranslate request is missing persisted source data."
            job.last_error = job.error
            job.status = "failed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        with self._retranslate_lock:
            job.status = "running"
            job.error = ""
            job.attempts += 1
            job.updated_at_wall = _utc_now_iso()
            self._retranslate_jobs[job.request_id] = job
            self._save_retranslate_job(job)
        if self._finish_retranslate_if_cache_changed(job):
            return
        try:
            with request_log_scope(
                self.cache,
                task_id=job.task_id,
                subtask_id=_retranslate_log_subtask_id(job.request_id),
                subtask_attempt=job.attempts,
                clock=_utc_now_iso,
            ):
                runner_result = asyncio.run(
                    self._call_runner_for_retranslate(
                        job.seg_data,
                        job.metadata,
                        model_id=job.model_id,
                        model_snapshot=job.model_snapshot,
                        prompt_preset_id=job.prompt_preset_id,
                        prompt_snapshot=job.prompt_snapshot,
                    )
                )
        except BridgeError as exc:
            job.error = f"{exc.code}: {exc.payload.message}"
            job.last_error = job.error
            job.status = "failed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        except Exception as exc:  # noqa: BLE001
            job.error = f"{type(exc).__name__}: {exc}"
            job.last_error = job.error
            job.status = "failed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return

        new_dst = runner_result.translations[job.segment_id]
        job.last_translation = new_dst
        try:
            with self._retranslate_task_lock(job.task_id):
                pre_review_snapshot = self._load_retranslate_snapshot(
                    job.task_id, (job.segment_id,)
                )
        except (TaskNotFoundError, OSError, ValueError):
            job.error = "task cache disappeared during retranslate."
            job.last_error = job.error
            job.status = "failed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        pre_review_segment = _find_segment_payload(
            pre_review_snapshot, job.segment_id
        )
        if pre_review_segment is None:
            job.error = "segment disappeared during retranslate."
            job.last_error = job.error
            job.status = "skipped"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        if (
            job.source_hash
            and _hash_text(_retranslate_source_text(pre_review_segment))
            != job.source_hash
        ):
            job.error = "source segment changed during retranslate."
            job.last_error = job.error
            job.status = "skipped"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        if _read_segment_dst(pre_review_snapshot, job.segment_id) != job.original_dst:
            job.status = "stale"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return
        source_text = _retranslate_source_text(job.seg_data)
        if not _is_preserved_nonprose_retranslation(source_text, new_dst):
            try:
                quality_decision = self._review_retranslation_candidate_batched(
                    source_text=source_text,
                    existing_dst=job.original_dst,
                    candidate_dst=new_dst,
                    metadata=job.metadata,
                    model_id=job.model_id,
                    model_snapshot=job.model_snapshot,
                )
            except Exception as exc:  # noqa: BLE001
                job.error = (
                    "quality review could not verify the new translation; "
                    f"kept the existing translation ({type(exc).__name__}: {exc})."
                )
                job.last_error = job.error
                job.status = "unresolved"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            if quality_decision.decision != "accept_new":
                job.error = (
                    "quality review kept the existing translation: "
                    f"{quality_decision.reason}"
                )
                job.last_error = job.error
                job.status = "unresolved"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return

        with self._retranslate_task_lock(job.task_id):
            try:
                snapshot = self._load_retranslate_snapshot(
                    job.task_id, (job.segment_id,)
                )
            except (TaskNotFoundError, OSError, ValueError):
                job.error = "task cache disappeared during retranslate."
                job.last_error = job.error
                job.status = "failed"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            current_seg = _find_segment_payload(snapshot, job.segment_id)
            if current_seg is None:
                job.error = "segment disappeared during retranslate."
                job.last_error = job.error
                job.status = "skipped"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            if (
                job.source_hash
                and _hash_text(_retranslate_source_text(current_seg)) != job.source_hash
            ):
                job.error = "source segment changed during retranslate."
                job.last_error = job.error
                job.status = "skipped"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            current_dst = _read_segment_dst(snapshot, job.segment_id)
            if current_dst != job.original_dst:
                job.status = "stale"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            try:
                _patch_segment_dst(self.cache, snapshot, job.segment_id, new_dst)
            except (TaskNotFoundError, OSError) as exc:
                job.error = f"failed to write cache: {exc}"
                job.last_error = job.error
                job.status = "failed"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return
            job.result_dst = new_dst
            job.cache_applied = True
            job.status = "completed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)

    def _review_retranslation_candidate_batched(
        self,
        *,
        source_text: str,
        existing_dst: str,
        candidate_dst: str,
        metadata: Mapping[str, object],
        model_id: str | None,
        model_snapshot: Mapping[str, object] | None,
    ) -> _RetranslateQualityDecision:
        settings = self.settings_store.load_all()
        effective_model_id = model_id or settings.app.active_translation_model_id
        snapshot = dict(model_snapshot) if model_snapshot is not None else None
        model = self._model_for_retranslate(
            effective_model_id,
            model_snapshot=snapshot,
            field=(
                "proofreading_model_id"
                if model_id
                else "active_translation_model_id"
            ),
        )
        group_key = _hash_text(
            json.dumps(
                {
                    "model_id": effective_model_id,
                    "model_snapshot": snapshot or {},
                },
                sort_keys=True,
                ensure_ascii=True,
            )
        )
        estimated_tokens = _estimate_dynamic_input_tokens(
            json.dumps(
                {
                    "source": source_text,
                    "existing_translation": existing_dst,
                    "new_candidate": candidate_dst,
                },
                ensure_ascii=False,
            )
        )
        item = _RetranslateQualityReviewItem(
            source_text=source_text,
            existing_dst=existing_dst,
            candidate_dst=candidate_dst,
            source_language=str(metadata.get("source_language", "")),
            target_language=str(metadata.get("target_language", "")),
            model_id=effective_model_id,
            model_snapshot=snapshot,
            estimated_tokens=estimated_tokens,
        )
        with self._retranslate_quality_gate:
            group = self._retranslate_quality_groups.get(group_key)
            is_leader = group is None
            if group is None:
                group = _RetranslateQualityReviewGroup()
                self._retranslate_quality_groups[group_key] = group
            group.items.append(item)
            self._retranslate_quality_gate.notify_all()

        if is_leader:
            self._drain_retranslate_quality_group(
                group_key,
                group,
                token_limit=max(0, int(model.input_token_limit)),
            )
        else:
            item.done.wait()
        if item.error is not None:
            raise item.error
        if item.result is None:
            raise RuntimeError("quality review completed without a decision")
        return item.result

    def _drain_retranslate_quality_group(
        self,
        group_key: str,
        group: _RetranslateQualityReviewGroup,
        *,
        token_limit: int,
    ) -> None:
        wait_for_more = True
        while True:
            with self._retranslate_quality_gate:
                if wait_for_more:
                    deadline = (
                        time.monotonic() + _RETRANSLATE_QUALITY_BATCH_WAIT_SECONDS
                    )
                    while (
                        len(group.items) < _RETRANSLATE_QUALITY_BATCH_MAX_ITEMS
                        and time.monotonic() < deadline
                    ):
                        self._retranslate_quality_gate.wait(
                            timeout=max(0.0, deadline - time.monotonic())
                        )
                batch: list[_RetranslateQualityReviewItem] = []
                batch_tokens = 0
                while (
                    group.items
                    and len(batch) < _RETRANSLATE_QUALITY_BATCH_MAX_ITEMS
                ):
                    next_item = group.items[0]
                    if (
                        batch
                        and token_limit > 0
                        and batch_tokens + next_item.estimated_tokens > token_limit
                    ):
                        break
                    batch.append(group.items.pop(0))
                    batch_tokens += next_item.estimated_tokens

            try:
                if len(batch) == 1:
                    only = batch[0]
                    decisions = {
                        "0": asyncio.run(
                            self._review_single_retranslation_candidate(
                                source_text=only.source_text,
                                existing_dst=only.existing_dst,
                                candidate_dst=only.candidate_dst,
                                metadata={
                                    "source_language": only.source_language,
                                    "target_language": only.target_language,
                                },
                                model_id=only.model_id,
                                model_snapshot=only.model_snapshot,
                            )
                        )
                    }
                else:
                    decisions = asyncio.run(
                        self._review_retranslation_candidates_batch(batch)
                    )
            except Exception as exc:  # noqa: BLE001
                for queued_item in batch:
                    queued_item.error = exc
            else:
                for item_index, queued_item in enumerate(batch):
                    queued_item.result = decisions[str(item_index)]
            finally:
                for queued_item in batch:
                    queued_item.done.set()

            with self._retranslate_quality_gate:
                if not group.items:
                    self._retranslate_quality_groups.pop(group_key, None)
                    return
            wait_for_more = False

    async def _review_retranslation_candidates_batch(
        self,
        items: Sequence[_RetranslateQualityReviewItem],
    ) -> dict[str, _RetranslateQualityDecision]:
        if not items:
            return {}
        first = items[0]
        model = self._model_for_retranslate(
            first.model_id,
            model_snapshot=first.model_snapshot,
            field="proofreading_model_id",
        )
        item_payloads = [
            {
                "id": str(index),
                "source_language": item.source_language,
                "target_language": item.target_language,
                "source": item.source_text,
                "existing_translation": item.existing_dst,
                "new_candidate": item.candidate_dst,
            }
            for index, item in enumerate(items)
        ]
        system_prompt = (
            _RETRANSLATE_QUALITY_COMPARATOR_PROMPT
            + " Review every item independently. Return exactly one JSON object "
            'matching {"decisions":{"<id>":{"decision":"accept_new|'
            'keep_existing|uncertain","reason":"..."}}}. Return every supplied '
            "id exactly once and no additional ids."
        )
        client = self.llm_client_factory()
        try:
            response = await client.chat(
                ChatRequest(
                    model=model,
                    system_prompt=system_prompt,
                    user_prompt=json.dumps(
                        {"items": item_payloads}, ensure_ascii=False
                    ),
                    log_label="proofreading retranslation quality review batch",
                )
            )
        finally:
            await client.aclose()
        return _decode_retranslate_quality_decisions(
            response.content,
            tuple(str(index) for index in range(len(items))),
        )

    async def _review_single_retranslation_candidate(
        self,
        *,
        source_text: str,
        existing_dst: str,
        candidate_dst: str,
        metadata: Mapping[str, object],
        model_id: str | None,
        model_snapshot: Mapping[str, object] | None,
    ) -> _RetranslateQualityDecision:
        settings = self.settings_store.load_all()
        model = self._model_for_retranslate(
            model_id or settings.app.active_translation_model_id,
            model_snapshot=model_snapshot,
            field=(
                "proofreading_model_id"
                if model_id
                else "active_translation_model_id"
            ),
        )
        source_language = str(metadata.get("source_language", ""))
        target_language = str(metadata.get("target_language", ""))
        system_prompt = (
            _RETRANSLATE_QUALITY_COMPARATOR_PROMPT
            + " Return exactly one JSON object with keys "
            '"decision" and "reason". The decision must be one of '
            '"accept_new", "keep_existing", or "uncertain".'
        )
        user_prompt = json.dumps(
            {
                "source_language": source_language,
                "target_language": target_language,
                "source": source_text,
                "existing_translation": existing_dst,
                "new_candidate": candidate_dst,
            },
            ensure_ascii=False,
        )
        client = self.llm_client_factory()
        try:
            response = await client.chat(
                ChatRequest(
                    model=model,
                    system_prompt=system_prompt,
                    user_prompt=user_prompt,
                    log_label="proofreading single retranslation quality review",
                )
            )
        finally:
            await client.aclose()
        return _decode_retranslate_quality_decision(response.content)

    def _retranslate_concurrency(self, job: RetranslateJob) -> tuple[str, int]:
        snapshot = job.model_snapshot
        if isinstance(snapshot, Mapping):
            try:
                model = ModelConfig.from_dict(snapshot)
            except (TypeError, ValueError):
                pass
            else:
                return model.id, effective_concurrency_limit(model)
        return job.model_id or "unknown", 1

    def _acquire_retranslate_slot(self, model_key: str, limit: int) -> None:
        with self._retranslate_gate:
            while self._active_retranslates_by_model.get(model_key, 0) >= limit:
                self._retranslate_gate.wait()
            self._active_retranslates_by_model[model_key] = (
                self._active_retranslates_by_model.get(model_key, 0) + 1
            )

    def _release_retranslate_slot(self, model_key: str) -> None:
        with self._retranslate_gate:
            active = self._active_retranslates_by_model.get(model_key, 0) - 1
            if active > 0:
                self._active_retranslates_by_model[model_key] = active
            else:
                self._active_retranslates_by_model.pop(model_key, None)
            self._retranslate_gate.notify_all()

    def _run_retranslate_batch(self, job: RetranslateJob) -> None:
        items = list(job.batch_items or [])
        if not items or job.metadata is None:
            job.error = "retranslate batch is missing persisted source data."
            job.last_error = job.error
            job.status = "failed"
            self._save_retranslate_job(job)
            return
        with self._retranslate_lock:
            job.status = "running"
            job.error = ""
            job.attempts += 1
            job.updated_at_wall = _utc_now_iso()
            self._retranslate_jobs[job.request_id] = job
            self._save_retranslate_job(job)

        ready: list[dict[str, object]] = []
        results: list[dict[str, object]] = []
        with self._retranslate_task_lock(job.task_id):
            try:
                snapshot = self._load_retranslate_snapshot(
                    job.task_id,
                    tuple(str(item.get("segment_id", "")) for item in items),
                )
            except (TaskNotFoundError, OSError, ValueError):
                job.error = "task cache disappeared during retranslate."
                job.last_error = job.error
                job.status = "failed"
                self._save_retranslate_job(job)
                return
        for item in items:
            segment_id = str(item.get("segment_id", ""))
            seg_data = item.get("seg_data")
            original_dst = str(item.get("original_dst", ""))
            current_seg = _find_segment_payload(snapshot, segment_id)
            if not isinstance(seg_data, Mapping) or current_seg is None:
                results.append(
                    {
                        "segment_id": segment_id,
                        "status": "skipped",
                        "error": "segment disappeared during retranslate.",
                    }
                )
                continue
            if str(item.get("source_hash", "")) and _hash_text(
                _retranslate_source_text(current_seg)
            ) != str(item.get("source_hash", "")):
                results.append(
                    {
                        "segment_id": segment_id,
                        "status": "skipped",
                        "error": "source segment changed during retranslate.",
                    }
                )
                continue
            if _read_segment_dst(snapshot, segment_id) != original_dst:
                results.append({"segment_id": segment_id, "status": "stale"})
                continue
            ready.append(item)

        quality_decisions: dict[str, _RetranslateQualityDecision] = {}
        quality_review_error = ""
        try:
            if ready:
                with request_log_scope(
                    self.cache,
                    task_id=job.task_id,
                    subtask_id=_retranslate_log_subtask_id(job.request_id),
                    subtask_attempt=job.attempts,
                    clock=_utc_now_iso,
                ):
                    runner_result = asyncio.run(
                        self._call_runner_for_retranslate_batch(
                            tuple(dict(item["seg_data"]) for item in ready),
                            job.metadata,
                            model_id=job.model_id,
                            model_snapshot=job.model_snapshot,
                            prompt_preset_id=job.prompt_preset_id,
                            prompt_snapshot=job.prompt_snapshot,
                        )
                    )
                    quality_decisions, quality_review_error = (
                        self._review_retranslate_batch_title_candidates(
                            job,
                            ready,
                            runner_result,
                        )
                    )
            else:
                runner_result = _RetranslateRunnerResult({}, {})
        except Exception as exc:  # noqa: BLE001
            job.error = (
                f"{exc.code}: {exc.payload.message}"
                if isinstance(exc, BridgeError)
                else f"{type(exc).__name__}: {exc}"
            )
            job.last_error = job.error
            results.extend(
                {
                    "segment_id": str(item.get("segment_id", "")),
                    "status": "failed",
                    "error": job.error,
                }
                for item in ready
            )
            job.batch_results = results
            job.status = "failed"
            job.updated_at_wall = _utc_now_iso()
            self._save_retranslate_job(job)
            return

        results.extend(
            self._apply_retranslate_batch_candidates(
                job,
                ready,
                runner_result,
                quality_decisions=quality_decisions,
                quality_review_error=quality_review_error,
            )
        )

        job.batch_results = results
        job.cache_applied = True
        primary = next(
            (item for item in results if item.get("segment_id") == job.segment_id),
            None,
        )
        if primary and primary.get("status") == "completed":
            job.result_dst = str(primary.get("result_dst", ""))
            job.last_translation = job.result_dst
        job.status = "completed"
        job.updated_at_wall = _utc_now_iso()
        self._save_retranslate_job(job)

    def _review_retranslate_batch_title_candidates(
        self,
        job: RetranslateJob,
        items: Sequence[Mapping[str, object]],
        runner_result: _RetranslateRunnerResult,
    ) -> tuple[dict[str, _RetranslateQualityDecision], str]:
        review_segment_ids: list[str] = []
        review_items: list[_RetranslateQualityReviewItem] = []
        metadata = job.metadata or {}
        for item in items:
            segment_id = str(item.get("segment_id", ""))
            if segment_id not in runner_result.quality_review_segments:
                continue
            seg_data = item.get("seg_data")
            if not isinstance(seg_data, Mapping):
                continue
            source_text = _retranslate_source_text(seg_data)
            existing_dst = str(item.get("original_dst", ""))
            candidate_dst = runner_result.translations[segment_id]
            review_segment_ids.append(segment_id)
            review_items.append(
                _RetranslateQualityReviewItem(
                    source_text=source_text,
                    existing_dst=existing_dst,
                    candidate_dst=candidate_dst,
                    source_language=str(metadata.get("source_language", "")),
                    target_language=str(metadata.get("target_language", "")),
                    model_id=job.model_id,
                    model_snapshot=(
                        dict(job.model_snapshot)
                        if job.model_snapshot is not None
                        else None
                    ),
                    estimated_tokens=_estimate_dynamic_input_tokens(
                        json.dumps(
                            {
                                "source": source_text,
                                "existing_translation": existing_dst,
                                "new_candidate": candidate_dst,
                            },
                            ensure_ascii=False,
                        )
                    ),
                )
            )
        if not review_items:
            return {}, ""
        try:
            indexed = asyncio.run(
                self._review_retranslation_candidates_batch(review_items)
            )
        except Exception as exc:  # noqa: BLE001
            return {}, f"{type(exc).__name__}: {exc}"
        return {
            segment_id: indexed[str(index)]
            for index, segment_id in enumerate(review_segment_ids)
        }, ""

    def _apply_retranslate_batch_candidates(
        self,
        job: RetranslateJob,
        items: Sequence[Mapping[str, object]],
        runner_result: _RetranslateRunnerResult,
        *,
        quality_decisions: Mapping[str, _RetranslateQualityDecision],
        quality_review_error: str,
    ) -> list[dict[str, object]]:
        results: list[dict[str, object]] = []
        updates: dict[str, str] = {}
        with self._retranslate_task_lock(job.task_id):
            try:
                snapshot = self._load_retranslate_snapshot(
                    job.task_id,
                    tuple(str(item.get("segment_id", "")) for item in items),
                )
            except (TaskNotFoundError, OSError, ValueError) as exc:
                return [
                    {
                        "segment_id": str(item.get("segment_id", "")),
                        "status": "failed",
                        "error": f"failed to read cache: {exc}",
                    }
                    for item in items
                ]
            for item in items:
                segment_id = str(item["segment_id"])
                original_dst = str(item.get("original_dst", ""))
                new_dst = runner_result.translations[segment_id]
                confidence_entry = runner_result.low_confidence.get(segment_id)
                current_seg = _find_segment_payload(snapshot, segment_id)
                if current_seg is None or (
                    str(item.get("source_hash", ""))
                    and _hash_text(_retranslate_source_text(current_seg))
                    != str(item.get("source_hash", ""))
                ):
                    results.append(
                        {
                            "segment_id": segment_id,
                            "status": "skipped",
                            "error": "source segment changed during retranslate.",
                        }
                    )
                    continue
                if _read_segment_dst(snapshot, segment_id) != original_dst:
                    results.append({"segment_id": segment_id, "status": "stale"})
                    continue
                if segment_id in runner_result.quality_review_segments:
                    decision = quality_decisions.get(segment_id)
                    if quality_review_error:
                        results.append(
                            {
                                "segment_id": segment_id,
                                "status": "unresolved",
                                "result_dst": new_dst,
                                "error": (
                                    "quality review could not verify the new "
                                    "title candidate; kept the existing "
                                    f"translation ({quality_review_error})."
                                ),
                            }
                        )
                        continue
                    if decision is None or decision.decision != "accept_new":
                        reason = (
                            decision.reason
                            if decision is not None
                            else "quality review returned no decision"
                        )
                        results.append(
                            {
                                "segment_id": segment_id,
                                "status": "unresolved",
                                "result_dst": new_dst,
                                "error": (
                                    "quality review did not approve the new "
                                    f"title candidate: {reason}"
                                ),
                            }
                        )
                        continue
                if _should_keep_existing_retranslation(
                    snapshot,
                    segment_id,
                    original_dst,
                    new_dst,
                    confidence_entry,
                ):
                    results.append(
                        {
                            "segment_id": segment_id,
                            "status": "unresolved",
                            "result_dst": new_dst,
                            "error": "retranslation did not improve source-language residue.",
                            "reasons": list(
                                (confidence_entry or {}).get("reasons", [])
                            ),
                            "tags": list((confidence_entry or {}).get("tags", [])),
                        }
                    )
                    continue
                updates[segment_id] = new_dst
                results.append(
                    {
                        "segment_id": segment_id,
                        "status": "completed",
                        "result_dst": new_dst,
                        "original_dst": original_dst,
                    }
                )
            try:
                _patch_segment_dsts(self.cache, snapshot, updates)
            except (TaskNotFoundError, OSError) as exc:
                for result in results:
                    segment_id = str(result.get("segment_id", ""))
                    if segment_id not in updates:
                        continue
                    result.clear()
                    result.update(
                        {
                            "segment_id": segment_id,
                            "status": "failed",
                            "error": f"failed to write cache: {exc}",
                        }
                    )
        return results

    def _finish_retranslate_if_cache_changed(self, job: RetranslateJob) -> bool:
        with self._retranslate_task_lock(job.task_id):
            try:
                snapshot = self._load_retranslate_snapshot(
                    job.task_id, (job.segment_id,)
                )
            except (TaskNotFoundError, OSError, ValueError):
                job.error = "task cache disappeared during retranslate."
                job.last_error = job.error
                job.status = "failed"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return True
            current_seg = _find_segment_payload(snapshot, job.segment_id)
            if current_seg is None:
                job.error = "segment disappeared during retranslate."
                job.last_error = job.error
                job.status = "skipped"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return True
            if (
                job.source_hash
                and _hash_text(_retranslate_source_text(current_seg)) != job.source_hash
            ):
                job.error = "source segment changed during retranslate."
                job.last_error = job.error
                job.status = "skipped"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return True
            if _read_segment_dst(snapshot, job.segment_id) != job.original_dst:
                job.status = "stale"
                job.updated_at_wall = _utc_now_iso()
                self._save_retranslate_job(job)
                return True
            return False

    async def _call_runner_for_retranslate(
        self,
        seg_data: Mapping[str, object],
        metadata: Mapping[str, object],
        *,
        model_id: str | None = None,
        model_snapshot: Mapping[str, object] | None = None,
        prompt_preset_id: str | None = None,
        prompt_snapshot: Mapping[str, object] | None = None,
    ) -> _RetranslateRunnerResult:
        return await self._call_runner_for_retranslate_batch(
            (seg_data,),
            metadata,
            model_id=model_id,
            model_snapshot=model_snapshot,
            prompt_preset_id=prompt_preset_id,
            prompt_snapshot=prompt_snapshot,
        )

    async def _call_runner_for_retranslate_batch(
        self,
        seg_datas: Sequence[Mapping[str, object]],
        metadata: Mapping[str, object],
        *,
        model_id: str | None = None,
        model_snapshot: Mapping[str, object] | None = None,
        prompt_preset_id: str | None = None,
        prompt_snapshot: Mapping[str, object] | None = None,
    ) -> _RetranslateRunnerResult:
        from transoria.workflows.translation.chunker import (  # noqa: PLC0415
            ChunkSegment,
            TranslationChunk,
        )
        from transoria.workflows.translation.rules import (  # noqa: PLC0415
            Glossary,
            ReplacementRule,
        )
        from transoria.workflows.translation.runner import (  # noqa: PLC0415
            TranslationSubtaskRunner,
            encode_subtask_payload,
            is_korean_latin_title_candidate,
            split_segment_payload_batches,
        )

        settings = self.settings_store.load_all()
        model = self._model_for_retranslate(
            model_id or settings.app.active_translation_model_id,
            model_snapshot=model_snapshot,
            field=(
                "proofreading_model_id"
                if model_id
                else "active_translation_model_id"
            ),
        )

        if prompt_snapshot is not None:
            try:
                preset = PromptPreset.from_dict(prompt_snapshot)
            except ValueError as exc:
                raise BridgeError.invalid_argument(
                    f"persisted prompt snapshot is invalid: {exc}",
                ) from exc
        elif prompt_preset_id:
            preset = self._resolve_prompt_preset(
                prompt_preset_id, kind=PromptKind.TRANSLATION
            )
        elif settings.app.active_translation_prompt_id:
            preset = self._resolve_prompt_preset(
                settings.app.active_translation_prompt_id,
                kind=PromptKind.TRANSLATION,
            )
        else:
            preset_data = metadata.get("prompt_preset")
            if not isinstance(preset_data, Mapping):
                raise BridgeError.invalid_argument(
                    "task metadata is missing prompt_preset snapshot (cache predates B.5.1).",
                )
            preset = PromptPreset.from_dict(preset_data)

        try:
            source_language = Language(str(metadata.get("source_language", "")))
            target_language = Language(str(metadata.get("target_language", "")))
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                f"task metadata has invalid language: {exc}",
            ) from exc

        glossary_records = metadata.get("glossary", [])
        glossary = (
            Glossary.from_records(glossary_records)
            if isinstance(glossary_records, list)
            else Glossary.empty()
        )
        post_records = metadata.get("post_replacements", [])
        post_replacements: tuple[ReplacementRule, ...] = (
            tuple(
                ReplacementRule(
                    src=str(r.get("src", "")),
                    dst=str(r.get("dst", "")),
                    regex=bool(r.get("regex", False)),
                    case_sensitive=bool(r.get("case_sensitive", False)),
                    note=str(r.get("note", "")),
                    enabled=bool(r.get("enabled", True)),
                )
                for r in post_records
                if isinstance(r, Mapping)
            )
            if isinstance(post_records, list)
            else ()
        )

        runner = TranslationSubtaskRunner(
            client=self.llm_client_factory(),
            model=model,
            prompt_preset=preset,
            source_language=source_language,
            target_language=target_language,
            post_replacements=post_replacements,
            enable_confidence_check=True,
            stream=True,
            low_confidence_max_retries=max(
                0, int(settings.translation.low_confidence_max_retries)
            ),
            transport_retry_attempts=max(
                0, int(settings.translation.request_retry_attempts)
            ),
            preserve_korean_latin_title_candidates=True,
        )
        translations: dict[str, str] = {}
        low_confidence: dict[str, dict[str, list[str]]] = {}
        try:
            for batch_index, batch in enumerate(
                split_segment_payload_batches(seg_datas), start=1
            ):
                chunk_segments = []
                segment_metadata = []
                source_texts = []
                for chunk_index, seg_data in enumerate(batch):
                    segment_id = str(seg_data["segment_id"])
                    original_text = str(seg_data.get("original_text", ""))
                    prompt_text = str(
                        seg_data.get("prompt_text", original_text)
                    )
                    source_texts.append(original_text)
                    chunk_segments.append(
                        ChunkSegment(
                            segment_id=segment_id,
                            chunk_index=chunk_index,
                            prompt_text=prompt_text,
                        )
                    )
                    segment_metadata.append(
                        {
                            "original_text": original_text,
                            "protection_spans": list(
                                seg_data.get("protection_spans", [])
                            ),
                            "leading_whitespace": str(
                                seg_data.get("leading_whitespace", "")
                            ),
                            "trailing_whitespace": str(
                                seg_data.get("trailing_whitespace", "")
                            ),
                        }
                    )
                chunk = TranslationChunk(
                    segments=tuple(chunk_segments),
                    context_lines=(),
                    glossary_entries=glossary.match_many(source_texts),
                )
                payload = encode_subtask_payload(
                    chunk, segment_metadata=segment_metadata
                )
                subtask = Subtask(
                    id=f"retranslate-batch-{batch_index:03d}",
                    task_id="retranslate-virtual",
                    request_payload=payload,
                )
                result = await runner.run(subtask)
                try:
                    response = json.loads(result.response_content)
                except json.JSONDecodeError as exc:
                    raise BridgeError(
                        "bridge.io_error",
                        f"runner returned invalid JSON: {exc}",
                        retryable=True,
                    ) from exc
                records = (
                    response.get("translations", {})
                    if isinstance(response, Mapping)
                    else {}
                )
                if isinstance(records, Mapping):
                    translations.update(
                        (str(segment_id), str(text))
                        for segment_id, text in records.items()
                    )
                if isinstance(response, Mapping):
                    low_confidence.update(low_confidence_by_segment(response))
        finally:
            await runner.client.aclose()
        expected = {str(seg_data["segment_id"]) for seg_data in seg_datas}
        missing = expected - set(translations)
        if missing:
            raise BridgeError(
                "bridge.io_error",
                f"runner returned no translation for segments: {sorted(missing)!r}.",
                retryable=True,
            )
        quality_review_segments = frozenset(
            segment_id
            for seg_data in seg_datas
            if (
                (segment_id := str(seg_data["segment_id"])) in translations
                and is_korean_latin_title_candidate(
                    str(seg_data.get("original_text", "")),
                    translations[segment_id],
                    source_language=source_language,
                    target_language=target_language,
                )
            )
        )
        return _RetranslateRunnerResult(
            translations,
            low_confidence,
            quality_review_segments,
        )

    def _model_for_retranslate(
        self,
        profile_id: str | None,
        *,
        model_snapshot: Mapping[str, object] | None,
        field: str,
    ) -> ModelConfig:
        if model_snapshot:
            try:
                snapshot_model = ModelConfig.from_dict(model_snapshot)
            except ValueError as exc:
                raise BridgeError.invalid_argument(
                    f"persisted model snapshot is invalid: {exc}",
                    field=field,
                ) from exc

            key_profile = self.profile_store.get(snapshot_model.id)
            if key_profile is None:
                raise BridgeError.invalid_argument(
                    f"model profile {snapshot_model.id!r} from the persisted "
                    "retranslate snapshot no longer exists.",
                    field=field,
                )
            if not key_profile.api_keys:
                raise BridgeError.invalid_argument(
                    f"model profile {snapshot_model.id!r} from the persisted "
                    "retranslate snapshot has no API key configured.",
                    field=field,
                )
            return snapshot_model.with_api_keys(tuple(key_profile.api_keys))

        return self._resolve_model_profile(profile_id, field=field)

    def _build_translation_config(
        self,
    ) -> tuple[TranslationConfig, ModelConfig, PromptPreset]:
        settings = self.settings_store.load_all()
        translation = settings.translation
        app = settings.app

        input_dir = _require_directory(translation.input_folder, field="input_folder")
        _require_input_with_supported_files(input_dir, field="input_folder")
        output_dir = _ensure_output_dir(
            translation.output_folder, field="output_folder"
        )
        _require_distinct_translation_folders(input_dir, output_dir)
        source_lang = _coerce_language(
            translation.source_language, field="source_language"
        )
        target_lang = _coerce_language(
            translation.target_language, field="target_language"
        )
        model = self._resolve_model_profile(
            app.active_translation_model_id, field="active_translation_model_id"
        )
        # Per-task timeout overrides any value persisted on the model
        # profile — this knob lives in the translation settings UI now.
        model = replace(model, timeout_seconds=float(translation.timeout_seconds))
        preset = self._resolve_prompt_preset(
            app.active_translation_prompt_id, kind=PromptKind.TRANSLATION
        )

        glossary = Glossary.from_records(translation.translation_glossary)
        text_preserve_rules = _coerce_text_preserve_rules(
            translation.text_preserve_rules
        )
        pre_replacements = _coerce_translation_replacements(
            translation.pre_replacements
        )
        post_replacements = _coerce_translation_replacements(
            translation.post_replacements
        )

        config = TranslationConfig(
            input_dir=input_dir,
            output_dir=output_dir,
            source_language=source_lang,
            target_language=target_lang,
            model=model,
            prompt_preset=preset,
            glossary=glossary,
            text_preserve_rules=text_preserve_rules,
            pre_replacements=pre_replacements,
            post_replacements=post_replacements,
            bilingual_enabled=translation.bilingual_enabled,
            bilingual_dedup_when_same=translation.bilingual_dedupe_identical,
            bilingual_subfolder=(
                translation.bilingual_subfolder_name or "bilingual outputs"
            ),
            context_line_count=max(0, int(translation.context_lines)),
            chunk_size=_derive_chunk_size(model.input_token_limit),
            dynamic_input_token_limit=max(0, int(model.input_token_limit)),
            dynamic_input_token_counter=_estimate_dynamic_input_tokens,
            low_confidence_max_retries=max(
                0, int(translation.low_confidence_max_retries)
            ),
            request_retry_attempts=max(0, int(translation.request_retry_attempts)),
        )
        return config, model, preset

    def start_translation(self, request_id: str) -> dict[str, object]:
        with self._start_locks["translation"]:
            return self._start_translation_locked(request_id)

    def _start_translation_locked(self, request_id: str) -> dict[str, object]:
        config, model, preset = self._build_translation_config()
        input_dir = config.input_dir
        output_dir = config.output_dir
        source_lang = config.source_language
        target_lang = config.target_language
        self._purge_kind_for_start(
            kind="translation", task_kind=TaskKind.TRANSLATION
        )

        task_id = _new_task_id("translation")
        started_at = _utc_now_iso()
        self._seed_placeholder(
            task_id,
            kind=TaskKind.TRANSLATION,
            started_at=started_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
                "source_language": source_lang.value,
                "target_language": target_lang.value,
                "model_id": model.id,
                "prompt_preset_id": preset.id,
                "timeout_seconds": float(model.timeout_seconds),
                "request_id": request_id,
            },
        )

        cache = self._cache_for_kind("translation")
        running = RunningTask(
            task_id=task_id,
            kind="translation",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._translation_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {
            "task_id": task_id,
            "started_at": started_at,
        }

    async def _translation_thread(
        self,
        task_id: str,
        config: TranslationConfig,
        running: RunningTask,
    ) -> None:
        client = self.llm_client_factory()
        cache = self._cache_for_kind("translation")
        config = replace(config, debug_log_dir=cache.task_dir(task_id) / "debug")

        def _capture(executor: TaskExecutor) -> None:
            running.set_executor(executor)

        def _touch_progress(_event: object) -> None:
            running.touch()

        def _finalize(result: TranslationRunResult) -> None:
            payload = _translation_result_payload(result, config=config)
            self._write_result(task_id, payload)

        try:
            orchestrator = TranslationOrchestrator(
                cache=cache,
                client=client,
                id_factory=lambda: task_id,
                progress=_touch_progress,
                on_executor_created=_capture,
                on_result_finalized=_finalize,
            )
            await orchestrator.run(config)
            self._maybe_cleanup_cache("translation", task_id)
        finally:
            await client.aclose()

    def _build_glossary_config(
        self,
    ) -> tuple[GlossaryConfig, ModelConfig, PromptPreset]:
        settings = self.settings_store.load_all()
        glossary = settings.glossary
        app = settings.app

        input_dir = _require_directory(glossary.input_folder, field="input_folder")
        _require_input_with_supported_files(input_dir, field="input_folder")
        output_dir = _ensure_output_dir(
            glossary.output_folder, field="output_folder"
        )
        source_lang = _coerce_language(
            glossary.source_language, field="source_language"
        )
        target_lang = _coerce_language(
            glossary.target_language, field="target_language"
        )
        model = self._resolve_model_profile(
            app.active_glossary_model_id, field="active_glossary_model_id"
        )
        # Per-task timeout overrides any value persisted on the model
        # profile — this knob lives in the glossary settings UI now.
        model = replace(model, timeout_seconds=float(glossary.timeout_seconds))
        preset = self._resolve_prompt_preset(
            app.active_glossary_prompt_id, kind=PromptKind.GLOSSARY
        )

        config = GlossaryConfig(
            input_dir=input_dir,
            output_dir=output_dir,
            source_language=source_lang,
            target_language=target_lang,
            model=model,
            prompt_preset=preset,
            reference_example_limit=max(
                0, int(glossary.reference_examples_per_term)
            ),
            max_term_display_length=max(1, int(glossary.max_term_display_length)),
            min_frequency=max(1, int(glossary.minimum_frequency)),
            chunk_token_limit=_effective_glossary_chunk_token_limit(
                int(glossary.chunk_token_limit)
            ),
            allow_src_eq_dst=bool(glossary.keep_identical_src_dst),
            combine_folder_glossary=bool(glossary.merge_folder_glossary),
            normalize_widths=bool(glossary.normalize_widths),
            novel_background=str(glossary.novel_background or ""),
            request_retry_attempts=max(0, int(glossary.request_retry_attempts)),
        )
        return config, model, preset

    def start_glossary(self, request_id: str) -> dict[str, object]:
        with self._start_locks["glossary"]:
            return self._start_glossary_locked(request_id)

    def _start_glossary_locked(self, request_id: str) -> dict[str, object]:
        config, model, preset = self._build_glossary_config()
        input_dir = config.input_dir
        output_dir = config.output_dir
        source_lang = config.source_language
        target_lang = config.target_language

        self._purge_kind_for_start(
            kind="glossary", task_kind=TaskKind.GLOSSARY
        )

        task_id = _new_task_id("glossary")
        started_at = _utc_now_iso()
        self._seed_placeholder(
            task_id,
            kind=TaskKind.GLOSSARY,
            started_at=started_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
                "source_language": source_lang.value,
                "target_language": target_lang.value,
                "model_id": model.id,
                "prompt_preset_id": preset.id,
                "timeout_seconds": float(model.timeout_seconds),
                "request_id": request_id,
            },
        )

        cache = self._cache_for_kind("glossary")
        running = RunningTask(
            task_id=task_id,
            kind="glossary",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._glossary_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    async def _glossary_thread(
        self,
        task_id: str,
        config: GlossaryConfig,
        running: RunningTask,
    ) -> None:
        client = self.llm_client_factory()
        cache = self._cache_for_kind("glossary")
        config = replace(config, debug_log_dir=cache.task_dir(task_id) / "debug")

        def _capture(executor: TaskExecutor) -> None:
            running.set_executor(executor)

        def _touch_progress(_event: object) -> None:
            running.touch()

        def _finalize(result: GlossaryExtractionResult) -> None:
            payload = _glossary_result_payload(result, config=config)
            self._write_result(task_id, payload)

        try:
            orchestrator = GlossaryOrchestrator(
                cache=cache,
                client=client,
                id_factory=lambda: task_id,
                progress=_touch_progress,
                on_executor_created=_capture,
                on_result_finalized=_finalize,
            )
            await orchestrator.run(config)
            self._maybe_cleanup_cache("glossary", task_id)
        finally:
            await client.aclose()

    def _build_glossary_review_config(
        self,
    ) -> tuple[GlossaryReviewConfig, ModelConfig, PromptPreset]:
        settings = self.settings_store.load_all()
        review = settings.glossary_review
        app = settings.app

        input_dir = _require_directory(review.input_folder, field="input_folder")
        try:
            output_filename = normalize_output_filename(review.output_filename)
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                str(exc),
                field="output_filename",
            ) from exc
        model = self._resolve_model_profile(
            app.active_glossary_review_model_id,
            field="active_glossary_review_model_id",
        )
        model = replace(model, timeout_seconds=float(review.timeout_seconds))
        source_lang = _coerce_language(
            review.source_language, field="source_language"
        )
        target_lang = _coerce_language(
            review.target_language, field="target_language"
        )
        preset = self._resolve_prompt_preset(
            app.active_glossary_review_prompt_id,
            kind=PromptKind.GLOSSARY_REVIEW,
        )
        config = GlossaryReviewConfig(
            input_dir=input_dir,
            selected_xlsx_path=Path(review.selected_xlsx_path)
            if review.selected_xlsx_path
            else None,
            selected_reference_paths=tuple(
                Path(path) for path in review.selected_reference_paths
            ),
            source_language=source_lang,
            target_language=target_lang,
            output_filename=output_filename,
            novel_background=review.novel_background,
            review_rounds=max(1, int(review.review_rounds)),
            batch_size=max(1, int(review.batch_size)),
            retry_attempts=max(0, int(review.retry_attempts)),
            model=model,
            prompt_preset=preset,
        )
        return config, model, preset

    def discover_glossary_review_inputs(
        self, *, input_folder: str, output_filename: str
    ) -> dict[str, object]:
        input_dir = _require_directory(input_folder, field="input_folder")
        try:
            output_name = normalize_output_filename(output_filename)
            candidates = discover_review_input_candidates(
                input_dir, output_filename=output_name
            )
        except ValueError as exc:
            raise BridgeError.invalid_argument(str(exc)) from exc
        return {
            "input_folder": str(input_dir),
            "xlsx_candidates": [
                {"path": str(path), "name": path.name}
                for path in candidates.xlsx_files
            ],
            "reference_candidates": [
                {"path": str(path), "name": path.name}
                for path in candidates.reference_files
            ],
        }

    def start_glossary_review(self, request_id: str) -> dict[str, object]:
        with self._start_locks["glossary_review"]:
            return self._start_glossary_review_locked(request_id)

    def _start_glossary_review_locked(self, request_id: str) -> dict[str, object]:
        config, model, preset = self._build_glossary_review_config()

        self._purge_kind_for_start(
            kind="glossary_review", task_kind=TaskKind.GLOSSARY_REVIEW
        )

        task_id = _new_task_id("glossary-review")
        started_at = _utc_now_iso()
        self._seed_placeholder(
            task_id,
            kind=TaskKind.GLOSSARY_REVIEW,
            started_at=started_at,
            metadata={
                "input_dir": str(config.input_dir),
                "output_dir": str(config.input_dir),
                "output_filename": config.output_filename,
                "source_language": config.source_language.value,
                "target_language": config.target_language.value,
                "review_rounds_total": config.review_rounds,
                "review_round_current": 0,
                "review_round_completed": 0,
                "review_round_total_batches": 0,
                "review_round_completed_batches": 0,
                "model_id": model.id,
                "prompt_preset_id": preset.id,
                "timeout_seconds": float(model.timeout_seconds),
                "request_id": request_id,
            },
        )

        cache = self._cache_for_kind("glossary_review")
        running = RunningTask(
            task_id=task_id,
            kind="glossary_review",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._glossary_review_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    async def _glossary_review_thread(
        self,
        task_id: str,
        config: GlossaryReviewConfig,
        running: RunningTask,
    ) -> None:
        client = self.llm_client_factory()
        cache = self._cache_for_kind("glossary_review")
        config = replace(config, debug_log_dir=cache.task_dir(task_id) / "debug")

        def _capture(executor: TaskExecutor) -> None:
            running.set_executor(executor)

        def _touch_progress(_event: object) -> None:
            running.touch()

        def _finalize(result: GlossaryReviewResult) -> None:
            payload = _glossary_review_result_payload(result, config=config)
            self._write_result(task_id, payload)

        try:
            orchestrator = GlossaryReviewOrchestrator(
                cache=cache,
                client=client,
                id_factory=lambda: task_id,
                progress=_touch_progress,
                on_executor_created=_capture,
                on_result_finalized=_finalize,
            )
            await orchestrator.run(config)
            self._maybe_cleanup_cache("glossary_review", task_id)
        finally:
            await client.aclose()

    def start_replacement(
        self,
        *,
        request_id: str,
        rules: Sequence[ReplacementRule],
        input_folder: str | None = None,
        output_folder: str | None = None,
    ) -> dict[str, object]:
        with self._start_locks["replacement"]:
            return self._start_replacement_locked(
                request_id=request_id,
                rules=rules,
                input_folder=input_folder,
                output_folder=output_folder,
            )

    def _start_replacement_locked(
        self,
        *,
        request_id: str,
        rules: Sequence[ReplacementRule],
        input_folder: str | None = None,
        output_folder: str | None = None,
    ) -> dict[str, object]:
        settings = self.settings_store.load_all()
        replacement = settings.replacement

        input_dir = _require_directory(
            input_folder or replacement.input_folder, field="input_folder"
        )
        resolved_output_folder = (
            output_folder or replacement.output_folder or str(input_dir)
        )
        output_dir = _ensure_output_dir(
            resolved_output_folder, field="output_folder"
        )

        if not rules:
            raise BridgeError.invalid_argument(
                "rules cannot be empty.",
                field="rules",
            )

        self._purge_kind_for_start(
            kind="replacement", task_kind=TaskKind.REPLACEMENT
        )

        task_id = _new_task_id("replacement")
        started_at = _utc_now_iso()
        files = scan_input_directory(input_dir)

        cache = self._cache_for_kind("replacement")
        running = RunningTask(
            task_id=task_id,
            kind="replacement",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)

        if not files:
            self._seed_placeholder(
                task_id,
                kind=TaskKind.REPLACEMENT,
                started_at=started_at,
                metadata={
                    "input_dir": str(input_dir),
                    "output_dir": str(output_dir),
                    "request_id": request_id,
                },
            )
            self._mark_status(task_id, TaskStatus.COMPLETED)
            self._write_result(
                task_id,
                {
                    "kind": "replacement",
                    "output_folder": str(output_dir),
                    "output_files": [],
                    "statistics_json_path": None,
                    "total_replacements": 0,
                },
            )
            running.mark_done()
            return {"task_id": task_id, "started_at": started_at}

        record = TaskRecord(
            id=task_id,
            kind=TaskKind.REPLACEMENT,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata={
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
                "request_id": request_id,
            },
        )
        subtasks = [
            Subtask(
                id=f"file-{index:04d}",
                task_id=task_id,
                request_payload={
                    "source_file": str(doc.path),
                    "format": doc.format.value,
                    "relative_path": doc.relative_path.as_posix(),
                },
            )
            for index, doc in enumerate(files)
        ]
        cache.write_seed(record, subtasks)

        stop_on_first_error = bool(replacement.stop_on_first_error)

        def _runner_target() -> None:
            self._run_replacement_loop(
                task_id=task_id,
                rules=tuple(rules),
                output_dir=output_dir,
                stop_on_first_error=stop_on_first_error,
                running=running,
            )

        self._spawn_thread(running, target=_runner_target, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _run_replacement_loop(
        self,
        *,
        task_id: str,
        rules: tuple[ReplacementRule, ...],
        output_dir: Path,
        stop_on_first_error: bool,
        running: RunningTask,
    ) -> None:
        running.touch()
        self._mark_status(task_id, TaskStatus.RUNNING)

        cache = self._cache_for_kind("replacement")
        snapshot = cache.load(task_id)
        outputs: list[Path] = []
        total_replacements = 0
        was_stopped = False
        early_break = False
        # Per-rule aggregation across files for the post-run report.
        # Each rule has a budget; once full we keep counting matches
        # but stop attaching new context snippets.
        occurrences_by_rule: dict[int, list[dict[str, object]]] = {
            i: [] for i in range(len(rules))
        }
        per_rule_total_count: dict[int, int] = {i: 0 for i in range(len(rules))}
        per_file_replacements: list[dict[str, object]] = []

        for subtask in snapshot.subtasks:
            running.touch()
            if running.stop_requested:
                was_stopped = True
                break

            running_state = replace(
                subtask,
                status=SubtaskStatus.RUNNING,
                attempt_count=subtask.attempt_count + 1,
                last_error="",
                last_error_at="",
            )
            cache.save_subtask(running_state)

            payload = subtask.request_payload
            source = Path(str(payload.get("source_file", "")))
            fmt = str(payload.get("format", "txt"))
            try:
                if fmt == DocumentFormat.EPUB.value:
                    result = replace_epub_file(
                        source, output_dir, list(rules), collect_occurrences=True
                    )
                else:
                    result = replace_txt_file(
                        source, output_dir, list(rules), collect_occurrences=True
                    )
                outputs.append(result.output_path)
                total_replacements += result.replacement_count
                per_file_replacements.append(
                    {
                        "source_path": str(source),
                        "output_path": str(result.output_path),
                        "replacement_count": result.replacement_count,
                    }
                )
                # Aggregate: tag each occurrence with the file it came
                # from before persisting. Per-rule budget keeps the
                # report bounded across many files.
                for occ in result.occurrences:
                    per_rule_total_count[occ.rule_index] = (
                        per_rule_total_count.get(occ.rule_index, 0) + 1
                    )
                    bucket = occurrences_by_rule.setdefault(occ.rule_index, [])
                    if len(bucket) >= _REPORT_MAX_OCCURRENCES_PER_RULE:
                        continue
                    bucket.append(
                        {
                            "file_path": str(source),
                            "char_offset": occ.char_offset,
                            "before_context": occ.before_context,
                            "match_text": occ.match_text,
                            "after_context": occ.after_context,
                            "replacement_text": occ.replacement_text,
                        }
                    )
                completed = replace(
                    running_state,
                    status=SubtaskStatus.COMPLETED,
                    response_content=json.dumps(
                        {
                            "output_path": str(result.output_path),
                            "replacement_count": result.replacement_count,
                            "errors": list(result.errors),
                        },
                        ensure_ascii=False,
                    ),
                )
                cache.save_subtask(completed)
            except Exception as exc:  # noqa: BLE001
                failed = replace(
                    running_state,
                    status=SubtaskStatus.FAILED,
                    last_error=f"{type(exc).__name__}: {exc}",
                    last_error_at=_utc_now_iso(),
                )
                cache.save_subtask(failed)
                if stop_on_first_error:
                    early_break = True
                    break

        # The rule-level chunk of the report carries the rule's own
        # parameters so the frontend can show "what was applied" without
        # having to re-look-up the user's settings.
        report_rules: list[dict[str, object]] = []
        for index, rule in enumerate(rules):
            captured = occurrences_by_rule.get(index, [])
            total_for_rule = per_rule_total_count.get(index, 0)
            report_rules.append(
                {
                    "rule_index": index,
                    "src": rule.src,
                    "dst": rule.dst,
                    "regex": rule.regex,
                    "case_sensitive": rule.case_sensitive,
                    "enabled": rule.enabled,
                    "total_count": total_for_rule,
                    "occurrences": captured,
                    "occurrences_truncated": total_for_rule > len(captured),
                }
            )

        report_path = self._write_replacement_report(
            task_id,
            {
                "task_id": task_id,
                "generated_at": _utc_now_iso(),
                "totals": {
                    "rules_active": sum(1 for r in rules if r.enabled),
                    "rules_with_matches": sum(
                        1 for r in report_rules if r["total_count"] > 0
                    ),
                    "total_replacements": total_replacements,
                    "files_processed": len(per_file_replacements),
                },
                "files": per_file_replacements,
                "rules": report_rules,
            },
        )

        statistics = {
            "kind": "replacement",
            "output_folder": str(output_dir),
            "output_files": [str(p) for p in outputs],
            "statistics_json_path": None,
            "replacement_report_path": str(report_path) if report_path else None,
            "total_replacements": total_replacements,
        }
        self._write_result(task_id, statistics)

        latest = cache.load(task_id)
        progress = latest.progress()
        if was_stopped or (early_break and progress.pending > 0):
            final = TaskStatus.STOPPED
        elif progress.failed > 0 and progress.pending == 0 and progress.running == 0:
            final = TaskStatus.FAILED
        elif progress.pending == 0 and progress.running == 0:
            final = TaskStatus.COMPLETED
        else:
            final = TaskStatus.STOPPED
        self._mark_status(task_id, final)
        self._maybe_cleanup_cache("replacement", task_id)

    def preview_epub_compress(
        self, *, input_path: str, mode: str, options: Mapping[str, object]
    ) -> dict[str, object]:
        try:
            config = EpubCompressOptions.from_mapping(options)
            return build_epub_compress_plan(
                Path(input_path), mode=mode, options=config
            ).to_dict()
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                str(exc),
                field="input_path",
            ) from exc

    def start_epub_compress(
        self,
        *,
        request_id: str,
        input_path: str,
        mode: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        with self._start_locks["epub_compress"]:
            return self._start_epub_compress_locked(
                request_id=request_id,
                input_path=input_path,
                mode=mode,
                options=options,
                actions=actions,
            )

    def _start_epub_compress_locked(
        self,
        *,
        request_id: str,
        input_path: str,
        mode: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        config = EpubCompressOptions.from_mapping(options)
        source_path = Path(input_path).expanduser().resolve()
        selected_actions = tuple(
            action
            for action in (
                EpubCompressAction.from_mapping(raw) for raw in actions
            )
            if action.selected
        )
        if not selected_actions:
            raise BridgeError.invalid_argument(
                "actions cannot be empty.",
                field="actions",
            )

        self._purge_kind_for_start(
            kind="epub_compress", task_kind=TaskKind.EPUB_COMPRESS
        )

        task_id = _new_task_id("epub-compress")
        started_at = _utc_now_iso()
        output_dir = (
            source_path.parent if source_path.is_file() else source_path
        )
        cache = self._cache_for_kind("epub_compress")
        running = RunningTask(
            task_id=task_id,
            kind="epub_compress",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)

        record = TaskRecord(
            id=task_id,
            kind=TaskKind.EPUB_COMPRESS,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata={
                "input_dir": str(source_path),
                "output_dir": str(output_dir),
                "request_id": request_id,
                "mode": mode,
                "options": config.to_dict(),
            },
        )
        subtasks = [
            Subtask(
                id=action.id or f"epub-{index:04d}",
                task_id=task_id,
                request_payload=action.to_dict(),
            )
            for index, action in enumerate(selected_actions)
        ]
        cache.write_seed(record, subtasks)

        def _runner_target() -> None:
            self._run_epub_compress_loop(
                task_id=task_id,
                input_path=source_path,
                mode=mode,
                options=config,
                running=running,
            )

        self._spawn_thread(running, target=_runner_target, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _run_epub_compress_loop(
        self,
        *,
        task_id: str,
        input_path: Path,
        mode: str,
        options: EpubCompressOptions,
        running: RunningTask,
    ) -> None:
        running.touch()
        self._mark_status(task_id, TaskStatus.RUNNING)

        cache = self._cache_for_kind("epub_compress")
        snapshot = cache.load(task_id)
        results = []
        was_stopped = False

        for subtask in snapshot.subtasks:
            running.touch()
            if running.stop_requested:
                was_stopped = True
                break

            running_state = replace(
                subtask,
                status=SubtaskStatus.RUNNING,
                attempt_count=subtask.attempt_count + 1,
                last_error="",
                last_error_at="",
            )
            cache.save_subtask(running_state)
            action = EpubCompressAction.from_mapping(subtask.request_payload)
            result = compress_epub_file(action, options)
            results.append(result)
            if result.status == "compressed":
                completed = replace(
                    running_state,
                    status=SubtaskStatus.COMPLETED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                )
                cache.save_subtask(completed)
            else:
                failed = replace(
                    running_state,
                    status=SubtaskStatus.FAILED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                    last_error=result.error,
                    last_error_at=_utc_now_iso(),
                )
                cache.save_subtask(failed)

        report = build_epub_compress_report(
            task_id=task_id,
            input_path=input_path,
            mode=mode,
            generated_at=_utc_now_iso(),
            results=results,
        )
        report_path = self._write_epub_compress_report(task_id, report)
        output_files = [
            row["output_path"]
            for row in report["results"]
            if row["status"] == "compressed"
        ]
        output_folder = input_path.parent if mode == "file" else input_path
        statistics = {
            "kind": "epub_compress",
            "output_folder": str(output_folder),
            "report_path": str(report_path) if report_path else None,
            "output_files": output_files,
            "compressed_count": report["totals"]["compressed"],
            "failed_count": report["totals"]["failed"],
        }
        self._write_result(task_id, statistics)

        latest = cache.load(task_id)
        progress = latest.progress()
        if was_stopped or progress.pending > 0 or progress.running > 0:
            final = TaskStatus.STOPPED
        elif progress.failed > 0:
            final = TaskStatus.FAILED
        else:
            final = TaskStatus.COMPLETED
        self._mark_status(task_id, final)
        self._maybe_cleanup_cache("epub_compress", task_id)

    def preview_epub_merge(
        self, *, input_dir: str, options: Mapping[str, object]
    ) -> dict[str, object]:
        try:
            config = EpubMergeOptions.from_mapping(options)
            return build_epub_merge_plan(
                Path(input_dir), options=config
            ).to_dict()
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                str(exc),
                field="input_dir",
            ) from exc

    def start_epub_merge(
        self,
        *,
        request_id: str,
        input_dir: str,
        output_path: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        with self._start_locks["epub_merge"]:
            return self._start_epub_merge_locked(
                request_id=request_id,
                input_dir=input_dir,
                output_path=output_path,
                options=options,
                actions=actions,
            )

    def _start_epub_merge_locked(
        self,
        *,
        request_id: str,
        input_dir: str,
        output_path: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        config = EpubMergeOptions.from_mapping({**dict(options), "output_path": output_path})
        source_dir = Path(input_dir).expanduser().resolve()
        selected_actions = tuple(
            action
            for action in (EpubMergeAction.from_mapping(raw) for raw in actions)
            if action.selected
        )
        if not selected_actions:
            raise BridgeError.invalid_argument(
                f"at least one {config.output_format.upper()} file must be selected.",
                field="actions",
            )
        if output_path.strip() or config.output_path.strip():
            output = Path(output_path or config.output_path).expanduser().resolve()
            output_suffix = ".txt" if config.output_format == "txt" else ".epub"
            if output.suffix.lower() != output_suffix:
                output = output.with_suffix(output_suffix)
        else:
            output = build_epub_merge_plan(source_dir, options=config).output_path

        self._purge_kind_for_start(
            kind="epub_merge", task_kind=TaskKind.EPUB_MERGE
        )

        task_id = _new_task_id("epub-merge")
        started_at = _utc_now_iso()
        cache = self._cache_for_kind("epub_merge")
        running = RunningTask(
            task_id=task_id,
            kind="epub_merge",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)

        record = TaskRecord(
            id=task_id,
            kind=TaskKind.EPUB_MERGE,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata={
                "input_dir": str(source_dir),
                "output_dir": str(output.parent),
                "output_path": str(output),
                "request_id": request_id,
                "options": config.to_dict(),
            },
        )
        subtask = Subtask(
            id="merge-0000",
            task_id=task_id,
            request_payload={
                "input_dir": str(source_dir),
                "output_path": str(output),
                "options": config.to_dict(),
                "actions": [action.to_dict() for action in selected_actions],
            },
        )
        cache.write_seed(record, [subtask])

        def _runner_target() -> None:
            self._run_epub_merge_loop(
                task_id=task_id,
                input_dir=source_dir,
                output_path=output,
                options=config,
                running=running,
            )

        self._spawn_thread(running, target=_runner_target, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _run_epub_merge_loop(
        self,
        *,
        task_id: str,
        input_dir: Path,
        output_path: Path,
        options: EpubMergeOptions,
        running: RunningTask,
    ) -> None:
        running.touch()
        self._mark_status(task_id, TaskStatus.RUNNING)
        cache = self._cache_for_kind("epub_merge")
        snapshot = cache.load(task_id)
        subtask = snapshot.subtasks[0]
        result: EpubMergeResult | None = None

        if running.stop_requested:
            self._mark_status(task_id, TaskStatus.STOPPED)
            self._maybe_cleanup_cache("epub_merge", task_id)
            return

        running_state = replace(
            subtask,
            status=SubtaskStatus.RUNNING,
            attempt_count=subtask.attempt_count + 1,
            last_error="",
            last_error_at="",
        )
        cache.save_subtask(running_state)
        payload = running_state.request_payload
        raw_actions = payload.get("actions", [])
        actions = [
            EpubMergeAction.from_mapping(raw)
            for raw in raw_actions
            if isinstance(raw, Mapping)
        ]
        result = merge_epub_files(
            action_id=running_state.id,
            input_dir=input_dir,
            output_path=output_path,
            actions=actions,
            options=options,
        )
        if result.status == "merged":
            cache.save_subtask(
                replace(
                    running_state,
                    status=SubtaskStatus.COMPLETED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                )
            )
        else:
            cache.save_subtask(
                replace(
                    running_state,
                    status=SubtaskStatus.FAILED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                    last_error=result.error,
                    last_error_at=_utc_now_iso(),
                )
            )

        report = build_epub_merge_report(
            task_id=task_id,
            input_dir=input_dir,
            generated_at=_utc_now_iso(),
            result=result,
        )
        report_path = self._write_epub_merge_report(task_id, report)
        statistics = {
            "kind": "epub_merge",
            "output_folder": str(output_path.parent),
            "report_path": str(report_path) if report_path else None,
            "output_files": [result.output_path] if result.status == "merged" else [],
            "merged_count": 1 if result.status == "merged" else 0,
            "failed_count": 1 if result.status == "failed" else 0,
        }
        self._write_result(task_id, statistics)

        latest = cache.load(task_id)
        progress = latest.progress()
        final = TaskStatus.FAILED if progress.failed > 0 else TaskStatus.COMPLETED
        self._mark_status(task_id, final)
        self._maybe_cleanup_cache("epub_merge", task_id)

    def preview_epub_convert(
        self,
        *,
        input_path: str,
        mode: str,
        options: Mapping[str, object],
    ) -> dict[str, object]:
        try:
            config = EpubConvertOptions.from_mapping(options)
            return build_epub_convert_plan(
                Path(input_path),
                mode=mode,
                options=config,
            ).to_dict()
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                str(exc),
                field="input_path",
            ) from exc

    def start_epub_convert(
        self,
        *,
        request_id: str,
        input_path: str,
        mode: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        with self._start_locks["epub_convert"]:
            return self._start_epub_convert_locked(
                request_id=request_id,
                input_path=input_path,
                mode=mode,
                options=options,
                actions=actions,
            )

    def _start_epub_convert_locked(
        self,
        *,
        request_id: str,
        input_path: str,
        mode: str,
        options: Mapping[str, object],
        actions: Sequence[Mapping[str, object]],
    ) -> dict[str, object]:
        config = EpubConvertOptions.from_mapping(options)
        source_path = Path(input_path).expanduser().resolve()
        selected_actions = tuple(
            action
            for action in (
                EpubConvertAction.from_mapping(raw) for raw in actions
            )
            if action.selected
        )
        if not selected_actions:
            raise BridgeError.invalid_argument(
                "actions cannot be empty.",
                field="actions",
            )

        self._purge_kind_for_start(
            kind="epub_convert", task_kind=TaskKind.EPUB_CONVERT
        )

        task_id = _new_task_id("epub-convert")
        started_at = _utc_now_iso()
        output_dir = Path(config.output_dir).expanduser().resolve() if config.output_dir.strip() else (
            source_path.parent if source_path.is_file() else source_path
        )
        cache = self._cache_for_kind("epub_convert")
        running = RunningTask(
            task_id=task_id,
            kind="epub_convert",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)

        record = TaskRecord(
            id=task_id,
            kind=TaskKind.EPUB_CONVERT,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata={
                "input_dir": str(source_path),
                "output_dir": str(output_dir),
                "request_id": request_id,
                "mode": mode,
                "options": config.to_dict(),
            },
        )
        subtasks = [
            Subtask(
                id=action.id or f"epub-{index:04d}",
                task_id=task_id,
                request_payload=action.to_dict(),
            )
            for index, action in enumerate(selected_actions)
        ]
        cache.write_seed(record, subtasks)

        def _runner_target() -> None:
            self._run_epub_convert_loop(
                task_id=task_id,
                input_path=source_path,
                mode=mode,
                running=running,
            )

        self._spawn_thread(running, target=_runner_target, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _run_epub_convert_loop(
        self,
        *,
        task_id: str,
        input_path: Path,
        mode: str,
        running: RunningTask,
    ) -> None:
        running.touch()
        self._mark_status(task_id, TaskStatus.RUNNING)

        cache = self._cache_for_kind("epub_convert")
        snapshot = cache.load(task_id)
        results = []
        was_stopped = False

        for subtask in snapshot.subtasks:
            running.touch()
            if running.stop_requested:
                was_stopped = True
                break

            running_state = replace(
                subtask,
                status=SubtaskStatus.RUNNING,
                attempt_count=subtask.attempt_count + 1,
                last_error="",
                last_error_at="",
            )
            cache.save_subtask(running_state)
            action = EpubConvertAction.from_mapping(subtask.request_payload)
            result = convert_epub_to_txt(action)
            results.append(result)
            if result.status == "converted":
                completed = replace(
                    running_state,
                    status=SubtaskStatus.COMPLETED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                )
                cache.save_subtask(completed)
            else:
                failed = replace(
                    running_state,
                    status=SubtaskStatus.FAILED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                    last_error=result.error,
                    last_error_at=_utc_now_iso(),
                )
                cache.save_subtask(failed)

        report = build_epub_convert_report(
            task_id=task_id,
            input_path=input_path,
            mode=mode,
            generated_at=_utc_now_iso(),
            results=results,
        )
        report_path = self._write_epub_convert_report(task_id, report)
        output_files = [
            row["output_path"]
            for row in report["results"]
            if row["status"] == "converted"
        ]
        output_folder = Path(str(cache.load_record(task_id).metadata.get("output_dir", "")))
        statistics = {
            "kind": "epub_convert",
            "output_folder": str(output_folder),
            "report_path": str(report_path) if report_path else None,
            "output_files": output_files,
            "converted_count": report["totals"]["converted"],
            "failed_count": report["totals"]["failed"],
        }
        self._write_result(task_id, statistics)

        latest = cache.load(task_id)
        progress = latest.progress()
        if was_stopped or progress.pending > 0 or progress.running > 0:
            final = TaskStatus.STOPPED
        elif progress.failed > 0:
            final = TaskStatus.FAILED
        else:
            final = TaskStatus.COMPLETED
        self._mark_status(task_id, final)
        self._maybe_cleanup_cache("epub_convert", task_id)

    def preview_txt_to_epub(
        self,
        *,
        options: Mapping[str, object],
    ) -> dict[str, object]:
        try:
            config = TxtToEpubOptions.from_mapping(options)
            return build_txt_to_epub_plan(config).to_dict()
        except ValueError as exc:
            raise BridgeError.invalid_argument(
                str(exc),
                field="options",
            ) from exc

    def start_txt_to_epub(
        self,
        *,
        request_id: str,
        options: Mapping[str, object],
    ) -> dict[str, object]:
        with self._start_locks["txt_to_epub"]:
            return self._start_txt_to_epub_locked(
                request_id=request_id,
                options=options,
            )

    def _start_txt_to_epub_locked(
        self,
        *,
        request_id: str,
        options: Mapping[str, object],
    ) -> dict[str, object]:
        config = TxtToEpubOptions.from_mapping(options)
        try:
            plan = build_txt_to_epub_plan(config)
        except ValueError as exc:
            raise BridgeError.invalid_argument(str(exc), field="options") from exc

        self._purge_kind_for_start(
            kind="txt_to_epub", task_kind=TaskKind.TXT_TO_EPUB
        )

        task_id = _new_task_id("txt-to-epub")
        started_at = _utc_now_iso()
        output_dir = plan.output_path.parent
        cache = self._cache_for_kind("txt_to_epub")
        running = RunningTask(
            task_id=task_id,
            kind="txt_to_epub",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)

        record = TaskRecord(
            id=task_id,
            kind=TaskKind.TXT_TO_EPUB,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata={
                "input_dir": str(plan.input_path),
                "output_dir": str(output_dir),
                "request_id": request_id,
                "options": config.to_dict(),
            },
        )
        cache.write_seed(
            record,
            [
                Subtask(
                    id=plan.action.id,
                    task_id=task_id,
                    request_payload=plan.action.to_dict(),
                )
            ],
        )

        def _runner_target() -> None:
            self._run_txt_to_epub_loop(
                task_id=task_id,
                input_path=plan.input_path,
                running=running,
            )

        self._spawn_thread(running, target=_runner_target, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _run_txt_to_epub_loop(
        self,
        *,
        task_id: str,
        input_path: Path,
        running: RunningTask,
    ) -> None:
        running.touch()
        self._mark_status(task_id, TaskStatus.RUNNING)

        cache = self._cache_for_kind("txt_to_epub")
        snapshot = cache.load(task_id)
        results = []
        was_stopped = False

        for subtask in snapshot.subtasks:
            running.touch()
            if running.stop_requested:
                was_stopped = True
                break
            running_state = replace(
                subtask,
                status=SubtaskStatus.RUNNING,
                attempt_count=subtask.attempt_count + 1,
                last_error="",
                last_error_at="",
            )
            cache.save_subtask(running_state)
            action = TxtToEpubAction.from_mapping(subtask.request_payload)
            result = convert_txt_to_epub(action)
            results.append(result)
            if result.status == "converted":
                completed = replace(
                    running_state,
                    status=SubtaskStatus.COMPLETED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                )
                cache.save_subtask(completed)
            else:
                failed = replace(
                    running_state,
                    status=SubtaskStatus.FAILED,
                    response_content=json.dumps(
                        result.to_dict(), ensure_ascii=False
                    ),
                    last_error=result.error,
                    last_error_at=_utc_now_iso(),
                )
                cache.save_subtask(failed)

        report = build_txt_to_epub_report(
            task_id=task_id,
            input_path=input_path,
            generated_at=_utc_now_iso(),
            results=results,
        )
        report_path = self._write_txt_to_epub_report(task_id, report)
        output_files = [
            row["output_path"]
            for row in report["results"]
            if row["status"] == "converted"
        ]
        output_folder = Path(str(cache.load_record(task_id).metadata.get("output_dir", "")))
        statistics = {
            "kind": "txt_to_epub",
            "output_folder": str(output_folder),
            "report_path": str(report_path) if report_path else None,
            "output_files": output_files,
            "converted_count": report["totals"]["converted"],
            "failed_count": report["totals"]["failed"],
        }
        self._write_result(task_id, statistics)

        latest = cache.load(task_id)
        progress = latest.progress()
        if was_stopped or progress.pending > 0 or progress.running > 0:
            final = TaskStatus.STOPPED
        elif progress.failed > 0:
            final = TaskStatus.FAILED
        else:
            final = TaskStatus.COMPLETED
        self._mark_status(task_id, final)
        self._maybe_cleanup_cache("txt_to_epub", task_id)

    def stop_task(self, *, kind: str, task_id: str) -> dict[str, object]:
        running = self.registry.get(task_id)
        if running is None:
            try:
                snapshot = self._cache_for_task(task_id).load(task_id)
            except TaskNotFoundError as exc:
                raise BridgeError.not_found(
                    f"task {task_id!r} not found.",
                    details={"task_id": task_id},
                ) from exc
            if snapshot.record.kind is not self._kind(kind):
                raise BridgeError.invalid_argument(
                    f"task {task_id!r} kind mismatch.",
                    field="task_id",
                )
            healed = self._reconcile_zombie(snapshot, self._cache_for_task(task_id))
            if healed.record.status is TaskStatus.STOPPED:
                return {"snapshot": _format_snapshot(healed)}
            raise BridgeError(
                "task.not_running",
                f"task {task_id!r} is not currently running.",
                retryable=False,
                details={"task_id": task_id},
            )
        if running.kind != kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind {running.kind!r} does not match {kind!r}.",
                field="task_id",
            )
        running.request_stop()
        self._mark_status(task_id, TaskStatus.STOPPING)
        return self.read_snapshot(kind=kind, task_id=task_id)

    def pause_task(self, *, kind: str, task_id: str) -> dict[str, object]:
        if kind in {
            "replacement",
            "epub_compress",
            "epub_merge",
            "epub_convert",
            "txt_to_epub",
        }:
            raise BridgeError(
                "task.invalid_transition",
                f"pause is not supported for {kind} (single-pass tool).",
                retryable=False,
                details={"reason": "single_pass"},
            )
        running = self.registry.get(task_id)
        if running is None:
            try:
                self._cache_for_task(task_id).load_record(task_id)
            except TaskNotFoundError as exc:
                raise BridgeError.not_found(
                    f"task {task_id!r} not found.",
                    details={"task_id": task_id},
                ) from exc
            raise BridgeError(
                "task.not_running",
                f"task {task_id!r} is not currently running.",
                retryable=False,
                details={"task_id": task_id},
            )
        if running.kind != kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind {running.kind!r} does not match {kind!r}.",
                field="task_id",
            )
        running.request_pause()
        return self.read_snapshot(kind=kind, task_id=task_id)

    def continue_task(self, *, kind: str, task_id: str) -> dict[str, object]:
        if kind in {
            "replacement",
            "epub_compress",
            "epub_merge",
            "epub_convert",
            "txt_to_epub",
        }:
            raise BridgeError(
                "task.invalid_transition",
                f"continue is not supported for {kind} (single-pass tool).",
                retryable=False,
                details={"reason": "single_pass"},
            )
        record_kind = self._kind(kind)
        cache = self._cache_for_task(task_id)
        try:
            snapshot = cache.load(task_id)
        except TaskNotFoundError as exc:
            mirrored = self._completed_snapshots.get(task_id)
            if mirrored is not None and mirrored.record.kind is record_kind:
                # The disk cache was wiped because this task already
                # finished cleanly; reject with the same error code we'd
                # return if the cache were still on disk so callers see
                # consistent behavior.
                raise BridgeError(
                    "task.invalid_transition",
                    f"continue requires status STOPPED, PAUSED, or FAILED; got {mirrored.record.status.value}.",
                    retryable=False,
                    details={"status": mirrored.record.status.value},
                ) from exc
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if snapshot.record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        snapshot = self._reconcile_zombie(snapshot, cache)
        existing = self._resolve_live_running(task_id, snapshot.record.status)
        if existing is not None and not existing.is_done:
            self._raise_live_task_conflict(
                existing,
                stall_seconds=self._live_task_stall_seconds_for_snapshot(snapshot),
            )
        if snapshot.record.status not in (
            TaskStatus.STOPPED,
            TaskStatus.PAUSED,
            TaskStatus.FAILED,
        ):
            raise BridgeError(
                "task.invalid_transition",
                f"continue requires status STOPPED, PAUSED, or FAILED; got {snapshot.record.status.value}.",
                retryable=False,
                details={"status": snapshot.record.status.value},
            )
        progress = snapshot.progress()
        if progress.pending == 0 and progress.failed == 0:
            if kind == "translation":
                return self._continue_translation(task_id)
            if kind == "glossary":
                return self._continue_glossary(task_id)
            if kind == "glossary_review":
                return self._continue_glossary_review(task_id)
            raise BridgeError(
                "task.invalid_transition",
                "task has no pending or failed subtasks; nothing to continue.",
                retryable=False,
                details={"reason": "no_remaining_work"},
            )

        if kind == "translation":
            return self._continue_translation(task_id)
        if kind == "glossary":
            return self._continue_glossary(task_id)
        return self._continue_glossary_review(task_id)

    def _continue_translation(self, task_id: str) -> dict[str, object]:
        config, _model, _preset = self._build_translation_config()
        started_at = _utc_now_iso()
        cache = self._cache_for_kind("translation")
        record = cache.load_record(task_id)
        metadata = dict(record.metadata)
        metadata["active_model_id"] = config.model.id
        metadata["active_prompt_id"] = config.prompt_preset.id
        cache.save_task(replace(record, metadata=metadata))
        running = RunningTask(
            task_id=task_id,
            kind="translation",
            cache=cache,
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._translation_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _continue_glossary(self, task_id: str) -> dict[str, object]:
        config, _model, _preset = self._build_glossary_config()
        started_at = _utc_now_iso()
        running = RunningTask(
            task_id=task_id,
            kind="glossary",
            cache=self._cache_for_kind("glossary"),
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._glossary_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def _continue_glossary_review(self, task_id: str) -> dict[str, object]:
        config, _model, _preset = self._build_glossary_review_config()
        started_at = _utc_now_iso()
        running = RunningTask(
            task_id=task_id,
            kind="glossary_review",
            cache=self._cache_for_kind("glossary_review"),
            created_at=started_at,
        )
        self.registry.add(running)
        self._mark_status(task_id, TaskStatus.RUNNING)

        def _async_runner() -> None:
            asyncio.run(self._glossary_review_thread(task_id, config, running))

        self._spawn_thread(running, target=_async_runner, task_id=task_id)
        return {"task_id": task_id, "started_at": started_at}

    def probe_continuable(self, *, kind: str) -> dict[str, object]:
        """Return whether a continuable cache exists for ``kind`` under
        the current settings (architecture § 1.3).

        Replacement is single-pass: always returns
        ``continuable=false`` regardless of cache state.
        """

        if kind in {
            "replacement",
            "epub_compress",
            "epub_merge",
            "epub_convert",
            "txt_to_epub",
        }:
            return {
                "continuable": False,
                "task_id": None,
                "status": None,
                "pending": 0,
                "failed": 0,
            }
        record_kind = self._kind(kind)
        settings = self.settings_store.load_all()
        if kind == "translation":
            input_folder = settings.translation.input_folder
            output_folder = settings.translation.output_folder
        elif kind == "glossary":
            input_folder = settings.glossary.input_folder
            output_folder = settings.glossary.output_folder
        else:
            input_folder = settings.glossary_review.input_folder
            output_folder = settings.glossary_review.input_folder

        cache = self._cache_for_kind(kind)
        candidates = sorted(
            (r for r in cache.list_tasks() if r.kind is record_kind),
            key=lambda r: r.created_at,
            reverse=True,
        )
        # Compare via the cache-key normalizer so a trailing slash,
        # double slashes, or NFD-vs-NFC (macOS readdir vs settings.json)
        # don't silently filter out a continuable task.
        expected_input = _normalized_dir(input_folder)
        expected_output = _normalized_dir(output_folder)
        for record in candidates:
            metadata = record.metadata
            if _normalized_dir(metadata.get("input_dir")) != expected_input:
                continue
            if _normalized_dir(metadata.get("output_dir")) != expected_output:
                continue
            try:
                snapshot = cache.load(record.id)
            except (TaskNotFoundError, ValueError, OSError):
                continue
            snapshot = self._reconcile_zombie(snapshot, cache)
            if snapshot.record.status not in (
                TaskStatus.STOPPED,
                TaskStatus.PAUSED,
                TaskStatus.FAILED,
            ):
                continue
            progress = snapshot.progress()
            finalize_only = (
                progress.pending == 0
                and progress.failed == 0
                and progress.running == 0
            )
            if progress.pending + progress.failed <= 0 and not finalize_only:
                continue
            return {
                "continuable": True,
                "task_id": record.id,
                "status": snapshot.record.status.value,
                "pending": progress.pending,
                "failed": progress.failed,
            }
        return {
            "continuable": False,
            "task_id": None,
            "status": None,
            "pending": 0,
            "failed": 0,
        }

    def read_snapshot(self, *, kind: str, task_id: str) -> dict[str, object]:
        record_kind = self._kind(kind)
        cache = self._cache_for_task(task_id)
        try:
            snapshot = cache.load(task_id)
        except TaskNotFoundError as exc:
            mirrored = self._completed_snapshots.get(task_id)
            if mirrored is not None and mirrored.record.kind is record_kind:
                # Disk cache was wiped after a clean COMPLETED run; serve
                # the in-memory mirror so the Run page keeps showing the
                # final stats until a new task is started.
                return {"snapshot": _format_snapshot(mirrored)}
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if snapshot.record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        snapshot = self._reconcile_zombie(snapshot, cache)
        snapshot = self._reconcile_glossary_review_completion(snapshot, cache)
        # Close the race window where disk shows terminal but the runner
        # thread hasn't flipped registry.is_done yet: the UI polls
        # read_snapshot, sees FAILED, opens the failure dialog; if the
        # user clicks Continue before the runner wrap-up completes,
        # continue_task hits a phantom "already running" conflict.
        # Force-clean here so by the time the UI sees terminal status
        # the registry is already settled.
        self._resolve_live_running(task_id, snapshot.record.status)
        return {"snapshot": _format_snapshot(snapshot)}

    def list_recent_tasks(
        self, *, kind: str, limit: int | None
    ) -> dict[str, object]:
        record_kind = self._kind(kind)
        cache = self._cache_for_kind(kind)
        # Pull from disk first. Reconcile transient orphan records before
        # formatting headers so a restart after early task seeding doesn't
        # surface a dead PENDING/RUNNING task as "Starting..." forever.
        on_disk: list[TaskRecord] = []
        for record in cache.list_tasks():
            if record.kind is not record_kind:
                continue
            if record.status in _ZOMBIE_TASK_STATES:
                try:
                    snapshot = self._reconcile_zombie(cache.load(record.id), cache)
                except (TaskNotFoundError, ValueError, OSError):
                    continue
                record = snapshot.record
            elif record.kind is TaskKind.GLOSSARY_REVIEW and record.status in (
                TaskStatus.COMPLETED,
                TaskStatus.STOPPED,
            ):
                try:
                    snapshot = self._reconcile_glossary_review_completion(
                        cache.load(record.id), cache
                    )
                except (TaskNotFoundError, ValueError, OSError):
                    continue
                record = snapshot.record
            on_disk.append(record)
        seen_ids = {r.id for r in on_disk}
        # Then merge in any completed tasks that have already had their
        # disk cache wiped — without this, the frontend's ``refreshActiveTask``
        # (which calls listRecentTasks(1) and uses the result to set
        # ``activeTaskId``) would lose track of the completed task and
        # zero out the Run page UI seconds after a clean COMPLETED run.
        for tid, snap in self._completed_snapshots.items():
            if snap.record.kind is not record_kind:
                continue
            if tid in seen_ids:
                continue
            on_disk.append(snap.record)
        on_disk.sort(key=lambda r: r.created_at, reverse=True)
        if limit is not None and limit >= 0:
            on_disk = on_disk[:limit]
        return {"tasks": [_format_header(r) for r in on_disk]}

    def list_failed_subtasks(
        self, *, kind: str, task_id: str
    ) -> dict[str, object]:
        record_kind = self._kind(kind)
        try:
            snapshot = self._cache_for_task(task_id).load(task_id)
        except TaskNotFoundError as exc:
            mirrored = self._completed_snapshots.get(task_id)
            if mirrored is not None and mirrored.record.kind is record_kind:
                # Disk cache wiped after a clean COMPLETED run; the
                # mirror has no subtasks, so failures = []. Returning
                # an empty list here (instead of bridge.not_found) is
                # what keeps the frontend's pollSnapshot from clearing
                # ``activeTaskId`` and zeroing the Run page.
                return {"failures": []}
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if snapshot.record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        return {"failures": _format_failures(snapshot)}

    def read_request_events(
        self,
        *,
        kind: str,
        task_id: str,
        limit: int | None = None,
        offset: int = 0,
        status: str = "",
    ) -> dict[str, object]:
        record_kind = self._kind(kind)
        cache = self._cache_for_task(task_id)
        try:
            record = cache.load_record(task_id)
        except TaskNotFoundError as exc:
            mirrored = self._completed_snapshots.get(task_id)
            if mirrored is not None and mirrored.record.kind is record_kind:
                return {"events": [], "total": 0}
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        if offset > 0 or status:
            events = cache.load_request_events(task_id)
            truncated = False
        else:
            events, truncated = cache.load_recent_request_events(task_id)
        try:
            snapshot = cache.load(task_id)
        except TaskNotFoundError:
            snapshot = None
        if snapshot is not None:
            snapshot = self._reconcile_zombie(snapshot, cache)
            events = (
                *events,
                *_failed_subtask_request_events(
                    cache,
                    record=record,
                    snapshot=snapshot,
                    request_events=events,
                ),
            )
        with self._retranslate_lock:
            live_retranslate_subtasks = {
                _retranslate_log_subtask_id(job.request_id)
                for job in self._retranslate_jobs.values()
                if job.task_id == task_id and job.status in {"pending", "running"}
            }
        formatted, total = _format_request_events(
            events,
            limit=limit,
            offset=offset,
            status=status,
            task_status=snapshot.record.status if snapshot is not None else record.status,
            subtask_statuses={
                subtask.id: subtask.status for subtask in snapshot.subtasks
            }
            if snapshot is not None
            else {},
            live_subtask_ids=live_retranslate_subtasks,
        )
        if snapshot is not None and record_kind is TaskKind.TRANSLATION:
            formatted = _translation_request_segment_refs(formatted, snapshot)
        return {
            "events": formatted,
            "total": total,
            "truncated": truncated,
        }

    def read_artifacts(self, *, kind: str, task_id: str) -> dict[str, object]:
        record_kind = self._kind(kind)
        cache = self._cache_for_task(task_id)
        try:
            record = cache.load_record(task_id)
        except TaskNotFoundError as exc:
            mirrored_result = self._completed_results.get(task_id)
            mirrored_snapshot = self._completed_snapshots.get(task_id)
            if mirrored_result is not None and mirrored_snapshot is not None:
                if mirrored_snapshot.record.kind is not record_kind:
                    raise BridgeError.invalid_argument(
                        f"task {task_id!r} kind mismatch.",
                        field="task_id",
                    ) from exc
                return mirrored_result
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        result = self._read_result(task_id)
        if result is None:
            snapshot = cache.load(task_id)
            return self._partial_result(record=record, snapshot=snapshot)
        return result

    @staticmethod
    def _kind(kind: str) -> TaskKind:
        record_kind = _KIND_TO_TASKKIND.get(kind)
        if record_kind is None:
            raise BridgeError.invalid_argument(
                f"unsupported task kind: {kind!r}",
                field="kind",
            )
        return record_kind

    def _resolve_model_profile(
        self, profile_id: str | None, *, field: str
    ) -> ModelConfig:
        if not profile_id:
            raise BridgeError.invalid_argument(
                f"{field} is not set; pick a model profile in App Settings.",
                field=field,
            )
        profile = self.profile_store.get(profile_id)
        if profile is None:
            raise BridgeError.invalid_argument(
                f"model profile {profile_id!r} does not exist.",
                field=field,
            )
        if not profile.api_keys:
            raise BridgeError.invalid_argument(
                f"model profile {profile_id!r} has no API key configured.",
                field=field,
                details={"reason": "missing_api_key"},
            )
        return profile

    def _resolve_prompt_preset(
        self, preset_id: str | None, *, kind: PromptKind
    ) -> PromptPreset:
        filename = f"prompts.{kind.value}.json"
        store = PromptPresetStore(
            path=self.prompts_cache_root / filename, kind=kind
        )
        return store.get_active(preset_id)

    def _purge_kind_for_start(
        self, *, kind: str, task_kind: TaskKind, join_timeout: float = 30.0
    ) -> None:
        """Stop any in-flight task on this kind so a new ``start`` can
        seed cache without colliding with a live writer thread.

        ``start`` is **no longer destructive to prior cache**: every
        run gets a fresh ``task_id`` so old runs coexist on disk and
        remain available for resume / proofreading / inspection. Stale
        runs are removed only via the user-driven cache cleanup UI.
        The fresh start uses a new id, so its records never collide
        with the prior one. Old in-memory completion mirrors are kept
        too — read_snapshot keys by task_id so the new task sees an
        empty mirror naturally.
        """

        live: list[RunningTask] = []
        for running in self.registry.list_by_kind(kind):
            if running.is_done:
                continue
            try:
                record = self._cache_for_task(running.task_id).load_record(
                    running.task_id
                )
            except (TaskNotFoundError, OSError, ValueError):
                live.append(running)
                continue
            resolved = self._resolve_live_running(running.task_id, record.status)
            if resolved is not None and not resolved.is_done:
                live.append(resolved)
        for running in live:
            running.request_stop()
        for running in live:
            if running.thread is not None:
                running.thread.join(timeout=join_timeout)
        for running in live:
            if running.is_done:
                continue
            if running.thread is not None and not running.thread.is_alive():
                running.mark_done()
                continue
            self._raise_live_task_conflict(running)

    def _spawn_thread(
        self,
        running: RunningTask,
        *,
        target: Callable[[], None],
        task_id: str,
    ) -> None:
        def _runner() -> None:
            try:
                target()
            except BaseException as exc:  # noqa: BLE001
                running.mark_done(error=exc)
                self._on_task_failure(task_id, exc)
                return
            running.mark_done()

        thread = threading.Thread(
            target=_runner, name=f"task-{task_id}", daemon=True
        )
        running.thread = thread
        thread.start()

    def _resolve_live_running(
        self, task_id: str, snapshot_status: TaskStatus
    ) -> "RunningTask | None":
        """Return the registry entry only when it's actually live.

        If the disk snapshot shows a terminal state but the registry still
        has a not-done entry, the entry is stale (thread exited without
        reaching mark_done — typically a host crash or unexpected
        BaseException path). Force-mark it done so subsequent
        continue/retranslate calls don't see a phantom "already running"
        conflict.
        """

        existing = self.registry.get(task_id)
        if existing is None or existing.is_done:
            return None
        if snapshot_status in _TERMINAL_TASK_STATES:
            existing.mark_done()
            return None
        return existing

    def _raise_live_task_conflict(
        self, running: RunningTask, *, stall_seconds: float | None = None
    ) -> None:
        stalled = self._live_task_is_stalled(
            running, stall_seconds=stall_seconds
        )
        heartbeat_age = running.heartbeat_age_seconds()
        stop_age = running.stop_requested_age_seconds()
        details: dict[str, object] = {
            "task_id": running.task_id,
            "stalled": stalled,
            "stop_requested": running.stop_requested,
            "heartbeat_age_seconds": round(heartbeat_age, 3),
            "stall_threshold_seconds": round(
                stall_seconds
                if stall_seconds is not None
                else _LIVE_TASK_STALL_SECONDS,
                3,
            ),
        }
        if stop_age is not None:
            details["stop_requested_age_seconds"] = round(stop_age, 3)
        if stalled:
            if running.stop_requested:
                message = (
                    f"task {running.task_id!r} is still stopping and appears "
                    "stalled; restart the app if it does not recover."
                )
            else:
                message = (
                    f"task {running.task_id!r} is already running and appears "
                    "stalled; stop it or restart the app before continuing."
                )
        else:
            message = f"task {running.task_id!r} is already running."
        raise BridgeError.conflict(message, details=details)

    def _live_task_is_stalled(
        self, running: RunningTask, *, stall_seconds: float | None = None
    ) -> bool:
        threshold = (
            _LIVE_TASK_STALL_SECONDS if stall_seconds is None else stall_seconds
        )
        return running.is_stalled(
            active_timeout_seconds=threshold,
            stopping_timeout_seconds=max(_STOP_REQUEST_STALL_SECONDS, threshold),
        )

    def _live_task_stall_seconds_for_snapshot(
        self, snapshot: TaskSnapshot
    ) -> float:
        timeout_seconds = _metadata_timeout_seconds(snapshot.record.metadata)
        if timeout_seconds is None:
            return _LIVE_TASK_STALL_SECONDS
        return max(
            _LIVE_TASK_STALL_SECONDS,
            timeout_seconds + _LIVE_TASK_STALL_TIMEOUT_HEADROOM_SECONDS,
        )

    def _seed_placeholder(
        self,
        task_id: str,
        *,
        kind: TaskKind,
        started_at: str,
        metadata: Mapping[str, object],
    ) -> None:
        record = TaskRecord(
            id=task_id,
            kind=kind,
            status=TaskStatus.PENDING,
            created_at=started_at,
            updated_at=started_at,
            metadata=dict(metadata),
        )
        self._cache_for_task(task_id).save_task(record)

    def _mark_status(self, task_id: str, status: TaskStatus) -> None:
        cache = self._cache_for_task(task_id)
        try:
            record = cache.load_record(task_id)
        except TaskNotFoundError:
            return
        if record.status is status:
            return
        now = _utc_now_iso()
        cache.save_task(
            replace(
                record,
                status=status,
                updated_at=now,
                metadata=runtime_metadata_for_status(record, status, now),
            )
        )

    def _reconcile_stale_running_subtasks(
        self, snapshot: TaskSnapshot, cache: TaskCache
    ) -> TaskSnapshot:
        if snapshot.record.status not in _STALE_RUNNING_SUBTASK_TASK_STATES:
            return snapshot
        healed_subtasks: list[Subtask] = []
        changed = False
        for subtask in snapshot.subtasks:
            if subtask.status is SubtaskStatus.RUNNING:
                pending = replace(
                    subtask,
                    status=SubtaskStatus.PENDING,
                    started_at="",
                )
                cache.save_subtask(pending)
                healed_subtasks.append(pending)
                changed = True
            else:
                healed_subtasks.append(subtask)
        if not changed:
            return snapshot
        return TaskSnapshot(record=snapshot.record, subtasks=tuple(healed_subtasks))

    def _reconcile_zombie(
        self, snapshot: TaskSnapshot, cache: TaskCache
    ) -> TaskSnapshot:
        # After a host-process crash (e.g. SIGBUS, OOM, kill -9) the runtime
        # never writes a terminal state, so the cache stays transient while
        # the in-memory executor is gone or no longer heartbeating. Flip it to
        # STOPPED here so the UI surfaces a continuable task instead of a
        # zombie.
        record = snapshot.record
        if record.status not in _ZOMBIE_TASK_STATES:
            return self._reconcile_stale_running_subtasks(snapshot, cache)
        live = self.registry.get(record.id)
        if live is not None and not live.is_done:
            stall_seconds = self._live_task_stall_seconds_for_snapshot(snapshot)
            if not self._live_task_is_stalled(
                live, stall_seconds=stall_seconds
            ):
                return snapshot
            live.mark_done()
        now = _utc_now_iso()
        healed = replace(
            record,
            status=TaskStatus.STOPPED,
            updated_at=now,
            metadata=runtime_metadata_for_status(record, TaskStatus.STOPPED, now),
        )
        cache.save_task(healed)
        healed_subtasks: list[Subtask] = []
        for subtask in snapshot.subtasks:
            if subtask.status is SubtaskStatus.RUNNING:
                pending = replace(
                    subtask,
                    status=SubtaskStatus.PENDING,
                    started_at="",
                )
                cache.save_subtask(pending)
                healed_subtasks.append(pending)
            else:
                healed_subtasks.append(subtask)
        return TaskSnapshot(record=healed, subtasks=tuple(healed_subtasks))

    def _reconcile_glossary_review_completion(
        self, snapshot: TaskSnapshot, cache: TaskCache
    ) -> TaskSnapshot:
        record = snapshot.record
        if record.kind is not TaskKind.GLOSSARY_REVIEW:
            return snapshot
        artifacts_exist = self._glossary_review_artifacts_exist(record)
        progress = snapshot.progress()
        if (
            record.status is TaskStatus.STOPPED
            and artifacts_exist
            and progress.total > 0
            and progress.pending == 0
            and progress.running == 0
            and progress.failed == 0
            and progress.completed == progress.total
        ):
            now = _utc_now_iso()
            healed = replace(
                record,
                status=TaskStatus.COMPLETED,
                updated_at=now,
                metadata=runtime_metadata_for_status(
                    record, TaskStatus.COMPLETED, now
                ),
            )
            cache.save_task(healed)
            return TaskSnapshot(record=healed, subtasks=snapshot.subtasks)
        if record.status is not TaskStatus.COMPLETED:
            return snapshot
        if artifacts_exist:
            return snapshot
        live = self.registry.get(record.id)
        if live is not None and not live.is_done:
            transient = record.with_status(TaskStatus.RUNNING)
            return TaskSnapshot(record=transient, subtasks=snapshot.subtasks)
        now = _utc_now_iso()
        healed = replace(
            record,
            status=TaskStatus.STOPPED,
            updated_at=now,
            metadata=runtime_metadata_for_status(record, TaskStatus.STOPPED, now),
        )
        cache.save_task(healed)
        return TaskSnapshot(record=healed, subtasks=snapshot.subtasks)

    def _glossary_review_artifacts_exist(self, record: TaskRecord) -> bool:
        if record.kind is not TaskKind.GLOSSARY_REVIEW:
            return False
        report_path = self.cache.task_dir(record.id) / REPORT_FILENAME
        output_path: Path | None = None
        result = self._read_result(record.id)
        if result is not None:
            raw_output = result.get("output_path")
            if isinstance(raw_output, str) and raw_output:
                output_path = Path(raw_output)
        if output_path is None:
            try:
                output_filename = normalize_output_filename(
                    str(record.metadata.get("output_filename", ""))
                )
            except ValueError:
                return False
            output_dir = Path(str(record.metadata.get("output_dir", "")))
            output_path = output_dir / output_filename
        return report_path.exists() and output_path.exists()

    def _on_task_failure(self, task_id: str, exc: BaseException) -> None:
        cache = self._cache_for_task(task_id)
        try:
            record = cache.load_record(task_id)
        except TaskNotFoundError:
            return
        message = f"{type(exc).__name__}: {exc}"
        metadata = dict(record.metadata)
        metadata["last_error"] = message
        metadata["last_traceback"] = "".join(
            traceback.format_exception(type(exc), exc, exc.__traceback__)
        )[-2000:]
        now = _utc_now_iso()
        timed_metadata = runtime_metadata_for_status(record, TaskStatus.FAILED, now)
        timed_metadata["last_error"] = metadata["last_error"]
        timed_metadata["last_traceback"] = metadata["last_traceback"]
        cache.save_task(
            replace(
                record,
                status=TaskStatus.FAILED,
                metadata=timed_metadata,
                updated_at=now,
            )
        )

    def _result_path(self, task_id: str) -> Path:
        return self._cache_for_task(task_id).task_dir(task_id) / _RESULT_FILENAME

    def _replacement_report_path(self, task_id: str) -> Path:
        return (
            self._cache_for_task(task_id).task_dir(task_id)
            / _REPLACEMENT_REPORT_FILENAME
        )

    def _epub_compress_report_path(self, task_id: str) -> Path:
        return (
            self._cache_for_task(task_id).task_dir(task_id)
            / _EPUB_COMPRESS_REPORT_FILENAME
        )

    def _epub_merge_report_path(self, task_id: str) -> Path:
        return (
            self._cache_for_task(task_id).task_dir(task_id)
            / _EPUB_MERGE_REPORT_FILENAME
        )

    def _epub_convert_report_path(self, task_id: str) -> Path:
        return (
            self._cache_for_task(task_id).task_dir(task_id)
            / _EPUB_CONVERT_REPORT_FILENAME
        )

    def _txt_to_epub_report_path(self, task_id: str) -> Path:
        return (
            self._cache_for_task(task_id).task_dir(task_id)
            / _TXT_TO_EPUB_REPORT_FILENAME
        )

    def _write_replacement_report(
        self, task_id: str, payload: Mapping[str, object]
    ) -> Path | None:
        try:
            path = self._replacement_report_path(task_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp = path.with_suffix(path.suffix + ".tmp")
            tmp.write_text(
                json.dumps(dict(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            os.replace(tmp, path)
            return path
        except OSError:
            # The report is observability, not load-bearing — never let
            # a write failure abort the task.
            return None

    def read_replacement_report(self, *, task_id: str) -> dict[str, object]:
        """Return the per-rule replacement report for a finished
        replacement task. Falls back to the in-memory mirror after the
        disk cache has been wiped on a clean completion. Raises
        ``BridgeError.not_found`` only when neither disk nor mirror
        has the report (older runs, or a task that never completed)."""

        path = self._replacement_report_path(task_id)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise BridgeError(
                    "bridge.io_error",
                    f"cannot read replacement report: {exc}",
                    retryable=False,
                    details={"task_id": task_id, "path": str(path)},
                ) from exc
        mirrored = self._completed_replacement_reports.get(task_id)
        if mirrored is not None:
            return mirrored
        raise BridgeError.not_found(
            f"replacement report not found for {task_id!r}",
            details={"task_id": task_id},
        )

    def _write_epub_compress_report(
        self, task_id: str, payload: Mapping[str, object]
    ) -> Path | None:
        try:
            path = self._epub_compress_report_path(task_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp = path.with_suffix(path.suffix + ".tmp")
            tmp.write_text(
                json.dumps(dict(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            os.replace(tmp, path)
            return path
        except OSError:
            return None

    def read_epub_compress_report(self, *, task_id: str) -> dict[str, object]:
        path = self._epub_compress_report_path(task_id)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise BridgeError(
                    "bridge.io_error",
                    f"cannot read EPUB compress report: {exc}",
                    retryable=False,
                    details={"task_id": task_id, "path": str(path)},
                ) from exc
        raise BridgeError.not_found(
            f"EPUB compress report not found for {task_id!r}",
            details={"task_id": task_id},
        )

    def _write_epub_merge_report(
        self, task_id: str, payload: Mapping[str, object]
    ) -> Path | None:
        try:
            path = self._epub_merge_report_path(task_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp = path.with_suffix(path.suffix + ".tmp")
            tmp.write_text(
                json.dumps(dict(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            os.replace(tmp, path)
            return path
        except OSError:
            return None

    def read_epub_merge_report(self, *, task_id: str) -> dict[str, object]:
        path = self._epub_merge_report_path(task_id)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise BridgeError(
                    "bridge.io_error",
                    f"cannot read EPUB merge report: {exc}",
                    retryable=False,
                    details={"task_id": task_id, "path": str(path)},
                ) from exc
        raise BridgeError.not_found(
            f"EPUB merge report not found for {task_id!r}",
            details={"task_id": task_id},
        )

    def _write_epub_convert_report(
        self, task_id: str, payload: Mapping[str, object]
    ) -> Path | None:
        try:
            path = self._epub_convert_report_path(task_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp = path.with_suffix(path.suffix + ".tmp")
            tmp.write_text(
                json.dumps(dict(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            os.replace(tmp, path)
            return path
        except OSError:
            return None

    def read_epub_convert_report(self, *, task_id: str) -> dict[str, object]:
        path = self._epub_convert_report_path(task_id)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise BridgeError(
                    "bridge.io_error",
                    f"cannot read EPUB convert report: {exc}",
                    retryable=False,
                    details={"task_id": task_id, "path": str(path)},
                ) from exc
        raise BridgeError.not_found(
            f"EPUB convert report not found for {task_id!r}",
            details={"task_id": task_id},
        )

    def _write_txt_to_epub_report(
        self, task_id: str, payload: Mapping[str, object]
    ) -> Path | None:
        try:
            path = self._txt_to_epub_report_path(task_id)
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp = path.with_suffix(path.suffix + ".tmp")
            tmp.write_text(
                json.dumps(dict(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            os.replace(tmp, path)
            return path
        except OSError:
            return None

    def read_txt_to_epub_report(self, *, task_id: str) -> dict[str, object]:
        path = self._txt_to_epub_report_path(task_id)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise BridgeError(
                    "bridge.io_error",
                    f"cannot read TXT to EPUB report: {exc}",
                    retryable=False,
                    details={"task_id": task_id, "path": str(path)},
                ) from exc
        raise BridgeError.not_found(
            f"TXT to EPUB report not found for {task_id!r}",
            details={"task_id": task_id},
        )

    def read_glossary_review_report(self, *, task_id: str) -> dict[str, object]:
        record_kind = TaskKind.GLOSSARY_REVIEW
        try:
            snapshot = self.cache.load(task_id)
        except TaskNotFoundError as exc:
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        record = snapshot.record
        if record.kind is not record_kind:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        snapshot = self._reconcile_glossary_review_completion(snapshot, self.cache)
        if snapshot.record.status is not TaskStatus.COMPLETED:
            raise BridgeError.conflict(
                "glossary review report is not ready yet.",
                details={"task_id": task_id, "status": snapshot.record.status.value},
            )
        path = self.cache.task_dir(task_id) / REPORT_FILENAME
        if not path.exists():
            raise BridgeError.not_found(
                f"glossary review report not found for {task_id!r}",
                details={"task_id": task_id},
            )
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise BridgeError(
                "bridge.io_error",
                f"cannot read glossary review report: {exc}",
                retryable=False,
                details={"task_id": task_id, "path": str(path)},
            ) from exc
        return payload if isinstance(payload, dict) else {"rows": []}

    def read_glossary_review_final(self, *, task_id: str) -> dict[str, object]:
        path = self._glossary_review_output_path(task_id)
        loaded = load_glossary_xlsx(path)
        return {
            "task_id": task_id,
            "path": str(path),
            "rows": [
                {
                    "row_index": row.row_index,
                    "src": row.src,
                    "dst": row.dst,
                    "info": row.info,
                    "frequency": row.frequency,
                }
                for row in loaded.rows
            ],
        }

    def update_glossary_review_final_row(
        self,
        *,
        task_id: str,
        row_index: int,
        src: str,
        dst: str,
        info: str,
        delete: bool = False,
    ) -> dict[str, object]:
        if row_index < 2:
            raise BridgeError.invalid_argument(
                "row_index must point to a spreadsheet data row.",
                field="row_index",
            )
        path = self._glossary_review_output_path(task_id)
        loaded = load_glossary_xlsx(path)
        if row_index not in {row.row_index for row in loaded.rows}:
            raise BridgeError.not_found(
                f"glossary review row {row_index!r} not found.",
                details={"task_id": task_id, "row_index": row_index},
            )
        workbook = load_workbook(path)
        sheet = workbook[loaded.sheet_name]
        if delete:
            sheet.delete_rows(row_index, 1)
        else:
            sheet.cell(row=row_index, column=loaded.source_col, value=src.strip())
            sheet.cell(row=row_index, column=loaded.target_col, value=dst.strip())
            sheet.cell(row=row_index, column=loaded.info_col, value=info.strip())
        workbook.save(path)
        return self.read_glossary_review_final(task_id=task_id)

    def delete_glossary_review_final_rows(
        self, *, task_id: str, row_indices: list[int]
    ) -> dict[str, object]:
        if not row_indices:
            raise BridgeError.invalid_argument(
                "row_indices must not be empty.",
                field="row_indices",
            )
        if any(row_index < 2 for row_index in row_indices):
            raise BridgeError.invalid_argument(
                "row_indices must point to spreadsheet data rows.",
                field="row_indices",
            )
        path = self._glossary_review_output_path(task_id)
        loaded = load_glossary_xlsx(path)
        existing = {row.row_index for row in loaded.rows}
        missing = sorted(set(row_indices) - existing)
        if missing:
            raise BridgeError.not_found(
                f"glossary review rows not found: {missing!r}.",
                details={"task_id": task_id, "row_indices": missing},
            )
        workbook = load_workbook(path)
        sheet = workbook[loaded.sheet_name]
        for row_index in sorted(set(row_indices), reverse=True):
            sheet.delete_rows(row_index, 1)
        workbook.save(path)
        return self.read_glossary_review_final(task_id=task_id)

    def restore_glossary_review_deleted_report_row(
        self,
        *,
        task_id: str,
        src: str,
        dst: str,
        info: str,
        frequency: int = 0,
    ) -> dict[str, object]:
        source = src.strip()
        target = dst.strip()
        category = info.strip()
        if not source or not target:
            raise BridgeError.invalid_argument(
                "src and dst must not be empty.",
                field="src",
            )
        path = self._glossary_review_output_path(task_id)
        loaded = load_glossary_xlsx(path)
        workbook = load_workbook(path)
        sheet = workbook[loaded.sheet_name]
        row_index = sheet.max_row + 1
        sheet.cell(row=row_index, column=loaded.source_col, value=source)
        sheet.cell(row=row_index, column=loaded.target_col, value=target)
        sheet.cell(row=row_index, column=loaded.info_col, value=category)
        if loaded.frequency_col is not None:
            sheet.cell(
                row=row_index,
                column=loaded.frequency_col,
                value=max(0, int(frequency)),
            )
        workbook.save(path)
        return self.read_glossary_review_final(task_id=task_id)

    def _glossary_review_output_path(self, task_id: str) -> Path:
        try:
            record = self.cache.load_record(task_id)
        except TaskNotFoundError as exc:
            raise BridgeError.not_found(
                f"task {task_id!r} not found.",
                details={"task_id": task_id},
            ) from exc
        if record.kind is not TaskKind.GLOSSARY_REVIEW:
            raise BridgeError.invalid_argument(
                f"task {task_id!r} kind mismatch.",
                field="task_id",
            )
        result = self._read_result(task_id)
        output_path = result.get("output_path") if result else None
        if isinstance(output_path, str) and output_path:
            path = Path(output_path)
        else:
            metadata = record.metadata
            output_dir = Path(str(metadata.get("output_dir", "")))
            output_filename = normalize_output_filename(
                str(metadata.get("output_filename", ""))
            )
            path = output_dir / output_filename
        if not path.exists():
            raise BridgeError.not_found(
                f"glossary review output not found for {task_id!r}.",
                details={"task_id": task_id, "path": str(path)},
            )
        return path

    def _write_result(
        self, task_id: str, payload: Mapping[str, object]
    ) -> None:
        path = self._result_path(task_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(
            json.dumps(dict(payload), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        os.replace(tmp, path)

    def _read_result(self, task_id: str) -> dict[str, object] | None:
        path = self._result_path(task_id)
        if not path.exists():
            return None
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except OSError as exc:
            raise BridgeError(
                "bridge.io_error",
                f"cannot read artifacts for task {task_id!r}: {exc}",
                retryable=False,
                details={"task_id": task_id, "path": str(path)},
            ) from exc
        except ValueError as exc:
            raise BridgeError(
                "task.artifact_corrupt",
                f"artifact payload for task {task_id!r} is not valid JSON.",
                retryable=False,
                details={"task_id": task_id, "path": str(path)},
            ) from exc
        if isinstance(payload, dict):
            return payload
        raise BridgeError(
            "task.artifact_corrupt",
            f"artifact payload for task {task_id!r} must be a JSON object.",
            retryable=False,
            details={"task_id": task_id, "path": str(path)},
        )

    def _partial_result(
        self, *, record: TaskRecord, snapshot: TaskSnapshot
    ) -> dict[str, object]:
        metadata = record.metadata
        output_dir = Path(str(metadata.get("output_dir", "")))
        if record.kind is TaskKind.TRANSLATION:
            return _partial_translation_payload(
                snapshot,
                output_dir=output_dir,
                statistics_dir=self._cache_for_kind("translation").task_dir(record.id),
            )
        if record.kind is TaskKind.GLOSSARY:
            input_dir = Path(str(metadata.get("input_dir", "")))
            return _partial_glossary_payload(
                output_dir=output_dir,
                input_folder_name=input_dir.name,
                statistics_dir=self._cache_for_kind("glossary").task_dir(record.id),
            )
        if record.kind is TaskKind.GLOSSARY_REVIEW:
            return _partial_glossary_review_payload(
                output_dir=output_dir,
                output_filename=str(metadata.get("output_filename", "")),
            )
        if record.kind is TaskKind.REPLACEMENT:
            return _partial_replacement_payload(snapshot, output_dir=output_dir)
        if record.kind is TaskKind.EPUB_COMPRESS:
            return _partial_epub_compress_payload(
                snapshot,
                output_folder=output_dir,
                report_path=self._epub_compress_report_path(record.id),
            )
        if record.kind is TaskKind.EPUB_MERGE:
            return _partial_epub_merge_payload(
                snapshot,
                output_folder=output_dir,
                report_path=self._epub_merge_report_path(record.id),
            )
        if record.kind is TaskKind.EPUB_CONVERT:
            return _partial_epub_convert_payload(
                snapshot,
                output_folder=output_dir,
                report_path=self._epub_convert_report_path(record.id),
            )
        if record.kind is TaskKind.TXT_TO_EPUB:
            return _partial_txt_to_epub_payload(
                snapshot,
                output_folder=output_dir,
                report_path=self._txt_to_epub_report_path(record.id),
            )
        raise BridgeError.invalid_argument(
            f"unsupported task kind: {record.kind.value!r}",
            field="kind",
        )


def default_llm_client_factory() -> LlmClient:
    """Default factory used in production: a real httpx-backed transport
    with no proxy. Use ``make_llm_client_factory(settings_store)`` from
    the router build path to honor the user's ``app.proxy_url`` setting."""

    return LlmClient(transport=HttpxChatTransport())


def make_llm_client_factory(
    settings_store: SettingsStore,
) -> LlmClientFactory:
    """Build an ``LlmClientFactory`` that reads the current
    ``app.proxy_url`` setting at every call.

    Reading on each construction (not once at startup) means the user
    can change the proxy in App Settings and the very next LLM call
    picks it up — no restart required. Empty / missing proxy falls
    back to httpx defaults (``HTTPS_PROXY`` env var if set).
    """

    def factory() -> LlmClient:
        proxy = ""
        try:
            proxy = (settings_store.load_all().app.proxy_url or "").strip()
        except Exception:  # noqa: BLE001 — settings hiccup must not block LLM calls
            proxy = ""
        return LlmClient(transport=HttpxChatTransport(proxy=proxy or None))

    return factory


__all__ = [
    "REPLACED_SUFFIX",
    "TaskService",
    "LlmClientFactory",
    "default_llm_client_factory",
    "make_llm_client_factory",
]
