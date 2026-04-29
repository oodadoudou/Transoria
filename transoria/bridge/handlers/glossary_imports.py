"""``glossary.import_rules`` bridge handler.

Reads a glossary file from disk (JSON or XLSX) and returns a list of
``{src, dst, info, regex, case_sensitive, enabled}`` entries the frontend
can splice into its in-memory glossary store.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Mapping

from transoria.bridge.errors import BridgeError
from transoria.bridge.handlers._utils import expect_string
from transoria.bridge.router import BridgeRouter


def _entry_from_record(raw: object) -> dict[str, object] | None:
    if not isinstance(raw, Mapping):
        return None
    src = str(raw.get("src", "")).strip()
    dst = str(raw.get("dst", "")).strip()
    if not src and not dst:
        return None
    info_value = raw.get("info") or raw.get("type") or raw.get("description") or ""
    return {
        "src": src,
        "dst": dst,
        "info": str(info_value).strip(),
        "regex": bool(raw.get("regex", False)),
        "case_sensitive": bool(raw.get("case_sensitive", False)),
        "enabled": raw.get("enabled", True) is not False,
    }


def _parse_json(path: Path) -> list[dict[str, object]]:
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8-sig")
    payload = json.loads(text)
    if not isinstance(payload, list):
        raise BridgeError.invalid_argument(
            "glossary JSON must be an array of entries.",
            details={"path": str(path)},
        )
    out: list[dict[str, object]] = []
    for raw in payload:
        entry = _entry_from_record(raw)
        if entry is not None:
            out.append(entry)
    return out


def _parse_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise BridgeError(
            "bridge.io_error",
            "openpyxl is required to import XLSX glossaries.",
            retryable=False,
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    keys = [str(value).strip().lower() if value is not None else "" for value in header]
    out: list[dict[str, object]] = []
    for row in rows:
        record: dict[str, object] = {}
        for key, value in zip(keys, row):
            if not key:
                continue
            record[key] = value
        entry = _entry_from_record(record)
        if entry is not None:
            out.append(entry)
    return out


def import_rules(payload: Mapping[str, object]) -> dict[str, object]:
    path_str = expect_string(payload, "path")
    path = Path(path_str)
    if not path.exists():
        raise BridgeError.not_found(
            f"glossary file does not exist: {path_str!r}",
            details={"path": path_str},
        )
    suffix = path.suffix.lower()
    try:
        if suffix in (".json",):
            entries = _parse_json(path)
        elif suffix in (".xlsx",):
            entries = _parse_xlsx(path)
        else:
            raise BridgeError.invalid_argument(
                f"unsupported glossary format: {suffix or 'unknown'}",
                details={"path": path_str, "suffix": suffix},
            )
    except json.JSONDecodeError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"glossary JSON is malformed: {exc}",
            retryable=False,
            details={"path": path_str},
        ) from exc
    except OSError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"cannot read glossary file: {exc}",
            retryable=True,
            details={"path": path_str},
        ) from exc
    return {"entries": entries}


def _coerce_export_entries(raw: object) -> list[dict[str, object]]:
    if not isinstance(raw, list):
        raise BridgeError.invalid_argument(
            "entries must be a list of objects.",
            field="entries",
        )
    out: list[dict[str, object]] = []
    for item in raw:
        if not isinstance(item, Mapping):
            continue
        out.append(
            {
                "src": str(item.get("src", "")),
                "dst": str(item.get("dst", "")),
                "info": str(item.get("info", "")),
                "regex": bool(item.get("regex", False)),
                "case_sensitive": bool(item.get("case_sensitive", False)),
                "enabled": item.get("enabled", True) is not False,
            }
        )
    return out


def _write_json(path: Path, entries: list[dict[str, object]]) -> None:
    path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_xlsx(path: Path, entries: list[dict[str, object]]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise BridgeError(
            "bridge.io_error",
            "openpyxl is required to export XLSX glossaries.",
            retryable=False,
        ) from exc
    workbook = Workbook()
    sheet = workbook.active
    headers = ("src", "dst", "info", "regex", "case_sensitive", "enabled")
    sheet.append(headers)
    for entry in entries:
        sheet.append(tuple(entry.get(key, "") for key in headers))
    workbook.save(path)


def export_rules(payload: Mapping[str, object]) -> dict[str, object]:
    path_str = expect_string(payload, "path")
    path = Path(path_str)
    suffix = path.suffix.lower()
    if suffix == "":
        path = path.with_suffix(".json")
        suffix = ".json"
    if suffix not in (".json", ".xlsx"):
        raise BridgeError.invalid_argument(
            f"unsupported export format: {suffix}",
            details={"path": path_str, "suffix": suffix},
        )
    entries = _coerce_export_entries(payload.get("entries"))
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        if suffix == ".json":
            _write_json(path, entries)
        else:
            _write_xlsx(path, entries)
    except OSError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"cannot write glossary file: {exc}",
            retryable=True,
            details={"path": str(path)},
        ) from exc
    return {"path": str(path), "count": len(entries)}


def _presets_dir(cache_root: Path) -> Path:
    return cache_root / "glossary_presets"


def list_presets_factory(cache_root: Path):
    def list_presets(_payload: Mapping[str, object]) -> dict[str, object]:
        directory = _presets_dir(cache_root)
        out: list[dict[str, object]] = []
        if directory.exists() and directory.is_dir():
            for entry_path in sorted(directory.iterdir()):
                if not entry_path.is_file():
                    continue
                if entry_path.suffix.lower() != ".json":
                    continue
                try:
                    entries = _parse_json(entry_path)
                except (BridgeError, OSError, json.JSONDecodeError):
                    continue
                out.append(
                    {
                        "id": entry_path.stem,
                        "name": entry_path.stem.replace("-", " ").replace(
                            "_", " "
                        ),
                        "entry_count": len(entries),
                        "entries": entries,
                    }
                )
        return {"directory": str(directory), "presets": out}

    return list_presets


def register(router: BridgeRouter, *, cache_root: Path) -> None:
    router.register("glossary.import_rules", import_rules)
    router.register("glossary.export_rules", export_rules)
    router.register("glossary.list_presets", list_presets_factory(cache_root))


__all__ = ["register", "import_rules", "export_rules", "list_presets_factory"]
