"""``rules.import_rules`` / ``rules.export_rules`` bridge handlers.

A single kind-aware entry point covers the three translation-side rule
families that share the JSON + XLSX wire convention:

* ``text_preserve``      — patterns that pass through untranslated
* ``pre_replacement``    — search/replace applied to source before LLM
* ``post_replacement``   — search/replace applied to model output

Replacement rules use the ``{src, dst, regex, case_sensitive, info}``
shape that matches third-party translator tooling, so users can import
files exported elsewhere without reformatting. Text-preserve rules use
the native ``{pattern, note, enabled}`` shape but accept ``src`` as a
fallback for ``pattern`` so the same exported file can be re-imported
into the preserve list when that's what the user intended.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Mapping

from transoria.bridge.errors import BridgeError
from transoria.bridge.handlers._utils import expect_string
from transoria.bridge.router import BridgeRouter
from transoria.formats.text import decode_text_bytes


_TEXT_PRESERVE = "text_preserve"
_PRE_REPLACEMENT = "pre_replacement"
_POST_REPLACEMENT = "post_replacement"
_KINDS = frozenset({_TEXT_PRESERVE, _PRE_REPLACEMENT, _POST_REPLACEMENT})


def _expect_kind(payload: Mapping[str, object]) -> str:
    kind = expect_string(payload, "kind")
    if kind not in _KINDS:
        raise BridgeError.invalid_argument(
            f"unsupported rule kind: {kind!r}",
            field="kind",
            details={"kind": kind, "allowed": sorted(_KINDS)},
        )
    return kind


def _normalize_text_preserve(raw: object) -> dict[str, object] | None:
    if not isinstance(raw, Mapping):
        return None
    # ``pattern`` is canonical; fall back to ``src`` so a file exported
    # from the replacement schema (or a third-party tool) still seeds
    # the preserve list when the user explicitly imports it here.
    pattern_raw = raw.get("pattern")
    if pattern_raw is None or pattern_raw == "":
        pattern_raw = raw.get("src", "")
    pattern = str(pattern_raw).strip()
    if not pattern:
        return None
    note_raw = raw.get("note")
    if note_raw is None or note_raw == "":
        note_raw = raw.get("info", "")
    return {
        "pattern": pattern,
        "note": str(note_raw).strip(),
        "enabled": raw.get("enabled", True) is not False,
    }


def _normalize_replacement(raw: object) -> dict[str, object] | None:
    if not isinstance(raw, Mapping):
        return None
    src = str(raw.get("src", "")).strip()
    if not src:
        return None
    dst = str(raw.get("dst", ""))
    note_raw = raw.get("note")
    if note_raw is None or note_raw == "":
        note_raw = raw.get("info", "")
    return {
        "src": src,
        "dst": dst,
        "regex": bool(raw.get("regex", False)),
        "case_sensitive": bool(raw.get("case_sensitive", False)),
        "note": str(note_raw).strip(),
        "enabled": raw.get("enabled", True) is not False,
    }


def _normalizer_for(kind: str):
    return _normalize_text_preserve if kind == _TEXT_PRESERVE else _normalize_replacement


def _parse_json(path: Path, kind: str) -> list[dict[str, object]]:
    text, _ = decode_text_bytes(path.read_bytes())
    payload = json.loads(text)
    if not isinstance(payload, list):
        raise BridgeError.invalid_argument(
            "rules JSON must be an array of objects.",
            details={"path": str(path)},
        )
    normalize = _normalizer_for(kind)
    out: list[dict[str, object]] = []
    for raw in payload:
        rule = normalize(raw)
        if rule is not None:
            out.append(rule)
    return out


def _parse_xlsx(path: Path, kind: str) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise BridgeError(
            "bridge.io_error",
            "openpyxl is required to import XLSX rules.",
            retryable=False,
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    keys = [
        str(value).strip().lower() if value is not None else ""
        for value in header
    ]
    normalize = _normalizer_for(kind)
    out: list[dict[str, object]] = []
    for row in rows:
        record: dict[str, object] = {}
        for key, value in zip(keys, row):
            if not key:
                continue
            record[key] = value
        rule = normalize(record)
        if rule is not None:
            out.append(rule)
    return out


def import_rules(payload: Mapping[str, object]) -> dict[str, object]:
    kind = _expect_kind(payload)
    path_str = expect_string(payload, "path")
    path = Path(path_str)
    if not path.exists():
        raise BridgeError.not_found(
            f"rules file does not exist: {path_str!r}",
            details={"path": path_str},
        )
    suffix = path.suffix.lower()
    try:
        if suffix == ".json":
            rules = _parse_json(path, kind)
        elif suffix == ".xlsx":
            rules = _parse_xlsx(path, kind)
        else:
            raise BridgeError.invalid_argument(
                f"unsupported rules format: {suffix or 'unknown'}",
                details={"path": path_str, "suffix": suffix},
            )
    except json.JSONDecodeError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"rules JSON is malformed: {exc}",
            retryable=False,
            details={"path": path_str},
        ) from exc
    except OSError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"cannot read rules file: {exc}",
            retryable=True,
            details={"path": path_str},
        ) from exc
    return {"kind": kind, "rules": rules}


def _coerce_export_rules(
    raw: object, kind: str
) -> list[dict[str, object]]:
    if not isinstance(raw, list):
        raise BridgeError.invalid_argument(
            "rules must be a list of objects.",
            field="rules",
        )
    normalize = _normalizer_for(kind)
    out: list[dict[str, object]] = []
    for item in raw:
        rule = normalize(item)
        if rule is not None:
            out.append(rule)
    return out


def _headers_for(kind: str) -> tuple[str, ...]:
    if kind == _TEXT_PRESERVE:
        return ("pattern", "note", "enabled")
    return ("src", "dst", "info", "regex", "case_sensitive", "enabled")


def _row_for(kind: str, rule: Mapping[str, object]) -> tuple[object, ...]:
    if kind == _TEXT_PRESERVE:
        return (
            rule.get("pattern", ""),
            rule.get("note", ""),
            rule.get("enabled", True),
        )
    return (
        rule.get("src", ""),
        rule.get("dst", ""),
        rule.get("note", ""),
        rule.get("regex", False),
        rule.get("case_sensitive", False),
        rule.get("enabled", True),
    )


def _payload_for(kind: str, rule: Mapping[str, object]) -> dict[str, object]:
    """Map normalized in-memory rule shape to the on-disk JSON shape.

    Replacement rules persist `note` as `info` so a file exported by
    Transoria round-trips with third-party translator tooling that
    keys on `info` rather than `note`.
    """

    if kind == _TEXT_PRESERVE:
        return {
            "pattern": rule.get("pattern", ""),
            "note": rule.get("note", ""),
            "enabled": rule.get("enabled", True),
        }
    return {
        "src": rule.get("src", ""),
        "dst": rule.get("dst", ""),
        "info": rule.get("note", ""),
        "regex": rule.get("regex", False),
        "case_sensitive": rule.get("case_sensitive", False),
        "enabled": rule.get("enabled", True),
    }


def _write_json(
    path: Path, kind: str, rules: list[dict[str, object]]
) -> None:
    payload = [_payload_for(kind, rule) for rule in rules]
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_xlsx(
    path: Path, kind: str, rules: list[dict[str, object]]
) -> None:
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise BridgeError(
            "bridge.io_error",
            "openpyxl is required to export XLSX rules.",
            retryable=False,
        ) from exc
    workbook = Workbook()
    sheet = workbook.active
    headers = _headers_for(kind)
    sheet.append(headers)
    for rule in rules:
        sheet.append(_row_for(kind, rule))
    workbook.save(path)


def export_rules(payload: Mapping[str, object]) -> dict[str, object]:
    kind = _expect_kind(payload)
    path_str = expect_string(payload, "path")
    path = Path(path_str)
    suffix = path.suffix.lower()
    if suffix == "":
        path = path.with_suffix(".json")
        suffix = ".json"
    if suffix not in (".json", ".xlsx"):
        raise BridgeError.invalid_argument(
            f"unsupported export format: {suffix}",
            details={"path": path_str, "suffix": suffix},
        )
    rules = _coerce_export_rules(payload.get("rules"), kind)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        if suffix == ".json":
            _write_json(path, kind, rules)
        else:
            _write_xlsx(path, kind, rules)
    except OSError as exc:
        raise BridgeError(
            "bridge.io_error",
            f"cannot write rules file: {exc}",
            retryable=True,
            details={"path": str(path)},
        ) from exc
    return {"kind": kind, "path": str(path), "count": len(rules)}


def register(router: BridgeRouter) -> None:
    router.register("rules.import_rules", import_rules)
    router.register("rules.export_rules", export_rules)


__all__ = ["register", "import_rules", "export_rules"]
