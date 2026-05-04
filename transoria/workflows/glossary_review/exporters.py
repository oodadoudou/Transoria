"""Final spreadsheet and cache report writers for glossary review."""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import load_workbook

from transoria.workflows.glossary_review.loader import (
    GlossaryReviewRow,
    LoadedGlossary,
    normalize_output_filename,
)

REPORT_FILENAME = "glossary-review-report.json"


def write_reviewed_xlsx(
    loaded: LoadedGlossary,
    rows: Sequence[GlossaryReviewRow],
    *,
    output_dir: Path,
    output_filename: str,
) -> Path:
    output_name = normalize_output_filename(output_filename)
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(loaded.workbook_path)
    sheet = workbook[loaded.sheet_name]
    if sheet.cell(row=1, column=loaded.info_col).value in (None, ""):
        sheet.cell(row=1, column=loaded.info_col, value="info")

    by_index = {row.row_index: row for row in rows}
    deleted: list[int] = []
    for row_index, row in by_index.items():
        if row.deleted:
            deleted.append(row_index)
            continue
        sheet.cell(row=row_index, column=loaded.target_col, value=row.dst)
        sheet.cell(row=row_index, column=loaded.info_col, value=row.info)
    for row_index in sorted(deleted, reverse=True):
        sheet.delete_rows(row_index, 1)

    path = output_dir / output_name
    tmp = path.with_suffix(path.suffix + ".tmp")
    workbook.save(tmp)
    os.replace(tmp, path)
    return path


def write_report(
    task_dir: Path,
    payload: Mapping[str, object],
) -> Path:
    task_dir.mkdir(parents=True, exist_ok=True)
    path = task_dir / REPORT_FILENAME
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(dict(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(tmp, path)
    return path


def read_report(task_dir: Path) -> dict[str, object] | None:
    path = task_dir / REPORT_FILENAME
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else None


__all__ = [
    "REPORT_FILENAME",
    "read_report",
    "write_report",
    "write_reviewed_xlsx",
]
