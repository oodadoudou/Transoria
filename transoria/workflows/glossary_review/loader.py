"""Input discovery and spreadsheet loading for glossary review."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook


SOURCE_HEADERS: frozenset[str] = frozenset(
    {"src", "source", "original", "term", "原文", "术语", "韩文"}
)
TARGET_HEADERS: frozenset[str] = frozenset(
    {"dst", "target", "translation", "translated", "译文", "中文"}
)
INFO_HEADERS: frozenset[str] = frozenset(
    {"info", "type", "category", "分类", "类型"}
)
FREQUENCY_HEADERS: frozenset[str] = frozenset(
    {"frequency", "freq", "count", "次数", "出现次数"}
)


class GlossaryReviewInputError(ValueError):
    pass


@dataclass(frozen=True)
class GlossaryReviewRow:
    row_index: int
    src: str
    dst: str
    info: str
    frequency: int
    context: str = ""
    deleted: bool = False


@dataclass(frozen=True)
class LoadedGlossary:
    workbook_path: Path
    sheet_name: str
    rows: tuple[GlossaryReviewRow, ...]
    source_col: int
    target_col: int
    info_col: int
    frequency_col: int | None


@dataclass(frozen=True)
class ReviewInput:
    glossary: LoadedGlossary
    reference_text: str
    reference_files: tuple[Path, ...]


def normalize_output_filename(value: str) -> str:
    name = (value or "glossary-review-final.xlsx").strip()
    if not name:
        name = "glossary-review-final.xlsx"
    if Path(name).name != name:
        raise GlossaryReviewInputError("output filename must not contain a path")
    if Path(name).suffix and Path(name).suffix.lower() != ".xlsx":
        raise GlossaryReviewInputError("output filename must end with .xlsx")
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name


def load_review_input(input_dir: Path, *, output_filename: str) -> ReviewInput:
    output_name = normalize_output_filename(output_filename)
    xlsx_path = _discover_glossary_xlsx(input_dir, output_filename=output_name)
    reference_files = _discover_reference_texts(input_dir)
    if not reference_files:
        raise GlossaryReviewInputError(
            "input folder must contain at least one .txt reference file"
        )
    return ReviewInput(
        glossary=load_glossary_xlsx(xlsx_path),
        reference_text=_merge_reference_texts(reference_files),
        reference_files=tuple(reference_files),
    )


def load_glossary_xlsx(path: Path) -> LoadedGlossary:
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = _headers(sheet)
    source_col = _find_column(headers, SOURCE_HEADERS, fallback=1)
    target_col = _find_column(headers, TARGET_HEADERS, fallback=2)
    if source_col == target_col:
        raise GlossaryReviewInputError("source and target columns must differ")
    info_col = _find_column(headers, INFO_HEADERS, fallback=0)
    if info_col <= 0:
        info_col = sheet.max_column + 1
        sheet.cell(row=1, column=info_col, value="info")
    frequency_col = _find_column(headers, FREQUENCY_HEADERS, fallback=0) or None

    rows: list[GlossaryReviewRow] = []
    for row_index in range(2, sheet.max_row + 1):
        src = _cell_text(sheet.cell(row=row_index, column=source_col).value)
        dst = _cell_text(sheet.cell(row=row_index, column=target_col).value)
        info = _cell_text(sheet.cell(row=row_index, column=info_col).value)
        if not src and not dst and not info:
            continue
        frequency = 1
        if frequency_col is not None:
            frequency = _cell_int(sheet.cell(row=row_index, column=frequency_col).value)
        rows.append(
            GlossaryReviewRow(
                row_index=row_index,
                src=src,
                dst=dst,
                info=info,
                frequency=frequency,
            )
        )
    if not rows:
        raise GlossaryReviewInputError("glossary spreadsheet contains no rows")
    return LoadedGlossary(
        workbook_path=path,
        sheet_name=sheet.title,
        rows=tuple(rows),
        source_col=source_col,
        target_col=target_col,
        info_col=info_col,
        frequency_col=frequency_col,
    )


def _discover_glossary_xlsx(input_dir: Path, *, output_filename: str) -> Path:
    candidates = [
        path
        for path in sorted(input_dir.iterdir(), key=lambda p: p.name.casefold())
        if path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not _is_ignored_xlsx(path, output_filename=output_filename)
    ]
    if not candidates:
        raise GlossaryReviewInputError("input folder contains no glossary .xlsx file")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise GlossaryReviewInputError(
            f"input folder contains multiple glossary .xlsx candidates: {names}"
        )
    return candidates[0]


def _discover_reference_texts(input_dir: Path) -> tuple[Path, ...]:
    return tuple(
        path
        for path in sorted(input_dir.iterdir(), key=lambda p: p.name.casefold())
        if path.is_file() and path.suffix.lower() == ".txt" and not path.name.startswith("~$")
    )


def _is_ignored_xlsx(path: Path, *, output_filename: str) -> bool:
    name = path.name
    lowered = name.casefold()
    return (
        name.startswith("~$")
        or lowered == output_filename.casefold()
        or lowered.endswith(".tmp.xlsx")
    )


def _merge_reference_texts(paths: Sequence[Path]) -> str:
    parts = [_read_text(path) for path in paths]
    return "\n\n".join(part for part in parts if part)


def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "cp932", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        text = _cell_text(sheet.cell(row=1, column=column).value)
        token = text.strip().lower()
        if token and token not in headers:
            headers[token] = column
    return headers


def _find_column(headers: dict[str, int], names: frozenset[str], *, fallback: int) -> int:
    for name in names:
        column = headers.get(name.lower())
        if column:
            return column
    return fallback


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_int(value: object) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0


__all__ = [
    "GlossaryReviewInputError",
    "GlossaryReviewRow",
    "LoadedGlossary",
    "ReviewInput",
    "load_glossary_xlsx",
    "load_review_input",
    "normalize_output_filename",
]
